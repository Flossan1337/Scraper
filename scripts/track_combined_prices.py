#!/usr/bin/env python3

import os
import re
import time
import math
import statistics
from datetime import date
from typing import List, Optional, Tuple

import requests
from bs4 import BeautifulSoup

# Excel (NYTT)
from openpyxl import Workbook, load_workbook

# ──────────────────────────────────────────────────────────────────────────────
# Selenium setup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager

# ──────────────────────────────────────────────────────────────────────────────
# 1) URLs & selectors

URL_INET = (
    "https://www.inet.se/kategori/48/"
    "datorlada-chassi?filter=%7B%22manufacturerIds%22%3A%5B293%5D%7D"
)
SELECTOR_INET = 'span[data-test-is-discounted-price="false"]'

URL_AMAZON = (
    "https://www.amazon.com/s?keywords=Computer+Cases"
    "&i=computers&rh=n%3A172282%2Cp_123%3A396156"
    "&s=exact-aware-popularity-rank"
    "&ref=sr_nr_p_123_3"
)
SELECTOR_AMAZON = "span.a-offscreen"

WEBHALLEN_URLS = [
    "https://www.webhallen.com/se/manufacturer/7782-Fractal-Design/36-Chassi?page=1",
    "https://www.webhallen.com/se/manufacturer/7782-Fractal-Design/36-Chassi?page=2",
]
AWD_IT_URL = "https://www.awd-it.co.uk/components/cases.html?brand=686&product_list_limit=64"
MEDIAMARKT_URLS = [
    "https://www.mediamarkt.de/de/search.html?query=fractal%20design&brand=FRACTAL%20DESIGN&productType=PC%20Geh%C3%A4use&page=1",
    "https://www.mediamarkt.de/de/search.html?query=fractal%20design&brand=FRACTAL%20DESIGN&productType=PC%20Geh%C3%A4use&page=2",
    "https://www.mediamarkt.de/de/search.html?query=fractal%20design&brand=FRACTAL%20DESIGN&productType=PC%20Geh%C3%A4use&page=3",
    "https://www.mediamarkt.de/de/search.html?query=fractal%20design&brand=FRACTAL%20DESIGN&productType=PC%20Geh%C3%A4use&page=4",
]
NEWEGG_URL = "https://www.newegg.com/Fractal-Design/BrandStore/ID-14581"

# Spara XLSX under ../data relativt till detta script (ditt repo har /data bredvid /scripts)
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.normpath(os.path.join(SCRIPT_DIR, "..", "data", "combined_prices.xlsx"))

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    )
}

# ──────────────────────────────────────────────────────────────────────────────
# 2) Helpers

def create_driver():
    opts = Options()
    opts.add_argument("--headless=new")
    opts.add_argument("--window-size=1920,1080")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--lang=sv-SE,sv;q=0.9,en-US;q=0.8,en;q=0.7,de;q=0.6")
    opts.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/115.0.0.0 Safari/537.36"
    )
    svc = Service(ChromeDriverManager().install())
    return webdriver.Chrome(service=svc, options=opts)

def fetch_prices_selenium(driver, url, css_selector, sleep_time=4, scroll_times=1) -> List[float]:
    driver.get(url)
    time.sleep(sleep_time)
    for _ in range(scroll_times):
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(2)
    soup = BeautifulSoup(driver.page_source, "html.parser")
    prices: List[float] = []
    for tag in soup.select(css_selector):
        txt = tag.get_text(strip=True)
        num = re.sub(r"[^\d,\.]", "", txt)
        num = num.replace(",", "")
        num = num.replace(",", ".")
        if not num or num.count(".") > 1:
            parts = re.findall(r"\d+\.\d+$", num)
            if parts:
                num = parts[-1]
            else:
                continue
        try:
            prices.append(float(num))
        except ValueError:
            continue
    return prices

def http_get(url: str, timeout: int = 25) -> requests.Response:
    r = requests.get(url, headers=HEADERS, timeout=timeout)
    r.raise_for_status()
    return r

def to_number(raw: str) -> Optional[float]:
    if not raw:
        return None
    s = raw.strip().replace("\xa0", " ").replace("\u202f", " ")
    keep = [ch for ch in s if ch.isdigit() or ch in ",."]
    s2 = "".join(keep)
    if not s2:
        return None
    if s2.count(",") and s2.count("."):
        if s2.rfind(",") > s2.rfind("."):
            s2 = s2.replace(".", "").replace(",", ".")
        else:
            s2 = s2.replace(",", "")
    else:
        if s2.count(",") == 1 and len(s2.split(",")[-1]) in (2, 3):
            s2 = s2.replace(",", ".")
        else:
            s2 = s2.replace(",", "")
    try:
        return float(s2)
    except ValueError:
        return None

def avg_median(nums: List[float]) -> Tuple[Optional[float], Optional[float]]:
    vals = [x for x in nums if isinstance(x, (int, float)) and not math.isnan(x)]
    if not vals:
        return None, None
    return round(sum(vals) / len(vals), 2), round(statistics.median(vals), 2)

def scroll_to_bottom(driver, times=2, pause=1.5):
    for _ in range(times):
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(pause)

# ──────────────────────────────────────────────────────────────────────────────
# 3) Scrapers (oförändrat)

def scrape_webhallen_selenium(driver) -> List[float]:
    prices: List[float] = []
    for url in WEBHALLEN_URLS:
        driver.get(url)
        time.sleep(3)
        scroll_to_bottom(driver, times=2, pause=1.2)
        soup = BeautifulSoup(driver.page_source, "html.parser")
        for span in soup.select("div.price-value._right span"):
            p = to_number(span.get_text())
            if p:
                prices.append(p)
    return prices

def scrape_mediamarkt_selenium(driver) -> List[float]:
    prices: List[float] = []
    for url in MEDIAMARKT_URLS:
        driver.get(url)
        time.sleep(3)
        scroll_to_bottom(driver, times=2, pause=1.0)
        soup = BeautifulSoup(driver.page_source, "html.parser")
        for span in soup.select('div[data-test*="cofr-price"] span'):
            p = to_number(span.get_text(strip=True))
            if p:
                prices.append(p)
    return prices

def scrape_newegg_selenium(driver) -> List[float]:
    driver.get(NEWEGG_URL)
    time.sleep(3)
    scroll_to_bottom(driver, times=3, pause=1.0)
    soup = BeautifulSoup(driver.page_source, "html.parser")
    prices: List[float] = []
    for node in soup.select("div.goods-price-current span.goods-price-value"):
        p = to_number(node.get_text(strip=True))
        if p:
            prices.append(p)
    if not prices:
        for card in soup.select("li.item-cell .price-current"):
            p = to_number(card.get_text(strip=True))
            if p:
                prices.append(p)
    return prices

def scrape_awd_it_requests() -> List[float]:
    prices: List[float] = []
    try:
        soup = BeautifulSoup(http_get(AWD_IT_URL).text, "html.parser")
        for node in soup.select("span.price-wrapper.price-including-tax > span.price"):
            p = to_number(node.get_text())
            if p:
                prices.append(p)
        if not prices:
            for node in soup.select("span.price-wrapper > span.price"):
                p = to_number(node.get_text())
                if p:
                    prices.append(p)
    except Exception as e:
        print(f"[AWD-IT] {e}")
    return prices

def scrape_inet_selenium(driver) -> List[float]:
    return fetch_prices_selenium(driver, URL_INET, SELECTOR_INET, sleep_time=4, scroll_times=1)

def scrape_amazon_selenium(driver) -> List[float]:
    return fetch_prices_selenium(driver, URL_AMAZON, SELECTOR_AMAZON, sleep_time=5, scroll_times=2)

# ──────────────────────────────────────────────────────────────────────────────
# 4) Excel-hantering (ersätter CSV)

def ensure_header_xlsx():
    os.makedirs(os.path.dirname(XLSX_PATH), exist_ok=True)
    if os.path.exists(XLSX_PATH):
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append([
        "date",
        "average_price_inet",
        "average_price_amazon",
        "average_price_webhallen",
        "average_price_awd_it",
        "average_price_mediamarkt",
        "average_price_newegg",
        "median_price_inet",
        "median_price_amazon",
        "median_price_webhallen",
        "median_price_awd_it",
        "median_price_mediamarkt",
        "median_price_newegg",
    ])
    wb.save(XLSX_PATH)

def append_row_xlsx(*, avg_inet, avg_amz, avg_webhallen, avg_awd, avg_mm, avg_newegg,
                    med_inet, med_amz, med_webhallen, med_awd, med_mm, med_newegg):
    today = date.today().isoformat()
    wb = load_workbook(XLSX_PATH)
    ws = wb.active
    ws.append([
        today,
        avg_inet, avg_amz, avg_webhallen,
        avg_awd, avg_mm, avg_newegg,
        med_inet, med_amz, med_webhallen,
        med_awd, med_mm, med_newegg,
    ])
    wb.save(XLSX_PATH)
    print(f"✓ Appended {today} → {XLSX_PATH}")

# ──────────────────────────────────────────────────────────────────────────────
# 5) Orchestrator

def main():
    ensure_header_xlsx()
    print(f"Working directory: {os.getcwd()}")
    print(f"Excel will be saved to: {XLSX_PATH}")

    driver = create_driver()
    try:
        inet_prices = scrape_inet_selenium(driver)
        print(f"  • Inet: {len(inet_prices)} prices")
        amz_prices = scrape_amazon_selenium(driver)
        print(f"  • Amazon: {len(amz_prices)} prices")
        webhallen_prices = scrape_webhallen_selenium(driver)
        print(f"  • Webhallen: {len(webhallen_prices)} prices")
        mm_prices = scrape_mediamarkt_selenium(driver)
        print(f"  • MediaMarkt: {len(mm_prices)} prices")
        newegg_prices = scrape_newegg_selenium(driver)
        print(f"  • Newegg: {len(newegg_prices)} prices")
    finally:
        driver.quit()

    awd_prices = scrape_awd_it_requests()
    print(f"  • AWD-IT: {len(awd_prices)} prices")

    # Stats
    avg_inet, med_inet = avg_median(inet_prices)
    avg_amz,  med_amz  = avg_median(amz_prices)
    avg_web,  med_web  = avg_median(webhallen_prices)
    avg_awd,  med_awd  = avg_median(awd_prices)
    avg_mm,   med_mm   = avg_median(mm_prices)
    avg_new,  med_new  = avg_median(newegg_prices)

    append_row_xlsx(
        avg_inet=avg_inet, avg_amz=avg_amz, avg_webhallen=avg_web,
        avg_awd=avg_awd, avg_mm=avg_mm, avg_newegg=avg_new,
        med_inet=med_inet, med_amz=med_amz, med_webhallen=med_web,
        med_awd=med_awd, med_mm=med_mm, med_newegg=med_new,
    )

if __name__ == "__main__":
    main()
