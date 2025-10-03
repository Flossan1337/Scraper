#!/usr/bin/env python3

import os
import csv
import re
import time
import statistics
from datetime import date
from bs4 import BeautifulSoup

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager

# NEW: Excel
from openpyxl import Workbook, load_workbook
from pathlib import Path

# ──────────────────────────────────────────────────────────────────────────────
URLS = [
    "https://www.rugvista.se/c/mattor/bastsaljare?page=1",
    "https://www.rugvista.se/c/mattor/bastsaljare?page=2",
    "https://www.rugvista.se/c/mattor/bastsaljare?page=3",
]

PRICE_SELECTOR = "div.text-sm.font-semibold"

# OLD: CSV_FILENAME = "rugvista_bestsellers.csv"
# NEW: write to ../data/rugvista_bestsellers.xlsx
SCRIPT_DIR = Path(__file__).resolve().parent
XLSX_PATH = (SCRIPT_DIR / ".." / "data" / "rugvista_bestsellers.xlsx").resolve()
# ──────────────────────────────────────────────────────────────────────────────

def create_driver():
    opts = Options()
    opts.add_argument("--headless")
    # force a desktop‐sized viewport so all products render immediately
    opts.add_argument("--window-size=1920,1080")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--no-sandbox")
    opts.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/115.0.0.0 Safari/537.36"
    )
    service = Service(ChromeDriverManager().install())
    return webdriver.Chrome(service=service, options=opts)

def fetch_prices(driver, url):
    driver.get(url)
    # give the JS a moment to load everything
    time.sleep(3)
    # scroll to bottom so any lazy items load
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(2)

    soup = BeautifulSoup(driver.page_source, "html.parser")
    prices = []
    for tag in soup.select(PRICE_SELECTOR):
        raw = tag.get_text(strip=True)             # e.g. "1 299 kr"
        num = re.sub(r"[^\d,\.]", "", raw)         # strip non-digits
        num = num.replace(",", ".")                # unify decimal
        try:
            prices.append(float(num))
        except ValueError:
            continue
    return prices

# ──────────────────────────────────────────────────────────────────────────────
# Excel handling (replaces CSV)

def ensure_header_xlsx():
    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    if XLSX_PATH.exists():
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["date", "median_price", "average_price"])
    wb.save(XLSX_PATH)

def append_to_xlsx(median_price, average_price):
    today = date.today().isoformat()
    if not XLSX_PATH.exists():
        ensure_header_xlsx()
    wb = load_workbook(XLSX_PATH)
    ws = wb.active
    ws.append([today, round(median_price, 2), round(average_price, 2)])
    wb.save(XLSX_PATH)
    print(f"✓ Appended {today}: median={median_price:.2f}, avg={average_price:.2f} → {XLSX_PATH}")

# ──────────────────────────────────────────────────────────────────────────────

def main():
    ensure_header_xlsx()
    driver = create_driver()
    all_prices = []

    for url in URLS:
        ps = fetch_prices(driver, url)
        print(f"  • Found {len(ps)} prices on {url}")
        all_prices.extend(ps)

    driver.quit()

    if not all_prices:
        print("❌ No prices found; double-check PRICE_SELECTOR and page content.")
        return

    med = statistics.median(all_prices)
    avg = statistics.mean(all_prices)
    append_to_xlsx(med, avg)

if __name__ == "__main__":
    main()
