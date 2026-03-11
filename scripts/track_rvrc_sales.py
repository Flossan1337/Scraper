#!/usr/bin/env python3
"""
track_rvrc_sales.py

Tracks RevolutionRace weekly and daily sales metrics at product-colour level
using the sale_last_week and sale_last_days counters from the Voyado Elevate API.

Methodology
-----------
  1. Queries the Elevate storefront API for all products across 5 markets
     (DE, SE, NO, UK, COM) and 3 categories (clothing, accessories, shoes).
  2. Groups variants by product-colour key (e.g. "10004_2001") using
     key.split("-", 1)[0] to strip the size suffix.
  3. sale_last_week and sale_last_days are global counters identical across
     all markets and sizes (verified by diagnostic on 2026-03-11).
  4. Prices come from the DE market (EUR) when available, falling back
     through SE, NO, UK, COM with FX conversion to EUR.

Output
------
  Excel: data/rvrc_sales.xlsx
    - "Latest Detail"  : replaced each run, one row per product-colour
    - "Daily Summary"   : appended each run, one row with aggregated totals

  State: data/rvrc_sales_state.json
    - daily_summaries array (one entry per run)
    - reference_products for day-over-day SLW tracking
"""

import json
import uuid
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Elevate API configuration
# ---------------------------------------------------------------------------
ELEVATE_CLUSTER_ID = "wA4BFC9F5"
ELEVATE_BASE_URL = f"https://{ELEVATE_CLUSTER_ID}.api.esales.apptus.cloud"
ELEVATE_ENDPOINT = "/api/storefront/v3/queries/landing-page"
ELEVATE_LIMIT = 600

ELEVATE_PRESENT_CUSTOM = (
    "allcategories|color_name|features_name|fit_name|gender_name|isdead|"
    "labeltext_name|googleanalyticsname|ratingcount|totalcolorcount|blobtext_name|"
    "sustainability|categorybreadcrumb|color|features|fit|gender|labeltext|blobtext|"
    "variant.campaign_id|variant.campaign_type|variant.campaign_has_conditions|"
    "variant.sale_last_days|variant.sale_last_week|variant.length_variant|variant.new_in"
)

ELEVATE_CATEGORIES = ["clothing", "accessories", "shoes"]

# ---------------------------------------------------------------------------
# Market configuration
# ---------------------------------------------------------------------------
MARKETS: dict[str, dict] = {
    "SE":  {"elevate_market": "SE", "locale": "sv-SE", "currency": "SEK"},
    "DE":  {"elevate_market": "DE", "locale": "de-DE", "currency": "EUR"},
    "NO":  {"elevate_market": "NO", "locale": "nb-NO", "currency": "NOK"},
    "UK":  {"elevate_market": "UK", "locale": "en-GB", "currency": "GBP"},
    "COM": {"elevate_market": "EU", "locale": "en-001", "currency": "EUR"},
}

MARKET_PRIORITY = ["DE", "SE", "NO", "UK", "COM"]

FX_FALLBACKS: dict[str, float] = {
    "SEK": 1.0,
    "EUR": 11.5,
    "NOK": 0.97,
    "GBP": 13.0,
}

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
SCRIPT_DIR = Path(__file__).resolve().parent
STATE_FILE = (SCRIPT_DIR / ".." / "data" / "rvrc_sales_state.json").resolve()
XLSX_PATH = (SCRIPT_DIR / ".." / "data" / "rvrc_sales.xlsx").resolve()

# ---------------------------------------------------------------------------
# Category keywords (breadcrumb -> display label)
# ---------------------------------------------------------------------------
_CATEGORY_KEYWORDS: list[tuple[str, str]] = [
    ("jacket", "Jackets & Vests"), ("vest", "Jackets & Vests"),
    ("trouser", "Trousers & Pants"), ("pant", "Trousers & Pants"),
    ("tight", "Trousers & Pants"),
    ("fleece", "Fleece & Midlayer"), ("midlayer", "Fleece & Midlayer"),
    ("sweater", "Fleece & Midlayer"),
    ("base layer", "Base Layer"), ("baselayer", "Base Layer"),
    ("t-shirt", "T-Shirts & Tops"), ("shirt", "T-Shirts & Tops"),
    ("top", "T-Shirts & Tops"),
    ("hoodie", "Hoodies & Sweatshirts"), ("sweatshirt", "Hoodies & Sweatshirts"),
    ("shoe", "Shoes"), ("boot", "Shoes"),
    ("sock", "Accessories"), ("cap", "Accessories"), ("hat", "Accessories"),
    ("glove", "Accessories"), ("bag", "Accessories"), ("backpack", "Accessories"),
    ("accessori", "Accessories"),
]


# ---------------------------------------------------------------------------
# State I/O
# ---------------------------------------------------------------------------

def load_state() -> dict:
    if STATE_FILE.exists():
        return json.loads(STATE_FILE.read_text(encoding="utf-8-sig"))
    return {"daily_summaries": [], "reference_products": []}


def save_state(state: dict) -> None:
    STATE_FILE.write_text(
        json.dumps(state, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


# ---------------------------------------------------------------------------
# FX rates
# ---------------------------------------------------------------------------

def fetch_fx_rates() -> dict[str, float]:
    non_sek = sorted({c["currency"] for c in MARKETS.values() if c["currency"] != "SEK"})
    rates: dict[str, float] = {"SEK": 1.0}
    if not non_sek:
        return rates
    try:
        resp = requests.get(
            f"https://api.frankfurter.app/latest?from=EUR&to=SEK,{','.join(non_sek)}",
            timeout=10,
        )
        resp.raise_for_status()
        data = resp.json()["rates"]
        eur_sek = float(data["SEK"])
        rates["EUR"] = eur_sek
        for ccy in non_sek:
            if ccy == "EUR":
                continue
            if ccy in data:
                rates[ccy] = eur_sek / float(data[ccy])
            else:
                rates[ccy] = FX_FALLBACKS.get(ccy, 1.0)
        print(f"  [FX] {rates}")
    except Exception as e:
        print(f"  [FX] API error ({e}) -- using fallbacks")
        for ccy in non_sek:
            rates[ccy] = FX_FALLBACKS.get(ccy, 1.0)
    return rates


# ---------------------------------------------------------------------------
# Elevate API helpers
# ---------------------------------------------------------------------------

def _get_price(price_raw) -> float:
    if isinstance(price_raw, dict):
        try:
            return float(price_raw.get("min") or 0.0)
        except (TypeError, ValueError):
            return 0.0
    try:
        return float(price_raw) if price_raw is not None else 0.0
    except (TypeError, ValueError):
        return 0.0


def _parse_category(breadcrumb_id: str, page_ref: str) -> str:
    if breadcrumb_id:
        parts = [p.strip() for p in breadcrumb_id.split(">")]
        raw = parts[1] if len(parts) >= 2 else parts[0]
        lower = raw.lower()
        for kw, label in _CATEGORY_KEYWORDS:
            if kw in lower:
                return label
        return raw
    return {"clothing": "Clothing", "accessories": "Accessories",
            "shoes": "Shoes"}.get(page_ref, page_ref.capitalize())


def fetch_elevate_page(
    elevate_market: str, locale: str, page_ref: str,
    skip: int = 0, customer_key: str = "", session_key: str = "",
) -> Optional[dict]:
    params = {
        "market": elevate_market, "locale": locale,
        "customerKey": customer_key, "sessionKey": session_key,
        "touchpoint": "desktop", "pageReference": page_ref,
        "limit": ELEVATE_LIMIT, "skip": skip,
        "presentCustom": ELEVATE_PRESENT_CUSTOM,
    }
    try:
        resp = requests.get(
            ELEVATE_BASE_URL + ELEVATE_ENDPOINT, params=params, timeout=(10, 60),
        )
        resp.raise_for_status()
        return resp.json()
    except requests.RequestException as exc:
        print(f"    [WARN] Elevate {page_ref} skip={skip}: {exc}")
        return None


def extract_variants(data: dict, page_ref: str = "") -> tuple[dict[str, dict], int, int]:
    """Extract variant data from an Elevate landing-page response."""
    variants: dict[str, dict] = {}
    pl = data.get("primaryList") or {}
    total_hits = int(pl.get("totalHits") or 0)
    groups = pl.get("productGroups") or []
    for group in groups:
        for product in group.get("products") or []:
            title = str(product.get("title") or product.get("name") or "")
            p_sell = _get_price(product.get("sellingPrice"))
            p_list = _get_price(product.get("listPrice")) or p_sell
            custom = product.get("custom") or {}
            breadcrumbs = custom.get("categorybreadcrumb") or []
            bc_id = breadcrumbs[0].get("id", "") if breadcrumbs else ""
            category = _parse_category(bc_id, page_ref)
            for variant in product.get("variants") or []:
                key = variant.get("key")
                if not key or not isinstance(key, str):
                    continue
                sell_price = _get_price(variant.get("sellingPrice")) or p_sell
                list_price = _get_price(variant.get("listPrice")) or p_list or sell_price
                v_custom = variant.get("custom") or {}

                slw = 0
                slw_list = v_custom.get("sale_last_week") or []
                try:
                    if slw_list:
                        slw = int(slw_list[0].get("label") or 0)
                except (TypeError, ValueError):
                    pass

                sld = 0
                sld_list = v_custom.get("sale_last_days") or []
                try:
                    if sld_list:
                        sld = int(sld_list[0].get("label") or 0)
                except (TypeError, ValueError):
                    pass

                variants[key] = {
                    "sale_last_week": slw,
                    "sale_last_days": sld,
                    "sell_price": sell_price,
                    "list_price": list_price,
                    "title": title,
                    "category": category,
                }
    return variants, total_hits, len(groups)


# ---------------------------------------------------------------------------
# Fetch all markets
# ---------------------------------------------------------------------------

def fetch_all_markets() -> dict[str, dict[str, dict]]:
    result: dict[str, dict[str, dict]] = {}
    for market_code, cfg in MARKETS.items():
        elevate_market = cfg["elevate_market"]
        locale = cfg["locale"]
        customer_key = str(uuid.uuid4())
        session_key = str(uuid.uuid4())
        print(f"  [{market_code}] Fetching (market={elevate_market}) ...")
        market_variants: dict[str, dict] = {}
        for cat in ELEVATE_CATEGORIES:
            skip, page = 0, 1
            while True:
                data = fetch_elevate_page(
                    elevate_market, locale, cat, skip, customer_key, session_key,
                )
                if data is None:
                    break
                new_v, total_hits, pg_count = extract_variants(data, cat)
                market_variants.update(new_v)
                print(f"    {cat} p{page}: +{len(new_v)} variants "
                      f"[{skip}-{skip + pg_count}/{total_hits}]")
                skip += pg_count
                if skip >= total_hits or pg_count == 0:
                    break
                page += 1
        print(f"  [{market_code}] {len(market_variants):,} unique variants")
        result[market_code] = market_variants
    return result


# ---------------------------------------------------------------------------
# Aggregate to product-colour level
# ---------------------------------------------------------------------------

def aggregate_product_colours(
    data_by_market: dict[str, dict[str, dict]],
    fx_rates: dict[str, float],
) -> list[dict]:
    """
    Deduplicate variants into product-colour rows with EUR prices.

    Each variant key like "10004_2001-M" maps to base_key "10004_2001"
    via key.split("-", 1)[0].  sale_last_week / sale_last_days are global
    (identical across sizes and markets), so we take the max per base_key
    as a safety measure.

    Price priority: DE (EUR natively) > SE > NO > UK > COM, converted to EUR.

    Returns one dict per product-colour with:
      base_key, title, category, sale_last_week, sale_last_days,
      sell_price_eur, list_price_eur
    """
    eur_sek = fx_rates.get("EUR", 11.5)
    seen: set[str] = set()
    rows: list[dict] = []

    for market_code in MARKET_PRIORITY:
        variants = data_by_market.get(market_code, {})
        currency = MARKETS[market_code]["currency"]
        fx_to_eur = fx_rates.get(currency, 1.0) / eur_sek

        # Group by base_key
        groups: dict[str, list[tuple[str, dict]]] = defaultdict(list)
        for key, v in variants.items():
            bk = key.split("-", 1)[0]
            groups[bk].append((key, v))

        for bk, items in groups.items():
            if bk in seen:
                continue
            seen.add(bk)

            slw = max((v.get("sale_last_week", 0) for _, v in items), default=0)
            sld = max((v.get("sale_last_days", 0) for _, v in items), default=0)

            # Pick representative variant (prefer one with positive sell_price)
            rep = next((v for _, v in items if v.get("sell_price", 0) > 0),
                       items[0][1])
            sell_eur = round(rep["sell_price"] * fx_to_eur, 2)
            list_raw = rep.get("list_price") or rep["sell_price"]
            list_eur = round(list_raw * fx_to_eur, 2)

            rows.append({
                "base_key":        bk,
                "title":           rep.get("title", ""),
                "category":        rep.get("category", ""),
                "sale_last_week":  slw,
                "sale_last_days":  sld,
                "sell_price_eur":  sell_eur,
                "list_price_eur":  list_eur,
            })

    return rows


# ---------------------------------------------------------------------------
# Compute daily summary
# ---------------------------------------------------------------------------

def compute_summary(rows: list[dict]) -> dict:
    """
    Compute the four aggregate revenue columns:
      - slw_x_sell:  sum(sale_last_week * sell_price_eur)
      - sld_x_sell:  sum(sale_last_days * sell_price_eur)
      - slw_x_list:  sum(sale_last_week * list_price_eur)
      - sld_x_list:  sum(sale_last_days * list_price_eur)
    """
    slw_x_sell = 0.0
    sld_x_sell = 0.0
    slw_x_list = 0.0
    sld_x_list = 0.0
    product_colors_total = len(rows)
    product_colors_with_sales = 0

    for r in rows:
        slw = r["sale_last_week"]
        sld = r["sale_last_days"]
        sp = r["sell_price_eur"]
        lp = r["list_price_eur"]

        slw_x_sell += slw * sp
        sld_x_sell += sld * sp
        slw_x_list += slw * lp
        sld_x_list += sld * lp

        if slw > 0 or sld > 0:
            product_colors_with_sales += 1

    return {
        "slw_x_sell_eur":  round(slw_x_sell, 2),
        "sld_x_sell_eur":  round(sld_x_sell, 2),
        "slw_x_list_eur":  round(slw_x_list, 2),
        "sld_x_list_eur":  round(sld_x_list, 2),
        "product_colors_total":      product_colors_total,
        "product_colors_with_sales": product_colors_with_sales,
    }


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

_HDR_FILL = PatternFill("solid", fgColor="1F497D")
_HDR_FONT = Font(bold=True, color="FFFFFF")


def _write_headers(ws, headers: list[str]) -> None:
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = Alignment(horizontal="center")


def _autofit(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 55)


def write_excel(today: str, rows: list[dict], summary: dict) -> None:
    if XLSX_PATH.exists():
        wb = load_workbook(XLSX_PATH)
    else:
        wb = Workbook()
        for name in list(wb.sheetnames):
            del wb[name]

    # -- Sheet 1: Daily Summary (append one row per run) ---------------------
    sum_name = "Daily Summary"
    if sum_name in wb.sheetnames:
        ws_sum = wb[sum_name]
    else:
        ws_sum = wb.create_sheet(sum_name)
        _write_headers(ws_sum, [
            "Date",
            "SLW x Sell Price (EUR)",
            "SLD x Sell Price (EUR)",
            "SLW x List Price (EUR)",
            "SLD x List Price (EUR)",
            "Product-Colors Total",
            "Product-Colors With Sales",
        ])

    ws_sum.append([
        today,
        summary["slw_x_sell_eur"],
        summary["sld_x_sell_eur"],
        summary["slw_x_list_eur"],
        summary["sld_x_list_eur"],
        summary["product_colors_total"],
        summary["product_colors_with_sales"],
    ])
    _autofit(ws_sum)

    # -- Sheet 2: Latest Detail (replaced each run) -------------------------
    det_name = "Latest Detail"
    if det_name in wb.sheetnames:
        del wb[det_name]
    ws_det = wb.create_sheet(det_name)
    _write_headers(ws_det, [
        "Product-Color Key",
        "Product Title",
        "Category",
        "Sold Past Week (SLW)",
        "Sold Past Days (SLD)",
        "Sell Price (EUR)",
        "List Price (EUR)",
        "SLW x Sell Price",
        "SLW x List Price",
        "SLD x Sell Price",
        "SLD x List Price",
    ])

    for r in sorted(rows, key=lambda x: -(x["sale_last_week"] * x["sell_price_eur"])):
        slw = r["sale_last_week"]
        sld = r["sale_last_days"]
        sp = r["sell_price_eur"]
        lp = r["list_price_eur"]
        ws_det.append([
            r["base_key"],
            r["title"],
            r["category"],
            slw,
            sld,
            sp,
            lp,
            round(slw * sp, 2),
            round(slw * lp, 2),
            round(sld * sp, 2),
            round(sld * lp, 2),
        ])
    _autofit(ws_det)

    wb.save(XLSX_PATH)
    print(f"  Saved -> {XLSX_PATH}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    today = date.today().isoformat()
    now = datetime.now().isoformat(timespec="seconds")
    print(f"[{now}] RVRC Sales Tracker (sale_last_week methodology)")
    print("=" * 60)

    state = load_state()

    # Guard: skip if already ran today
    if state.get("daily_summaries") and state["daily_summaries"][-1]["date"] == today:
        print(f"Already ran today ({today}). Delete the last entry in "
              f"daily_summaries from {STATE_FILE.name} to re-run.")
        return

    print("\nFetching FX rates ...")
    fx_rates = fetch_fx_rates()

    print("\nFetching inventory from Elevate API ...")
    data_by_market = fetch_all_markets()

    total_v = sum(len(v) for v in data_by_market.values())
    unique_v = len({k for m in data_by_market.values() for k in m})
    print(f"\n  Variants: {total_v:,} raw, {unique_v:,} unique across markets")

    if total_v == 0:
        print("  ERROR: No variants fetched. Check network / API.")
        return

    print("\nAggregating to product-colour level ...")
    rows = aggregate_product_colours(data_by_market, fx_rates)
    print(f"  Product-colour groups: {len(rows):,}")

    print("\nComputing summary ...")
    summary = compute_summary(rows)

    with_sales = summary["product_colors_with_sales"]
    print(f"  Product-colors with sales:  {with_sales:,} / {len(rows):,}")
    print(f"  SLW x Sell Price (EUR):     {summary['slw_x_sell_eur']:>14,.2f}")
    print(f"  SLD x Sell Price (EUR):     {summary['sld_x_sell_eur']:>14,.2f}")
    print(f"  SLW x List Price (EUR):     {summary['slw_x_list_eur']:>14,.2f}")
    print(f"  SLD x List Price (EUR):     {summary['sld_x_list_eur']:>14,.2f}")

    # Save state
    if not state.get("daily_summaries"):
        state["daily_summaries"] = []
    state["daily_summaries"].append({
        "date":      today,
        "timestamp": now,
        "fx_rates":  {k: round(v, 4) for k, v in fx_rates.items()},
        **summary,
    })

    # Store top-10 reference products for day-over-day SLW tracking
    ref = sorted(rows, key=lambda x: -x["sale_last_week"])[:10]
    state["reference_products"] = [
        {"base_key": r["base_key"], "title": r["title"],
         "sale_last_week": r["sale_last_week"], "date": today}
        for r in ref
    ]

    save_state(state)
    print(f"\n  State saved -> {STATE_FILE.name}")

    print("\nWriting Excel ...")
    write_excel(today, rows, summary)

    print("\nDone.")


if __name__ == "__main__":
    main()
