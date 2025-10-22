# scripts/adtraction_epc_combined.py
# Kombinerar Finance + Non-Finance. Robust kategorifångst (onclick/category-id ELLER direkta länkar).
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
from statistics import median
from datetime import date
import os, re, time, json, argparse, requests, urllib.parse as _url
from openpyxl import Workbook, load_workbook

BASE_ROOT = "https://secure.adtraction.com"
BASE = f"{BASE_ROOT}/partner"
STATE_PATH = "adtraction_state.json"

DATA_DIR = "data"
XLSX_PATH = os.path.join(DATA_DIR, "adtraction_epc_medians.xlsx")
SHEET_FIN = "Finance"
SHEET_NON = "Non-Finance"

COUNTRY_URLS = {
    "Sweden":       f"{BASE}/programs.htm?cid=1&asonly=false",
    "Denmark":      f"{BASE}/programs.htm?cid=12&asonly=false",
    "Finland":      f"{BASE}/programs.htm?cid=14&asonly=false",
    "Norway":       f"{BASE}/programs.htm?cid=33&asonly=false",
    "France":       f"{BASE}/programs.htm?cid=15&asonly=false",
    "Germany":      f"{BASE}/programs.htm?cid=16&asonly=false",
    "Italy":        f"{BASE}/programs.htm?cid=22&asonly=false",
    "Spain":        f"{BASE}/programs.htm?cid=42&asonly=false",
    "Netherlands":  f"{BASE}/programs.htm?cid=32&asonly=false",
    "Poland":       f"{BASE}/programs.htm?cid=34&asonly=false",
    "Switzerland":  f"{BASE}/programs.htm?cid=44&asonly=false",
}

COUNTRY_CCY = {
    "Sweden": "SEK","Denmark": "DKK","Finland": "EUR","Norway": "NOK",
    "France": "EUR","Germany": "EUR","Italy": "EUR","Spain": "EUR",
    "Netherlands": "EUR","Poland": "PLN","Switzerland": "CHF",
}

COUNTRY_ORDER = [
    "Sweden","Denmark","Finland","Norway",
    "France","Germany","Italy","Spain",
    "Netherlands","Poland","Switzerland"
]

# ---------- Parsning ----------
NUMBER_RE   = re.compile(r"(\d[\d\s\u00A0]*[.,]\d+|\d[\d\s\u00A0]*)")
CURR_CODE_RE = re.compile(r"\b(SEK|EUR|DKK|NOK|PLN|CHF)\b", re.IGNORECASE)
HAS_EUR_SYM  = re.compile(r"€")
HAS_PLN_SYM  = re.compile(r"zł", re.IGNORECASE)
HAS_CHF_SYM  = re.compile(r"\bCHF\b|(?<!\w)Fr(?!\w)")
HAS_KR_SYM   = re.compile(r"\bkr\.?\b", re.IGNORECASE)

FINANCE_LABELS = ["finans", "finance", "finanzen", "finanza", "finanze", "financiën", "finanse", "finanzas", "rahoitus"]

def parse_number(text: str):
    if not text: return None
    t = " ".join(text.split()).strip().lower()
    if t in {"ingen data","no data","-","—",""}: return None
    m = NUMBER_RE.search(t)
    if not m: return None
    raw = m.group(1).replace("\u00A0"," ").replace(" ","")
    if "," in raw and "." in raw: raw = raw.replace(".","")
    try: val = float(raw.replace(",", "."))
    except ValueError: return None
    return val if abs(val) >= 1e-12 else None

def detect_currency(cell_text: str, country_name: str):
    if not cell_text: return None
    t = " ".join(cell_text.split()).strip()
    m = CURR_CODE_RE.search(t)
    if m: return m.group(1).upper()
    if HAS_EUR_SYM.search(t): return "EUR"
    if HAS_PLN_SYM.search(t): return "PLN"
    if HAS_CHF_SYM.search(t): return "CHF"
    if HAS_KR_SYM.search(t) or "SEK" in t.upper():
        if "SEK" in t.upper(): return "SEK"
        if country_name == "Sweden": return "SEK"
        if country_name == "Denmark": return "DKK"
        if country_name == "Norway":  return "NOK"
        return "SEK"
    return None

def scrape_epc_values_from_table(page, country_name: str):
    out = []
    tables = page.locator("table")
    for i in range(tables.count()):
        table = tables.nth(i)
        headers = table.locator("thead tr th")
        if not headers.count(): continue
        heads = [headers.nth(j).inner_text().strip().lower() for j in range(headers.count())]
        if not any("epc" in h for h in heads): continue
        epc_idx = next(idx for idx, h in enumerate(heads) if "epc" in h)
        try: table.locator("tbody tr").first.wait_for(state="visible", timeout=6000)
        except PWTimeout: pass
        rows = table.locator("tbody tr")
        for r in range(rows.count()):
            cells = rows.nth(r).locator("td")
            if cells.count() <= epc_idx: continue
            txt = cells.nth(epc_idx).inner_text()
            val = parse_number(txt)
            if val is None: continue
            ccy = detect_currency(txt, country_name)
            out.append((val, ccy))
    return out

# ---------- Kategorier ----------
def extract_category_items(page):
    """
    Returnerar en lista med dicts:
      {"label": <text_lower>, "cid": <id_str> or None, "url": <abs_url> or None}
    Täcker både onclick('category(id)') OCH direkta länkar till listadvertprograms.htm.
    """
    items = []

    # 1) Plocka onclick/category-id, href med category(id), och data-attributes
    data = page.evaluate("""
(() => {
  const out = [];
  const rx = /category\\(\\s*['"]?(-?\\d+)['"]?\\s*\\)/i;
  document.querySelectorAll('*').forEach(el => {
    const text = (el.textContent || '').trim().replace(/\\s+/g, ' ');
    let cid = null, url = null;

    const oc = el.getAttribute && el.getAttribute('onclick');
    if (oc && rx.test(oc)) cid = rx.exec(oc)[1];

    const href = el.getAttribute && el.getAttribute('href');
    if (href) {
      if (rx.test(href)) cid = rx.exec(href)[1];
      if (/listadvertprograms\\.htm/i.test(href)) url = href;
    }

    const ds = el.dataset || {};
    for (const k of ['category','categoryId','cid','cat','cId']) {
      if (ds[k] && !cid) cid = String(ds[k]);
    }

    if (cid || url) out.push({text, cid, url});
  });
  return out;
})()
""") or []

    # 2) Normalisera: absolut URL, label lower-case
    for d in data:
        label = (d.get("text") or "").strip().lower()
        cid   = d.get("cid"); 
        url   = d.get("url")
        if url:
            if url.startswith("/"): url = BASE_ROOT + url
            elif not url.startswith("http"): url = BASE_ROOT + "/" + url.lstrip("./")
        if cid or url:
            items.append({"label": label, "cid": (str(cid) if cid is not None else None), "url": url})

    # 3) Deduplicera på (cid,url)
    seen = set(); uniq=[]
    for it in items:
        key = (it["cid"], it["url"])
        if key in seen: continue
        seen.add(key); uniq.append(it)
    return uniq

def pick_finance(items):
    """Välj finance-post via label. Returnerar ett dict från items eller None."""
    for it in items:
        if any(tok in (it["label"] or "") for tok in FINANCE_LABELS):
            return it
    return None

def make_list_url_from_cid(cid):
    # standardiserad list-URL om vi bara har ett category-id
    return f"{BASE}/listadvertprograms.htm?cId={cid}&asonly=false"

# ---------- Paginering ----------
def discover_pagination_urls(page, any_list_url):
    parsed = _url.urlparse(any_list_url)
    q = _url.parse_qs(parsed.query)
    cid_key = "cId" if "cId" in q else ("cid" if "cid" in q else None)
    cid_val = (q.get(cid_key, [""])[0] if cid_key else "")

    hrefs = page.evaluate("""() => Array.from(document.querySelectorAll('a[href]'), a => a.getAttribute('href'))""") or []
    urls = set([any_list_url])

    def abs_url(h):
        if not h: return None
        if h.startswith("http"): return h
        if h.startswith("/"):    return BASE_ROOT + h
        return BASE_ROOT + "/" + h.lstrip("./")

    for h in hrefs:
        u = abs_url(h)
        if not u or "listadvertprograms.htm" not in u.lower(): 
            continue
        pq = _url.parse_qs(_url.urlparse(u).query)
        # samma kategori?
        if cid_key and pq.get(cid_key, [""])[0] != cid_val:
            continue
        # någon sidparameter?
        if "page" in pq or "p" in pq:
            urls.add(u)

    def page_key(u):
        pq = _url.parse_qs(_url.urlparse(u).query)
        for k in ("page", "p"):
            if k in pq:
                try: return int(pq[k][0])
                except: pass
        return 1

    return sorted(urls, key=page_key)

# ---------- Login ----------
def looks_like_login(page):
    url = page.url.lower()
    if "login" in url or "signin" in url: return True
    if page.locator('input[type="password"]').count() > 0: return True
    if page.locator('text=/logga in|sign in|log in/i').count() > 0: return True
    return False

def auto_login_and_save_state(p, email: str, password: str, headless: bool = True):
    browser = p.chromium.launch(headless=headless)
    context = browser.new_context()
    page = context.new_page()
    page.goto(f"{BASE}/programs.htm?asonly=false", wait_until="domcontentloaded")

    if not looks_like_login(page):
        context.storage_state(path=STATE_PATH)
        context.close(); browser.close()
        return True

    def fill(sel, val):
        loc = page.locator(sel)
        if loc.count(): loc.first.fill(val); return True
        return False

    try: page.locator('a:has-text("Partner")').first.click(timeout=2000)
    except Exception: pass

    ok_email = (fill('input[type="email"]', email) or
                fill('input[name="email"]', email) or
                fill('input#email', email) or
                fill('input[id*="email" i]', email) or
                fill('input[name*="user" i]', email))
    ok_pwd = (fill('input[type="password"]', password) or
              fill('input[name="password"]', password) or
              fill('input#password', password) or
              fill('input[id*="pass" i]', password))
    if not (ok_email and ok_pwd):
        page.goto(f"{BASE}/login.htm", wait_until="domcontentloaded")
        ok_email = (fill('input[type="email"]', email) or
                    fill('input[name="email"]', email) or
                    fill('input#email', email) or
                    fill('input[id*="email" i]', email) or
                    fill('input[name*="user" i]', email))
        ok_pwd = (fill('input[type="password"]', password) or
                  fill('input[name="password"]', password) or
                  fill('input#password', password) or
                  fill('input[id*="pass" i]', password))

    clicked = False
    for sel in ['button[type="submit"]','input[type="submit"]',
                'button:has-text("Logga in")','button:has-text("Sign in")',
                'text=Logga in','text=Sign in']:
        try: page.locator(sel).first.click(timeout=1500); clicked=True; break
        except Exception: continue
    if not clicked:
        try: page.keyboard.press("Enter")
        except Exception: pass

    try: page.wait_for_load_state("networkidle", timeout=10000)
    except PWTimeout: pass

    page.goto(f"{BASE}/programs.htm?asonly=false", wait_until="domcontentloaded")
    ok = not looks_like_login(page)
    if ok: context.storage_state(path=STATE_PATH)
    context.close(); browser.close()
    return ok

# ---------- FX ----------
def fetch_fx_local_to_sek(currencies):
    need = sorted({c for c in currencies if c and c != "SEK"})
    out = {c: 1.0 for c in currencies}
    if not need: return out
    try:
        r = requests.get("https://open.er-api.com/v6/latest/SEK", timeout=15)
        r.raise_for_status(); data = r.json()
        if data.get("result") == "success":
            rates = data.get("rates", {})
            for c in need:
                per_sek = float(rates[c]); out[c] = 1.0/per_sek if per_sek else 1.0
            print("\nFX (open.er-api.com):")
            for c in need: print(f"  {c}→SEK = {out[c]:.4f}")
            return out
    except Exception as e:
        print(f"FX provider 1 failed ({e}); trying fallback…")
    symbols = ",".join(need)
    r = requests.get(f"https://api.exchangerate.host/latest?base=SEK&symbols={symbols}", timeout=15)
    r.raise_for_status(); rates = r.json().get("rates", {})
    for c in need:
        per_sek = float(rates[c]); out[c] = 1.0/per_sek if per_sek else 1.0
    print("\nFX (exchangerate.host):")
    for c in need: print(f"  {c}→SEK = {out[c]:.4f}")
    return out

# ---------- Excel ----------
def ensure_book_and_sheets(path):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    if not os.path.exists(path):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = SHEET_FIN
        ws2 = wb.create_sheet(SHEET_NON)
        header = ["date", "All (SEK)"] + [f"{c} (SEK)" for c in COUNTRY_ORDER]
        ws1.append(header); ws2.append(header); wb.save(path)
    else:
        wb = load_workbook(path)
        if SHEET_FIN not in wb.sheetnames:
            wb.create_sheet(SHEET_FIN).append(["date","All (SEK)"]+[f"{c} (SEK)" for c in COUNTRY_ORDER])
        if SHEET_NON not in wb.sheetnames:
            wb.create_sheet(SHEET_NON).append(["date","All (SEK)"]+[f"{c} (SEK)" for c in COUNTRY_ORDER])
        wb.save(path)

def append_row(path, sheet_name, dt, all_median, per_country):
    wb = load_workbook(path)
    ws = wb[sheet_name]
    row = [dt, (None if all_median is None else round(all_median,2))]
    for c in COUNTRY_ORDER:
        v = per_country.get(c)
        row.append(None if v is None else round(v,2))
    ws.append(row); wb.save(path)

def debug_dump(page, prefix):
    """Spara PNG + HTML för felsökning."""
    os.makedirs("pages", exist_ok=True)
    try:
        page.screenshot(path=f"pages/{prefix}.png", full_page=True)
    except Exception:
        pass
    try:
        html = page.content()
        with open(f"pages/{prefix}.html", "w", encoding="utf-8") as f:
            f.write(html)
    except Exception:
        pass    

# ---------- Core per-land ----------
def run_for_country(page, country, fx):
    results = {"finance": {"n":0, "median": None}, "non": {"n":0, "median": None}}
    ccy_expected = COUNTRY_CCY[country]

    # Landingssida
    page.goto(COUNTRY_URLS[country], wait_until="domcontentloaded")
    try: page.wait_for_load_state("networkidle", timeout=8000)
    except PWTimeout: pass

    items = extract_category_items(page)
    # välj finance
    fin_item = pick_finance(items)
    # skapa finance-url om bara cid fanns
    fin_url = None
    if fin_item:
        fin_url = fin_item["url"] or (make_list_url_from_cid(fin_item["cid"]) if fin_item["cid"] else None)

    # -- Finance --
    if fin_url:
        values_local = []
        page.goto(fin_url, wait_until="domcontentloaded")
        try: page.wait_for_selector("table", timeout=8000)
        except PWTimeout: pass
        for u in discover_pagination_urls(page, fin_url):
            page.goto(u, wait_until="domcontentloaded")
            try: page.wait_for_selector("table", timeout=8000)
            except PWTimeout: continue
            values_local.extend(scrape_epc_values_from_table(page, country))

        if values_local:
            values_sek = [val * fx.get((ccy or ccy_expected).upper(), 1.0) for val, ccy in values_local]
            results["finance"]["n"] = len(values_local)
            results["finance"]["median"] = median(values_sek)
            print(f"[{country}] n={results['finance']['n']}  median={results['finance']['median']:.2f} SEK")
        else:
            print(f"[{country}] No EPC values found (Finance).")
    else:
        print(f"[{country}] No Finance category link found.")

    # -- Non-Finance --
    values_local_nf = []
    non_links = []
    for it in items:
        if fin_item and ((it["url"] == fin_item.get("url")) or (it["cid"] and it["cid"] == fin_item.get("cid"))):
            continue
        url = it["url"] or (make_list_url_from_cid(it["cid"]) if it["cid"] else None)
        if url:
            non_links.append(url)

    # deduplicera & logga vad vi faktiskt tänker besöka
    non_links = sorted(set(non_links))
    print(f"[{country}] Non-Finance categories detected: {len(non_links)}")
    for i, lk in enumerate(non_links[:10], 1):
        print(f"   NF link {i}: {lk}")
    if len(non_links) > 10:
        print(f"   ... +{len(non_links)-10} more")

    for idx, link in enumerate(non_links, 1):
        page.goto(link, wait_until="domcontentloaded")
        try:
            page.wait_for_selector("table", timeout=15000)
        except PWTimeout:
            # om ingen tabell dök upp—dumpa för att se sidan
            debug_dump(page, f"{country}_nf_{idx}_no_table")
            continue

        # samla alla pagineringssidor
        page_urls = discover_pagination_urls(page, link)
        if len(page_urls) == 1:
            print(f"   {country} NF page has no pagination: {link}")
        else:
            print(f"   {country} NF pagination pages: {len(page_urls)}")

        found_any = False
        for pidx, u in enumerate(page_urls, 1):
            page.goto(u, wait_until="domcontentloaded")
            try:
                page.wait_for_selector("table", timeout=15000)
            except PWTimeout:
                debug_dump(page, f"{country}_nf_{idx}_p{pidx}_no_table")
                continue

            vals_before = len(values_local_nf)
            values_local_nf.extend(scrape_epc_values_from_table(page, country))
            if len(values_local_nf) == vals_before:
                # ingen EPC hittad på just denna sida—dumpa för att se kolumnrubriker etc.
                debug_dump(page, f"{country}_nf_{idx}_p{pidx}_no_epc")
            else:
                found_any = True

    if values_local_nf:
        values_sek_nf = [val * fx.get((ccy or ccy_expected).upper(), 1.0) for val, ccy in values_local_nf]
        results["non"]["n"] = len(values_local_nf)
        results["non"]["median"] = median(values_sek_nf)
        print(f"[{country}] n={results['non']['n']} median_ex_fin={results['non']['median']:.2f} SEK")
    else:
        print(f"[{country}] No EPC values found (Non-Finance).")

    return results

# ---------- Main ----------
def main():
    parser = argparse.ArgumentParser(add_help=False)
    parser.add_argument("countries", nargs="*", help="Optional filter: run only matching countries (partial ok)")
    parser.add_argument("--headful", action="store_true")
    args = parser.parse_args()

    countries = COUNTRY_ORDER
    if args.countries:
        sel=[]
        for r in args.countries:
            sel.extend([c for c in COUNTRY_ORDER if r.lower() in c.lower()])
        seen=set(); countries=[c for c in sel if not (c in seen or seen.add(c))] or COUNTRY_ORDER

    fx = fetch_fx_local_to_sek({COUNTRY_CCY[c] for c in countries})
    ensure_book_and_sheets(XLSX_PATH)
    today = date.today().isoformat()

    finance_by_country = {}
    non_by_country = {}

    with sync_playwright() as p:
        # login
        need_login = not os.path.exists(STATE_PATH)
        if not need_login:
            tb = p.chromium.launch(headless=True)
            tc = tb.new_context(storage_state=STATE_PATH)
            tp = tc.new_page()
            tp.goto(f"{BASE}/programs.htm?asonly=false", wait_until="domcontentloaded")
            need_login = looks_like_login(tp)
            tc.close(); tb.close()
        if need_login:
            email = os.environ.get("ADTRACTION_EMAIL","")
            pwd   = os.environ.get("ADTRACTION_PASSWORD","")
            if not (email and pwd) or not auto_login_and_save_state(p, email, pwd, headless=True):
                print("Auto-login failed or credentials missing. Set ADTRACTION_EMAIL & ADTRACTION_PASSWORD.")
                return

        browser = p.chromium.launch(headless=not args.headful)
        context = browser.new_context(storage_state=STATE_PATH, viewport={"width":1440,"height":900} if args.headful else None)
        page = context.new_page()

        for country in countries:
            try:
                res = run_for_country(page, country, fx)
                if res["finance"]["n"] > 0: finance_by_country[country] = res["finance"]["median"]
                if res["non"]["n"] > 0:     non_by_country[country]     = res["non"]["median"]
            except Exception as e:
                print(f"[{country}] Error: {e}")

        context.close(); browser.close()

    fin_all = median(list(finance_by_country.values())) if finance_by_country else None
    non_all = median(list(non_by_country.values())) if non_by_country else None

    append_row(XLSX_PATH, SHEET_FIN, today, fin_all, finance_by_country)
    append_row(XLSX_PATH, SHEET_NON, today, non_all, non_by_country)
    print(f"\nWrote {XLSX_PATH} → sheets [{SHEET_FIN}] & [{SHEET_NON}] for {today}")

if __name__ == "__main__":
    main()
