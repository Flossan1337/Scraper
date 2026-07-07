#!/usr/bin/env python3
"""
track_ahlsell_led_panel_inventory.py

Spårar lagernivåer för infällda armaturer (LED-paneler) på Ahlsell.se.
Aggregerar lagret per varumärke och sparar en daglig tidsserie i Excel.

Centrallager (Logistikcentrum i Hallsberg) är INTE tillgängligt via det
publika API:et – endast de ~100 lokala butikerna returneras av
/api/warehouses/stock.

Flöde
-----
  1. Hämta alla produkter i kategorin "Infällda armaturer" via sök-API
     (paginerat, filtrerar klientsidan på item_category5)
  2. Expandera varianter per produkt (alla artikelnummer)
  3. Hämta och cachelagra butikskatalog
  4. Hämta lagersaldo per artikelnummer
  5. Summera lager per varumärke
  6. Spara daglig snapshot i JSON-tillståndsfil
  7. Exportera till Excel (laddas och uppdateras inkrementellt):
       • "Varumärken"  — tidsserie: rader = datum, kolumner = varumärken
       • "Artiklar"    — senaste snapshotens artikeldetaljer (skrivs om)

API:er (ingen autentisering krävs)
-----------------------------------
  Sök:       GET https://www.ahlsell.se/api/search
             ?searchPhrase=infallda+armaturer&pageSize=200&page={n}
  Varianter: GET https://www.ahlsell.se/api/search/variants
             ?productCode={code}&activeVariantNumber={variantNumber}
  Butiker:   GET https://www.ahlsell.se/api/warehouses
  Lager:     GET https://www.ahlsell.se/api/warehouses/stock
             ?variantNumber={articleNumber}

Tillståndsfil  : data/ahlsell_led_panel_state.json
Excel-utdata   : data/ahlsell_led_panel_inventory.xlsx
"""

import json
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date
from pathlib import Path

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Konfiguration ──────────────────────────────────────────────────────────────
BASE_URL       = "https://www.ahlsell.se"
SEARCH_URL     = f"{BASE_URL}/api/search"
VARIANTS_URL   = f"{BASE_URL}/api/search/variants"
WAREHOUSES_URL = f"{BASE_URL}/api/warehouses"
STOCK_URL      = f"{BASE_URL}/api/warehouses/stock"

SEARCH_PHRASE    = "infallda armaturer"
TARGET_CAT5      = "Infällda armaturer"   # item_category5 filter
PAGE_SIZE        = 200
REQUEST_DELAY    = 0.15  # sekunder mellan sekventiella anrop
STOCK_WORKERS    = 24    # parallella trådar för lagerhämtning

STATE_FILE = Path(__file__).parent.parent / "data" / "ahlsell_led_panel_state.json"
EXCEL_FILE = Path(__file__).parent.parent / "data" / "ahlsell_led_panel_inventory.xlsx"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept": "application/json",
    "Accept-Language": "sv-SE,sv;q=0.9",
}

PLEJD_BRAND = "Plejd"

# ── Stilar ─────────────────────────────────────────────────────────────────────
_EVEN_FILL   = PatternFill("solid", fgColor="F2F7FD")
_CENTER      = Alignment(horizontal="center", vertical="center")
_PLEJD_FILL  = PatternFill("solid", fgColor="E8F5E9")   # ljusgrön rad för Plejd-datum (ej använd här)
_HDR_DATE_FILL = PatternFill("solid", fgColor="375623")  # mörkgrön – datumkolumn
_HDR_PLEJD_FILL = PatternFill("solid", fgColor="1A5276") # mörkblå – Plejd-kolumn
_HDR_OTHER_FILL = PatternFill("solid", fgColor="4A4A4A") # antracit – övriga varumärken


# ── API-anrop ──────────────────────────────────────────────────────────────────

def fetch_category_products() -> list[dict]:
    """
    Hämtar alla produktkort i kategorin 'Infällda armaturer' via sök-API.
    Paginerar tills alla sidor hämtats och filtrerar på item_category5.
    """
    all_cards: list[dict] = []
    page = 1
    total_count: int | None = None

    while True:
        time.sleep(REQUEST_DELAY)
        resp = requests.get(
            SEARCH_URL,
            params={"searchPhrase": SEARCH_PHRASE, "pageSize": PAGE_SIZE, "page": page},
            headers=HEADERS,
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()

        if total_count is None:
            total_count = data.get("productCount", 0)

        cards = data.get("productCards") or []
        for card in cards:
            try:
                attrs = json.loads(card.get("trackingProductAttributes", "{}"))
            except (json.JSONDecodeError, TypeError):
                attrs = {}
            if attrs.get("item_category5") == TARGET_CAT5:
                all_cards.append(card)

        fetched = (page - 1) * PAGE_SIZE + len(cards)
        if not cards or fetched >= total_count:
            break
        page += 1

    return all_cards


def fetch_all_variant_numbers(product_code: str, active_variant: str) -> list[str]:
    """Returnerar artikelnummer för samtliga varianter av en produkt."""
    resp = requests.get(
        VARIANTS_URL,
        params={"productCode": product_code, "activeVariantNumber": active_variant},
        headers=HEADERS,
        timeout=30,
    )
    resp.raise_for_status()
    return [item["code"] for item in resp.json().get("items", []) if item.get("code")]


def fetch_warehouses() -> dict[str, dict]:
    """Hämtar butikskatalog och returnerar som {warehouseId: metadata}."""
    resp = requests.get(WAREHOUSES_URL, headers=HEADERS, timeout=30)
    resp.raise_for_status()
    return {
        str(w["id"]): {
            "name":        w.get("name", ""),
            "city":        w.get("city", ""),
            "address":     w.get("address", ""),
            "phone":       w.get("phoneNumber", ""),
            "url_segment": w.get("urlSegment", ""),
        }
        for w in resp.json()
    }


def fetch_stock_total(variant_number: str) -> float:
    """
    Hämtar lagersaldo för ett artikelnummer och returnerar total kvantitet
    över samtliga butiker. Returnerar 0.0 vid fel.
    """
    resp = requests.get(
        STOCK_URL,
        params={"variantNumber": variant_number},
        headers=HEADERS,
        timeout=30,
    )
    resp.raise_for_status()
    return sum(
        entry.get("stock", {}).get("quantity") or 0.0
        for entry in resp.json()
        if (entry.get("stock", {}).get("quantity") or 0.0) > 0
    )


# ── Insamling ──────────────────────────────────────────────────────────────────

def collect_snapshot() -> tuple[dict[str, dict], dict[str, float], dict[str, float]]:
    """
    Kör alla API-anrop och returnerar:
      products       : {articleNumber: {product_name, brand, product_code, page_url}}
      stock_by_article: {articleNumber: total_quantity}
      stock_by_brand  : {brand: total_quantity}
    """
    # 1. Produktlista
    print("Hämtar produkter i kategorin 'Infällda armaturer'...")
    product_cards = fetch_category_products()
    print(f"  {len(product_cards)} produkter hittade i kategorin")

    # 2. Expandera varianter (parallellt för produkter med >1 variant)
    def expand_card(card: dict) -> list[tuple[str, dict]]:
        """Returnerar lista av (articleNumber, metadata) för ett produktkort."""
        product_code  = card["code"]
        product_name  = card["name"]
        brand         = card.get("brand", "Okänt")
        most_relevant = str(card["variantNumber"])
        num_variants  = card.get("numberOfVariants", 1)
        page_url      = card.get("firstVariationPageUrl", "")
        meta = {"product_name": product_name, "brand": brand,
                "product_code": product_code, "page_url": page_url}

        if num_variants > 1:
            try:
                variant_numbers = fetch_all_variant_numbers(product_code, most_relevant)
            except Exception as exc:
                print(f"  Varning: kunde ej hämta varianter för {product_name}: {exc}")
                variant_numbers = [most_relevant]
        else:
            variant_numbers = [most_relevant]

        return [(str(vn), meta) for vn in variant_numbers]

    print(f"  Expanderar varianter ({STOCK_WORKERS} parallella trådar)...")
    products: dict[str, dict] = {}
    with ThreadPoolExecutor(max_workers=STOCK_WORKERS) as pool:
        for entries in pool.map(expand_card, product_cards):
            for art_num, meta in entries:
                products[art_num] = meta

    print(f"  Totalt {len(products)} artikelnummer (inkl. alla varianter)")

    # 3. Lagersaldo per artikel (parallellt)
    print(f"Hämtar lagersaldo ({len(products)} artiklar, {STOCK_WORKERS} parallella trådar)...")
    stock_by_article: dict[str, float] = {}
    article_list = list(products.keys())
    completed = 0

    with ThreadPoolExecutor(max_workers=STOCK_WORKERS) as pool:
        futures = {pool.submit(fetch_stock_total, art): art for art in article_list}
        for future in as_completed(futures):
            art_num = futures[future]
            try:
                stock_by_article[art_num] = future.result()
            except Exception as exc:
                print(f"  Varning: lagerfel för {art_num}: {exc}")
                stock_by_article[art_num] = 0.0
            completed += 1
            if completed % 100 == 0:
                print(f"  {completed}/{len(article_list)} artiklar klara...")

    total = sum(stock_by_article.values())
    print(f"  Klart — {total:.0f} enheter totalt i lager")

    # 4. Aggregera per varumärke
    stock_by_brand: dict[str, float] = {}
    for art_num, qty in stock_by_article.items():
        brand = products[art_num]["brand"]
        stock_by_brand[brand] = stock_by_brand.get(brand, 0.0) + qty

    return products, stock_by_article, stock_by_brand


# ── Tillstånd ──────────────────────────────────────────────────────────────────

def load_state() -> dict:
    if STATE_FILE.exists():
        return json.loads(STATE_FILE.read_text(encoding="utf-8"))
    return {"warehouses": {}, "products": {}, "snapshots": {}}


def save_state(state: dict) -> None:
    STATE_FILE.write_text(
        json.dumps(state, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def update_state(
    state: dict,
    products: dict,
    warehouses: dict,
    stock_by_article: dict[str, float],
    stock_by_brand: dict[str, float],
) -> None:
    today = date.today().isoformat()
    state["warehouses"]   = warehouses
    state["products"]     = products
    state["last_updated"] = today
    state.setdefault("snapshots", {})[today] = {
        "by_brand":   {b: round(v) for b, v in stock_by_brand.items()},
        "by_article": {a: round(v) for a, v in stock_by_article.items()},
    }


# ── Excel ──────────────────────────────────────────────────────────────────────

def _sorted_brands(snapshots: dict) -> list[str]:
    """
    Returnerar en brandlista sorterad efter lagerstorlek (senaste snapshot),
    störst till vänster. Plejd placeras i sin naturliga position.
    """
    all_brands: set[str] = set()
    for snap in snapshots.values():
        all_brands.update(snap.get("by_brand", {}).keys())

    # Använd senaste snapshotens lagernivåer som sorteringsnyckeln
    latest = snapshots[max(snapshots.keys())].get("by_brand", {})
    return sorted(all_brands, key=lambda b: latest.get(b, 0), reverse=True)


def _write_brands_sheet(ws, snapshots: dict, brands: list[str]) -> None:
    """
    Skriver/ersätter innehållet i sheet 'Varumärken'.
    Rader = datum (kronologisk ordning), kolumner = varumärken.
    """
    # Rubrikrad
    ws.cell(1, 1, "Datum").fill  = PatternFill("solid", fgColor="375623")
    ws.cell(1, 1).font           = Font(color="FFFFFF", bold=True, size=11)
    ws.cell(1, 1).alignment      = _CENTER

    for col, brand in enumerate(brands, start=2):
        cell = ws.cell(1, col, brand)
        fill_color = "1A5276" if brand == PLEJD_BRAND else "4A4A4A"
        cell.fill      = PatternFill("solid", fgColor=fill_color)
        cell.font      = Font(color="FFFFFF", bold=True, size=11)
        cell.alignment = _CENTER

    ws.row_dimensions[1].height = 22

    # Datarader
    sorted_dates = sorted(snapshots.keys())
    for row, d in enumerate(sorted_dates, start=2):
        by_brand = snapshots[d].get("by_brand", {})
        ws.cell(row, 1, d).alignment = _CENTER
        for col, brand in enumerate(brands, start=2):
            val = by_brand.get(brand)
            ws.cell(row, col, val)
        if row % 2 == 0:
            for col in range(1, 2 + len(brands)):
                ws.cell(row, col).fill = _EVEN_FILL

    # Kolumnbredder
    ws.column_dimensions["A"].width = 13
    for col in range(2, 2 + len(brands)):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = 18

    ws.freeze_panes = "B2"


def _write_articles_sheet(ws, snapshots: dict, products: dict) -> None:
    """
    Skriver senaste snapshotens artikeldetaljer.
    """
    if not snapshots:
        return

    latest_date = max(snapshots.keys())
    by_article  = snapshots[latest_date].get("by_article", {})

    hdr_fill = PatternFill("solid", fgColor="4A4A4A")
    hdr_font = Font(color="FFFFFF", bold=True, size=11)
    headers  = ["Artikel", "Produktnamn", "Varumärke", f"Lager ({latest_date})"]

    for col, label in enumerate(headers, start=1):
        cell            = ws.cell(1, col, label)
        cell.fill       = hdr_fill
        cell.font       = hdr_font
        cell.alignment  = _CENTER

    ws.row_dimensions[1].height = 22

    # Sortera: Plejd-artiklar överst, sedan per varumärke + produktnamn
    def sort_key(item):
        art, meta = item
        brand = meta.get("brand", "")
        return (0 if brand == PLEJD_BRAND else 1, brand, meta.get("product_name", ""))

    sorted_articles = sorted(products.items(), key=sort_key)

    for row, (art, meta) in enumerate(sorted_articles, start=2):
        brand = meta.get("brand", "")
        stock = by_article.get(art)
        for col, val in enumerate(
            [art, meta.get("product_name", ""), brand, stock], start=1
        ):
            ws.cell(row, col, val).alignment = _CENTER if col in (1, 4) else Alignment()
        if row % 2 == 0:
            for col in range(1, 5):
                ws.cell(row, col).fill = _EVEN_FILL

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 18
    ws.freeze_panes = "A2"


def write_excel(state: dict) -> None:
    snapshots = state.get("snapshots", {})
    products  = state.get("products", {})

    if not snapshots:
        print("Ingen data att exportera.")
        return

    brands = _sorted_brands(snapshots)

    # Ladda befintlig arbetsbok eller skapa ny
    if EXCEL_FILE.exists():
        wb = load_workbook(EXCEL_FILE)
    else:
        wb = Workbook()
        # Ta bort standard-sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # ── Sheet 1: Varumärken ────────────────────────────────────────────────────
    if "Varumärken" in wb.sheetnames:
        del wb["Varumärken"]
    ws_brands = wb.create_sheet("Varumärken", 0)
    _write_brands_sheet(ws_brands, snapshots, brands)

    # ── Sheet 2: Artiklar ──────────────────────────────────────────────────────
    if "Artiklar" in wb.sheetnames:
        del wb["Artiklar"]
    ws_articles = wb.create_sheet("Artiklar")
    _write_articles_sheet(ws_articles, snapshots, products)

    wb.save(EXCEL_FILE)
    print(f"Excel sparad: {EXCEL_FILE}")


# ── Main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    today = date.today().isoformat()
    print(f"=== Ahlsell LED-panel lageruppföljning — {today} ===\n")

    state = load_state()

    if today in state.get("snapshots", {}):
        print(f"Snapshot för {today} finns redan. Exporterar Excel...")
        write_excel(state)
        return

    products, stock_by_article, stock_by_brand = collect_snapshot()

    print("\nHämtar butikskatalog...")
    try:
        warehouses = fetch_warehouses()
        print(f"  {len(warehouses)} butiker")
    except Exception as exc:
        print(f"  Varning: kunde ej hämta butiker: {exc}")
        warehouses = state.get("warehouses", {})

    update_state(state, products, warehouses, stock_by_article, stock_by_brand)
    save_state(state)
    print(f"\nTillstånd sparat: {STATE_FILE}")

    print("\nVarumärkessammanfattning:")
    plejd_total = stock_by_brand.get(PLEJD_BRAND, 0)
    grand_total = sum(stock_by_brand.values())
    for brand in ([PLEJD_BRAND] if PLEJD_BRAND in stock_by_brand else []) + sorted(
        b for b in stock_by_brand if b != PLEJD_BRAND
    ):
        print(f"  {brand:<30} {stock_by_brand[brand]:>6.0f} enheter")
    print(f"  {'TOTALT':<30} {grand_total:>6.0f} enheter")
    print(f"\n  Plejd andel: {plejd_total/grand_total*100:.1f}% av totalt lager")

    write_excel(state)
    print("\nKlart!")


if __name__ == "__main__":
    main()
