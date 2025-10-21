# scripts/litium_radar_bigquery.py
# -*- coding: utf-8 -*-

import os
import datetime as dt
from collections import defaultdict

import pandas as pd
from google.cloud import bigquery
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# --- konfig ---
DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data")
XLSX_PATH = os.path.join(DATA_DIR, "litium_radar.xlsx")
SINCE = dt.date(2024, 1, 1)
UNTIL = dt.date.today().replace(day=1)  # nuvarande månad
TLD_FILTER = (".se", ".no", ".dk", ".fi")  # ändra/utöka v.b.

# HTTP Archive tabellmönster:
#   `httparchive.technologies.*` där _TABLE_SUFFIX = YYYY_MM_DD (månadens crawl)
# Kolumner (relevanta): page (URL), app (tekniknamn)
# Vi extraherar host och filtrerar på app='Litium'.

SQL = """
SELECT
  _TABLE_SUFFIX AS crawl_date,
  REGEXP_EXTRACT(LOWER(page), r'^https?://([^/]+)') AS host
FROM `httparchive.technologies.*`
WHERE app = 'Litium'
  AND _TABLE_SUFFIX BETWEEN @start_suffix AND @end_suffix
"""

def month_iter(start_date, end_date):
    y, m = start_date.year, start_date.month
    last = (end_date.year, end_date.month)
    while (y, m) <= last:
        yield y, m
        if m == 12:
            y += 1; m = 1
        else:
            m += 1

def suffix_range(start_date, end_date):
    # HA använder ungefär första dagen i månaden (YYYY_MM_01) som suffix
    s = f"{start_date.year:04d}_{start_date.month:02d}_01"
    e = f"{end_date.year:04d}_{end_date.month:02d}_31"  # safe upper bound
    return s, e

def ensure_data_dir():
    os.makedirs(DATA_DIR, exist_ok=True)

def main():
    print(f"[i] Querying HTTP Archive (BigQuery) for Litium: {SINCE:%Y-%m} → {UNTIL:%Y-%m}")
    client = bigquery.Client()

    start_suffix, end_suffix = suffix_range(SINCE, UNTIL)
    job_config = bigquery.QueryJobConfig(
        query_parameters=[
            bigquery.ScalarQueryParameter("start_suffix", "STRING", start_suffix),
            bigquery.ScalarQueryParameter("end_suffix", "STRING", end_suffix),
        ]
    )

    # Kör en enda query över hela perioden (billigare/snabbare)
    df = client.query(SQL, job_config=job_config).result().to_dataframe()
    # Rensa ev None och extrahera TLD-filter
    df = df.dropna(subset=["host"])
    df["host"] = df["host"].str.strip()

    if TLD_FILTER:
        df = df[df["host"].str.endswith(TLD_FILTER)]

    # Normalisera suffix → YYYY-MM
    # _TABLE_SUFFIX = YYYY_MM_DD
    df["Month"] = df["crawl_date"].str.slice(0, 7).str.replace("_", "-", regex=False)  # "YYYY_MM" -> "YYYY-MM"

    # Antal unika domäner per månad
    monthly_hosts = (
        df.groupby("Month")["host"]
          .nunique()
          .reset_index(name="HighCount")  # "HighCount" ~ Litium-detekterad i HA (Wappalyzer)
          .sort_values("Month")
          .reset_index(drop=True)
    )

    # Hitta "nya" domäner per månad: första gången en host syns i hela serien
    first_seen = (
        df.groupby("host")["Month"]
          .min()
          .reset_index(name="FirstMonth")
    )
    # bygga lista per månad
    new_by_month = defaultdict(list)
    for _, row in first_seen.iterrows():
        new_by_month[row["FirstMonth"]].append(row["host"])

    # Se till att alla månader i intervallet finns med (även om 0)
    all_months = [f"{y:04d}-{m:02d}" for (y, m) in month_iter(SINCE, UNTIL)]
    monthly_hosts = monthly_hosts.set_index("Month").reindex(all_months, fill_value=0).reset_index()
    monthly_hosts.columns = ["Month", "HighCount"]

    # --- skriv Excel ---
    ensure_data_dir()
    wb = Workbook()

    # Summary
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Month", "HighCount"])
    for _, r in monthly_hosts.iterrows():
        ws.append([r["Month"], int(r["HighCount"])])
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12

    # NewDomains
    ws2 = wb.create_sheet("NewDomains")
    for col_idx, ym in enumerate(all_months, start=1):
        ws2.cell(row=1, column=col_idx, value=ym)
        doms = sorted(new_by_month.get(ym, []))
        for r, host in enumerate(doms, start=2):
            ws2.cell(row=r, column=col_idx, value=host)
        ws2.column_dimensions[get_column_letter(col_idx)].width = 28

    wb.save(XLSX_PATH)
    print(f"[ok] Wrote Excel → {XLSX_PATH}")

if __name__ == "__main__":
    main()
