import asyncio
import re
import os
from datetime import date
from pathlib import Path
from playwright.async_api import async_playwright, TimeoutError as PWTimeout
import openpyxl

# ── KONFIGURATION ──────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).resolve().parent
DATA_DIR = (SCRIPT_DIR / ".." / "data").resolve()
DATA_DIR.mkdir(parents=True, exist_ok=True)
XLSX_PATH = str((DATA_DIR / "fractal_amazon_ranks_us.xlsx").resolve())
SHEET_NAME = "Rankings_US"

# Produkter och kategori
PRODUCTS = [
    # Headsets
    {"name": "Scape Dark",          "asin": "B0D5HGK3C2", "category": "Computer Headsets"},
    {"name": "Scape Light",         "asin": "B0D5HK6JRS", "category": "Computer Headsets"},
    
    # Gaming Chairs (Refine)
    {"name": "Refine Mesh Light",   "asin": "B0CSYXY8FD", "category": "Computer Gaming Chairs"},
    {"name": "Refine Fabric Light", "asin": "B0CSYXYX39", "category": "Computer Gaming Chairs"},
    {"name": "Refine Mesh Dark",    "asin": "B0CSYYMTT4", "category": "Computer Gaming Chairs"},
    {"name": "Refine Fabric Dark",  "asin": "B0CSYWWRSV", "category": "Computer Gaming Chairs"},
]

HEADLESS = True 

# ───────────────────────────────────────────────────────────────────────────

def parse_rank_from_string(text: str, category: str) -> int:
    """
    Extraherar siffran från strängen "#15 in Computer Headsets".
    """
    # Vi letar efter mönstret: #123 (någonting) Kategori
    # Exempel: "#15 in Computer Headsets"
    # Vi är strikta med att kategorinamnet måste finnas med för att undvika sub-kategorier
    pattern = r"#([0-9,]+)\s+in\s+.*?" + re.escape(category)
    match = re.search(pattern, text, re.IGNORECASE)
    
    if match:
        clean_number = match.group(1).replace(",", "")
        return int(clean_number)
    return 0

async def handle_blockers(page):
    """Klickar bort popups"""
    try:
        blockers = ['text="Continue shopping"', 'input.a-button-input']
        for sel in blockers:
            if await page.locator(sel).is_visible(timeout=1000):
                await page.click(sel)
                await page.wait_for_timeout(1000)
                return
    except:
        pass

async def get_exact_rank(page, category: str) -> int:
    """
    Letar upp det exakta HTML-elementet baserat på användarens bild.
    """
    rank = 0
    
    # --- METOD 1: Tabell-layout (Från din bild) ---
    # Vi letar efter en tabellrad (tr) som innehåller texten "Best Sellers Rank"
    # Sedan hämtar vi texten från den raden.
    try:
        # Locator: Hitta en 'tr' som har en 'th' med texten "Best Sellers Rank"
        # Detta är extremt specifikt och bör undvika "Compare"-tabeller
        table_row = page.locator("tr:has(th:text-is('Best Sellers Rank'))").first
        
        if await table_row.count() > 0:
            row_text = await table_row.inner_text()
            # row_text ser ut typ: "Best Sellers Rank #15 in Computer Headsets..."
            val = parse_rank_from_string(row_text, category)
            if val > 0:
                return val
    except Exception:
        pass

    # --- METOD 2: Bullet List-layout (Alternativ Amazon-design) ---
    # Ibland använder Amazon punktlistor istället för tabeller.
    try:
        # Leta efter list-item som innehåller texten
        list_item = page.locator("li:has-text('Best Sellers Rank')").first
        if await list_item.count() > 0:
            li_text = await list_item.inner_text()
            val = parse_rank_from_string(li_text, category)
            if val > 0:
                return val
    except Exception:
        pass

    return 0

async def get_product_rank(context, asin: str, category: str, name: str) -> int:
    page = await context.new_page()
    url = f"https://www.amazon.com/dp/{asin}?th=1&psc=1&language=en_US"
    print(f"  Fetching rank for {name} ({category})...")

    try:
        await page.goto(url, wait_until="domcontentloaded", timeout=30000)
        await handle_blockers(page)

        # Använd den nya "kirurgiska" funktionen
        rank = await get_exact_rank(page, category)
        
        if rank > 0:
            print(f"    ✅ Rank: #{rank}")
        else:
            print(f"    ⚠️ Could not find specific rank for '{category}'.")
            # Om du vill debugga, avkommentera raden nedan för att se vad scriptet ser:
            # await page.screenshot(path=f"debug_{name.replace(' ', '_')}.png")
        
        return rank

    except Exception as e:
        print(f"    ❌ Error: {e}")
        return 0
    finally:
        await page.close()

def append_to_excel(data_dict):
    file_exists = os.path.exists(XLSX_PATH)
    
    if not file_exists:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        headers = ["Date"] + [p["name"] for p in PRODUCTS]
        ws.append(headers)
    else:
        wb = openpyxl.load_workbook(XLSX_PATH)
        if SHEET_NAME in wb.sheetnames:
            ws = wb[SHEET_NAME]
        else:
            ws = wb.create_sheet(SHEET_NAME)
            headers = ["Date"] + [p["name"] for p in PRODUCTS]
            ws.append(headers)

    row = [data_dict["Date"]]
    for p in PRODUCTS:
        row.append(data_dict.get(p["name"], 0))
    
    ws.append(row)
    wb.save(XLSX_PATH)
    print(f"💾 Data saved to {XLSX_PATH}")

async def run():
    today = date.today().isoformat()
    results = {"Date": today}

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=HEADLESS)
        # Vi sätter fönstret stort för att undvika mobil-layout
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
            viewport={"width": 1920, "height": 1080}, 
            locale="en-US"
        )
        
        print(f"--- Starting Rank Scraping ({today}) ---")

        for prod in PRODUCTS:
            rank = await get_product_rank(context, prod["asin"], prod["category"], prod["name"])
            results[prod["name"]] = rank
        
        await browser.close()

    append_to_excel(results)

if __name__ == "__main__":
    asyncio.run(run())