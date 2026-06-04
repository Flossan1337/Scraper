#!/usr/bin/env python3
"""
track_duroc_machines.py

Spårar antalet "Maskiner installerade" på Duroc Machine Tools startsida.

Flöde
-----
  1. Hämtar https://www.durocmachinetool.se/ med requests
  2. Parsar ut h2-taggen med antalet maskiner via BeautifulSoup
  3. Sparar daglig snapshot i JSON-tillståndsfil
  4. Exporterar tidsserie till Excel

Tillståndsfil : data/duroc_machines_state.json
Excel-utdata  : data/duroc_machines.xlsx
"""

import json
from datetime import date
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Konfiguration ──────────────────────────────────────────────────────────────
URL = "https://www.durocmachinetool.se/"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "sv-SE,sv;q=0.9,en;q=0.8",
}

SCRIPT_DIR  = Path(__file__).resolve().parent
STATE_FILE  = SCRIPT_DIR.parent / "data" / "duroc_machines_state.json"
EXCEL_FILE  = SCRIPT_DIR.parent / "data" / "duroc_machines.xlsx"

# ── Scraping ───────────────────────────────────────────────────────────────────

def fetch_machine_count() -> int:
    """Hämtar och returnerar antalet maskiner installerade från Durocs startsida."""
    resp = requests.get(URL, headers=HEADERS, timeout=30)
    resp.raise_for_status()

    soup = BeautifulSoup(resp.text, "lxml")

    # Hitta <p> med texten "Maskiner installerade" och gå upp till föräldern
    # för att hitta syskon-<h2>-taggen med värdet.
    for p in soup.find_all("p"):
        if "Maskiner installerade" in p.get_text():
            h2 = p.find_previous_sibling("h2")
            if h2:
                return int(h2.get_text(strip=True).replace("\xa0", "").replace(" ", ""))

    raise ValueError("Kunde inte hitta 'Maskiner installerade' på sidan.")

# ── Tillstånd ──────────────────────────────────────────────────────────────────

def load_state() -> dict:
    if STATE_FILE.exists():
        return json.loads(STATE_FILE.read_text(encoding="utf-8"))
    return {"history": []}


def save_state(state: dict) -> None:
    STATE_FILE.parent.mkdir(parents=True, exist_ok=True)
    STATE_FILE.write_text(json.dumps(state, indent=2, ensure_ascii=False), encoding="utf-8")

# ── Excel ──────────────────────────────────────────────────────────────────────

HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT   = Font(bold=True, color="FFFFFF")
ALT_FILL      = PatternFill("solid", fgColor="D6E4F0")

def write_excel(history: list[dict]) -> None:
    """Skriver (eller uppdaterar) Excel-filen med hela tidsserien."""
    EXCEL_FILE.parent.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(EXCEL_FILE) if EXCEL_FILE.exists() else Workbook()

    # ── Blad: Tidsserie ────────────────────────────────────────────────────────
    sheet_name = "Maskiner installerade"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.active
        ws.title = sheet_name

    headers = ["Datum", "Maskiner installerade", "Dag-för-dag förändring"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for row_idx, entry in enumerate(history, start=2):
        prev_value = history[row_idx - 3]["value"] if row_idx > 2 else None
        delta = (entry["value"] - prev_value) if prev_value is not None else None

        ws.cell(row=row_idx, column=1, value=entry["date"]).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=2, value=entry["value"]).alignment = Alignment(horizontal="center")
        delta_cell = ws.cell(row=row_idx, column=3, value=delta)
        delta_cell.alignment = Alignment(horizontal="center")

        if row_idx % 2 == 0:
            for col in range(1, 4):
                ws.cell(row=row_idx, column=col).fill = ALT_FILL

    # Kolumnbredder
    for col, width in enumerate([14, 24, 26], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    wb.save(EXCEL_FILE)
    print(f"  Excel sparad → {EXCEL_FILE}")

# ── Huvudprogram ───────────────────────────────────────────────────────────────

def main() -> None:
    today = date.today().isoformat()

    print(f"Hämtar Duroc maskinsida …")
    count = fetch_machine_count()
    print(f"  Maskiner installerade: {count}")

    state = load_state()

    # Uppdatera bara om det inte redan finns ett värde för idag
    existing_dates = {e["date"] for e in state["history"]}
    if today in existing_dates:
        print(f"  Redan registrerat för {today}, hoppar över.")
    else:
        state["history"].append({"date": today, "value": count})
        save_state(state)
        print(f"  Tillstånd sparat → {STATE_FILE}")

    write_excel(state["history"])
    print("Klar.")


if __name__ == "__main__":
    main()
