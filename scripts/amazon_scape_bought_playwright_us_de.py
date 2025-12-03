import asyncio
import random
import re
import os
from datetime import date
from pathlib import Path
from playwright.async_api import async_playwright

try:
    import openpyxl
except ImportError:
    print("Du måste installera openpyxl: pip install openpyxl")
    exit()

# ── KONFIGURATION ──────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).resolve().parent
DATA_DIR = (SCRIPT_DIR / ".." / "data").resolve()
DATA_DIR.mkdir(parents=True, exist_ok=True)

# Filnamn
XLSX_PATH = str((DATA_DIR / "fractal_scape_refine_data.xlsx").resolve())
SHEET_NAME = "Tracking"

HEADLESS = True

# --- PRODUKTER ---
PRODUCTS = [
    # --- SCAPE ---
    ("Scape Dark",        "B0D5HGK3C2"),
    ("Scape Light",       "B0D5HK6JRS"),
    # --- REFINE ---
    ("Refine Mesh Light",   "B0CSYXY8FD"),
    ("Refine Fabric Light", "B0CSYXYX39"),
    ("Refine Mesh Dark",    "B0CSYYMTT4"),
    ("Refine Fabric Dark",  "B0CSYWWRSV"),
]

# --- LÄNDER ---
COUNTRIES = [
    # code, domain, Accept-Language, gl, locale, cookie_name, cookie_value
    ("US", "amazon.com", "en-US,en;q=0.9", "US", "en-US", "i18n-prefs", "USD"),
    ("DE", "amazon.de",  "de-DE,de;q=0.9,en;q=0.6", "DE", "de-DE", "lc-acbde", "de_DE"), 
]

UAS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_0) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4 Safari/605.1.15",
]

# --- HÄR ÄR DIN HTML-STRUKTUR ---
BOUGHT_SELECTORS = [
    "#social-proofing-faceout-title-tk_bought span.a-text-bold", 
    "#social-proofing-faceout-title-tk_bought",                  
    "div.social-proofing-faceout span.a-text-bold",              
]

# ───────────────────────────────────────────────────────────────────────────

def parse_number(text: str) -> int:
    m = re.search(r"([0-9][0-9.,]*)", text)
    if not m: return 0
    return int(re.sub(r"[^\d]", "", m.group(1)) or "0")

def extract_bought_from_html(html_content: str) -> int:
    match = re.search(r">([0-9.,]+)\+?\s*bought", html_content, re.IGNORECASE)
    if match:
        return parse_number(match.group(1))
    return 0

def extract_rank_from_row_text(row_text: str) -> int:
    clean_text = re.sub(r"top\s*100", "", row_text, flags=re.IGNORECASE)
    regex = r"(?:(?:Nr\.?|#)\s*([0-9.,]+))|([0-9.,]+)\s+in\s+"
    matches = re.findall(regex, clean_text, re.IGNORECASE)
    
    candidates = []
    for m in matches:
        raw_num = m[0] if m[0] else m[1]
        clean_str = raw_num.replace(",", "").replace(".", "")
        if clean_str.isdigit():
            val = int(clean_str)
            if 0 < val < 10000000:
                candidates.append(val)
                
    if candidates:
        return min(candidates)
    return 0

# --- HÄR ÄR DEN TOTALT OMGJORDA FUNKTIONEN FÖR ATT KLARA TYSKLAND ---
async def handle_amazon_blockers(page, domain="amazon.com"):
    """
    Försöker aggressivt hitta och klicka bort cookie-banners och popups.
    Speciellt anpassad för Amazon DE overlay.
    """
    print("    -> Checking for blockers/cookies...")
    
    # Lista på alla möjliga knappar för att acceptera cookies
    cookie_selectors = [
        "#sp-cc-accept",             # Standard ID (oftast denna)
        "input[name='accept']",      # Ibland en input
        "button[name='accept']",     # Ibland en button
        "text=Akzeptieren",          # Text-baserad (Tyska)
        "text=Accept Cookies",       # Text-baserad (Engelska)
    ]

    # Försök hitta och klicka på cookie-knappen
    for selector in cookie_selectors:
        try:
            # Vi kollar om den är synlig
            if await page.locator(selector).is_visible(timeout=1000):
                print(f"    -> 🍪 Hittade cookie-knapp: {selector}. Klickar...")
                await page.click(selector)
                
                # VIKTIGT: Vänta på att overlayn faktiskt försvinner
                try:
                    await page.locator(selector).wait_for(state="hidden", timeout=3000)
                    print("    -> 🍪 Banner borta.")
                except:
                    pass
                
                # Vänta en kort stund för att sidan ska "sätta sig"
                await page.wait_for_timeout(1000)
                break # Vi har klickat, gå vidare
        except Exception:
            continue

    # Hantera "Fortsätt handla" (Upsell)
    upsells = ['text="Continue shopping"', 'text="Weiter shoppen"', 'text="Weiter einkaufen"']
    for sel in upsells:
        try:
            if await page.locator(sel).is_visible(timeout=500):
                await page.click(sel)
                await page.wait_for_timeout(500)
        except:
            pass

async def get_product_data(context, asin: str, domain: str, gl: str, name: str):
    page = await context.new_page()
    url = f"https://www.{domain}/dp/{asin}?th=1&psc=1&gl={gl}"
    print(f"  Fetching {name} ({domain})...")

    bought_val = 0
    rank_val = 0

    try:
        await page.goto(url, wait_until="domcontentloaded", timeout=30000)
        
        # Kör den nya blocker-hanteraren
        await handle_amazon_blockers(page, domain)

        # Vi hämtar HTML först nu, EFTER att vi (förhoppningsvis) klickat bort bannern
        full_html = await page.content()

        # --- 1. BOUGHT COUNT ---
        found_bought_text = ""
        for sel in BOUGHT_SELECTORS:
            try:
                el = await page.wait_for_selector(sel, state="attached", timeout=1000)
                if el:
                    found_bought_text = (await el.inner_text()).strip()
                    if found_bought_text: break
            except:
                continue
        
        if found_bought_text:
            bought_val = parse_number(found_bought_text)
            print(f"    🛒 Bought (CSS): {bought_val}+")
        else:
            bought_val = extract_bought_from_html(full_html)
            if bought_val > 0:
                print(f"    🛒 Bought (Regex): {bought_val}+")
            else:
                print(f"    🛒 Bought: 0")

        # --- 2. RANKING ---
        try:
            rank_header = page.locator("text=/Best Sellers Rank|Bestseller-Rang|Best Seller Rank/i").first
            
            if await rank_header.count() > 0:
                row_element = rank_header.locator("xpath=ancestor::tr | ancestor::li | ancestor::div[contains(@class, 'db_row')]").first
                
                if await row_element.count() > 0:
                    row_text = await row_element.inner_text()
                    rank_val = extract_rank_from_row_text(row_text)
                else:
                    row_text = await rank_header.evaluate("el => el.parentElement.innerText")
                    rank_val = extract_rank_from_row_text(row_text)

            if rank_val > 0:
                print(f"    🏆 Rank: #{rank_val}")
            else:
                body_text = await page.inner_text("body")
                rank_val = extract_rank_from_row_text(body_text)
                if rank_val > 0:
                    print(f"    🏆 Rank (Fallback): #{rank_val}")
                else:
                    print(f"    ⚠️ Rank: Hittades ej")

        except Exception as e:
            print(f"    ⚠️ Rank Error: {e}")

        # Om vi fick 0 på allt, dumpa HTML för debug
        if bought_val == 0 and rank_val == 0:
             print(f"    📸 Sparar debug-HTML för {name}...")
             debug_file = DATA_DIR / f"debug_{name.replace(' ', '_')}.html"
             debug_file.write_text(full_html, encoding="utf-8")

        return bought_val, rank_val

    except Exception as e:
        print(f"    ❌ Error: {e}")
        return 0, 0
    finally:
        await page.close()

def append_to_excel(data_dict):
    file_exists = os.path.exists(XLSX_PATH)
    
    headers = ["Date"]
    keys_order = []
    
    for country_code, *_ in COUNTRIES:
        for prod_name, _ in PRODUCTS:
            key_bought = f"{prod_name} {country_code} Bought"
            key_rank = f"{prod_name} {country_code} Rank"
            headers.append(key_bought)
            headers.append(key_rank)
            keys_order.append(key_bought)
            keys_order.append(key_rank)

    if not file_exists:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(headers)
    else:
        wb = openpyxl.load_workbook(XLSX_PATH)
        if SHEET_NAME in wb.sheetnames:
            ws = wb[SHEET_NAME]
        else:
            ws = wb.create_sheet(SHEET_NAME)
            if ws.max_row == 1: pass 

    row = [data_dict["Date"]]
    for key in keys_order:
        row.append(data_dict.get(key, 0))
    
    ws.append(row)
    wb.save(XLSX_PATH)
    print(f"💾 Data saved to {XLSX_PATH}")

async def run_once():
    today = date.today().isoformat()
    results = {"Date": today}

    async with async_playwright() as pw:
        # VIKTIGT: Sätt args för att starta maximerat, hjälper ibland med layouts
        browser = await pw.chromium.launch(headless=HEADLESS, args=["--start-maximized"])
        print(f"--- Starting Final Fix V4 ({today}) ---")

        for code, domain, accept_lang, gl, locale, ck_name, ck_value in COUNTRIES:
            print(f"\n--- Market: {code} ({domain}) ---")
            
            context = await browser.new_context(
                locale=locale,
                user_agent=random.choice(UAS),
                java_script_enabled=True,
                viewport={"width": 1920, "height": 1080}, # Större skärm minskar risken för mobil-layout
                extra_http_headers={"Accept-Language": accept_lang},
            )
            await context.add_cookies([{
                "name": ck_name, "value": ck_value,
                "domain": f".{domain}", "path": "/",
                "secure": True, "httpOnly": False
            }])

            for prod_name, asin in PRODUCTS:
                full_name = f"{prod_name} {code}"
                bought, rank = await get_product_data(context, asin, domain, gl, full_name)
                results[f"{full_name} Bought"] = bought
                results[f"{full_name} Rank"] = rank
                await asyncio.sleep(random.randint(2, 5))

            await context.close()
        await browser.close()
    append_to_excel(results)

if __name__ == "__main__":
    asyncio.run(run_once())