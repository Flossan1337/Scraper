"""
Fetch EU procurement data from TED (Tenders Electronic Daily) Search API.

For each company in the COMPANIES list, queries the TED Search API for
procurement notices mentioning that company, then writes structured results
to an Excel workbook with one sheet per company.

The Search API is fully anonymous – no API key or registration required.

API docs:  https://docs.ted.europa.eu/api/latest/search.html
Swagger:   https://api.ted.europa.eu/swagger
Query ref: https://ted.europa.eu/en/help/search-browse#expert-search
"""

from __future__ import annotations

import re
import sys
import time
import logging
from io import BytesIO
from pathlib import Path
from typing import Any, Union

import pandas as pd
import requests
from lxml import etree
from openpyxl.styles import Font

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
# Companies to track.  Each entry is either a plain string (company name)
# or a dict with:
#   display_name  – used as the Excel sheet name
#   search_terms  – list of names/brands used in the FT query
#   org_numbers   – list of Swedish/Norwegian/Danish org numbers for
#                   precise role-detection (winner / buyer matching)
#
# Org numbers are normalised internally (dashes and spaces stripped)
# before comparison with TED identifier fields.
COMPANIES: list[Union[str, dict]] = [
    "Exsitec AB",
    "EQL Pharma AB",
]

# TED API endpoint (anonymous, no auth needed)
API_URL = "https://api.ted.europa.eu/v3/notices/search"

# Fields to retrieve from the API.  Each adds to the "fields per page" budget
# (max 10 000 per page).  We keep a focused set of the most relevant fields.
API_FIELDS = [
    "publication-number",
    "notice-title",
    "notice-type",
    "publication-date",
    "dispatch-date",
    # Buyer
    "buyer-name",
    "buyer-country",
    "buyer-city",
    "buyer-identifier",
    # Winner / award
    "winner-name",
    "winner-country",
    "winner-identifier",
    "winner-selection-status",
    "contract-conclusion-date",
    # Values
    "estimated-value-lot",
    "estimated-value-cur-lot",
    "tender-value",
    "tender-value-cur",
    "total-value",
    "total-value-cur",
    # Procedure & classification
    "procedure-type",
    "contract-nature",
    "classification-cpv",
    "place-of-performance",
    # Deadlines
    "deadline-receipt-tender-date-lot",
    # Text
    "title-proc",
    "description-proc",
    # Tenderers
    "organisation-name-tenderer",
    # Links
    "links",
]

PAGE_SIZE = 100          # notices per API page (max 250)
REQUEST_DELAY = 0.5      # seconds between API calls (fair-use)
MAX_PAGES = 150          # safety limit (100 x 150 = 15 000 notices)

# Only fetch notices published on or after this date (YYYYMMDD).
# Reduces result size significantly for broad queries like Ambea.
MIN_PUBLICATION_DATE = "20210101"

# Columns to include in the Excel output (in this order).
# "Won by Company" is computed internally but intentionally excluded here;
# it is still used to apply green font formatting in the Excel writer.
OUTPUT_COLUMNS = [
    "Status",
    "Publication Date",
    "Awarded Value",
    "Buyer Name",
    "Estimated Value",
    "Winner Name",
    "Awarded Currency",
    "Estimated Currency",
]

# Paths
REPO_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = REPO_ROOT / "data"
DATA_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_FILE = DATA_DIR / "ted_procurements.xlsx"

# Logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _extract_text(value: Any, preferred_langs: tuple[str, ...] = ("eng", "swe")) -> str:
    """Extract a readable string from a TED API field value.

    TED returns multilingual text fields as ``{"swe": ["text"], "eng": ["text"]}``
    and simple fields as plain strings, lists, or numbers.
    """
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, str):
        return value
    if isinstance(value, list):
        # List of plain values (e.g. CPV codes, country codes)
        return ", ".join(str(v) for v in value)
    if isinstance(value, dict):
        # Multilingual dict - pick preferred language
        for lang in preferred_langs:
            if lang in value:
                v = value[lang]
                return ", ".join(str(i) for i in v) if isinstance(v, list) else str(v)
        # Fall back to first available language
        for lang, v in value.items():
            if isinstance(v, list):
                return ", ".join(str(i) for i in v)
            return str(v)
    return str(value)


def _extract_link(links_field: Any) -> str:
    """Return the English HTML link for a notice, or the first available."""
    if not isinstance(links_field, dict):
        return ""
    html = links_field.get("html", {})
    for lang in ("ENG", "SWE"):
        if lang in html:
            return html[lang]
    # First available
    for url in html.values():
        return url
    return ""


def _company_in_field(field_value: Any, company_name: str) -> bool:
    """Check if *company_name* appears (case-insensitive) in a field value."""
    text = _extract_text(field_value).lower()
    # Check both full name and name without corporate suffix
    name_lower = company_name.lower()
    core = re.sub(
        r"\s+(ab|as|a/s|gmbh|ltd|inc|oy|aps|bv|nv|sa|sl|srl|ag)\s*$",
        "", name_lower,
    ).strip()
    return name_lower in text or (core != name_lower and core in text)


def _parse_numeric(val: Any) -> object:
    """Try to extract a single numeric value; return None if unavailable.

    If the API returns a list of values (e.g. multiple lots), they are summed.
    The actual thousand-separator formatting is applied in the Excel writer.
    """
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return val
    if isinstance(val, str):
        try:
            return float(val)
        except (ValueError, TypeError):
            return None
    if isinstance(val, list):
        nums = []
        for v in val:
            try:
                nums.append(float(v))
            except (ValueError, TypeError):
                pass
        return sum(nums) if nums else None
    return None


def _normalize_org_num(num: str) -> str:
    """Normalise an org/company number to alphanumeric-only for comparison.

    Strips spaces and dashes so that ``556668-4345`` and ``5566684345`` are
    treated as equal.
    """
    return re.sub(r"[^0-9A-Za-z]", "", num)


def _normalize_company_config(company: Union[str, dict]) -> dict:
    """Return a unified config dict regardless of whether *company* is a
    plain string or an already-structured dict."""
    if isinstance(company, str):
        return {
            "display_name": company,
            "search_terms": [company],
            "org_numbers": set(),
        }
    org_numbers = {
        _normalize_org_num(n)
        for n in company.get("org_numbers", [])
        if n
    }
    return {
        "display_name": company["display_name"],
        "search_terms": company.get("search_terms", [company["display_name"]]),
        "org_numbers": org_numbers,
    }


def _org_in_identifier_field(field_value: Any, org_nums: set) -> bool:
    """Return True if any normalised org number in *org_nums* appears in
    *field_value* (a TED identifier field, possibly a list)."""
    if not org_nums or field_value is None:
        return False
    values = field_value if isinstance(field_value, list) else [field_value]
    for v in values:
        if _normalize_org_num(str(v)) in org_nums:
            return True
    return False


# ---------------------------------------------------------------------------
# XML lot-level parsing (eForms)
# ---------------------------------------------------------------------------

TED_XML_URL = "https://ted.europa.eu/en/notice/{pub_num}/xml"


def fetch_notice_xml(pub_num: str) -> bytes | None:
    """Download the eForms XML for a single notice. Returns bytes or None.

    Retries once on 429 (Too Many Requests) with a longer back-off.
    """
    url = TED_XML_URL.format(pub_num=pub_num)
    for attempt in range(2):
        try:
            resp = requests.get(url, timeout=30)
            if resp.status_code == 429 and attempt == 0:
                log.info("  Rate-limited on %s, waiting 3 s and retrying...", pub_num)
                time.sleep(3)
                continue
            resp.raise_for_status()
            return resp.content
        except requests.RequestException as exc:
            if attempt == 0:
                time.sleep(2)
                continue
            log.warning("  XML download failed for %s: %s", pub_num, exc)
            return None
    return None


def parse_eforms_xml(xml_bytes: bytes, company_name: str) -> list[dict]:
    """Parse an eForms XML and extract lot-level tender data for *company_name*.

    Returns a list of dicts, one per lot tender where the company participated.
    Each dict contains: lot_id, lot_title, tender_value, currency, result
    (won/lost/pending), contract_title, num_tenders (total on that lot),
    plus the framework-agreement max value if available.
    """
    tree = etree.parse(BytesIO(xml_bytes))
    root = tree.getroot()
    nsmap = root.nsmap.copy()
    # Ensure common prefixes are available
    nsmap.setdefault("cbc", "urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2")
    nsmap.setdefault("cac", "urn:oasis:names:specification:ubl:schema:xsd:CommonAggregateComponents-2")
    nsmap.setdefault("efac", "urn:oasis:names:specification:ubl:schema:xsd:CommonExtensionComponents-2")

    # Try to find efac/efbc namespaces from the document
    for prefix, uri in root.nsmap.items():
        if prefix and "efac" in prefix.lower():
            nsmap["efac"] = uri
        if prefix and "efbc" in prefix.lower():
            nsmap["efbc"] = uri

    # Check for NoticeResult — only eForms award notices have this
    nr = root.find(".//efac:NoticeResult", nsmap)
    if nr is None:
        return []

    # Procedure-level total estimated value
    proc_est_el = root.find(
        ".//cac:ProcurementProject/cac:RequestedTenderTotal/"
        "cbc:EstimatedOverallContractAmount", nsmap,
    )
    proc_estimated_value = (
        float(proc_est_el.text) if proc_est_el is not None and proc_est_el.text else None
    )
    proc_estimated_currency = (
        proc_est_el.get("currencyID", "") if proc_est_el is not None else ""
    )

    # NoticeResult total awarded value
    total_amount_el = nr.find("cbc:TotalAmount", nsmap)
    if total_amount_el is None:
        total_amount_el = nr.find("efbc:OverallMaximumFrameworkContractsAmount", nsmap)
    total_awarded_value = (
        float(total_amount_el.text) if total_amount_el is not None and total_amount_el.text else None
    )
    total_awarded_currency = (
        total_amount_el.get("currencyID", "") if total_amount_el is not None else ""
    )

    company_lower = company_name.lower()
    core = re.sub(
        r"\s+(ab|as|a/s|gmbh|ltd|inc|oy|aps|bv|nv|sa|sl|srl|ag)\s*$",
        "", company_lower,
    ).strip()

    # 1. Build org_map: ORG-ID → company name
    org_map: dict[str, str] = {}
    for org in root.findall(".//efac:Organization", nsmap):
        company_el = org.find("efac:Company", nsmap)
        if company_el is None:
            continue
        org_id_el = company_el.find("cac:PartyIdentification/cbc:ID", nsmap)
        org_name_el = company_el.find(
            "cac:PartyName/cbc:Name", nsmap
        )
        if org_id_el is not None:
            name = org_name_el.text if org_name_el is not None else ""
            org_map[org_id_el.text] = name or ""

    # 2. Build party_map: TPA-ID → set of ORG-IDs
    party_map: dict[str, set[str]] = {}
    for tp in root.findall(".//efac:TenderingParty", nsmap):
        tpa_id_el = tp.find("cbc:ID", nsmap)
        if tpa_id_el is None:
            continue
        tpa_id = tpa_id_el.text
        org_ids: set[str] = set()
        for tenderer in tp.findall("efac:Tenderer", nsmap):
            oid_el = tenderer.find("cbc:ID", nsmap)
            if oid_el is not None:
                org_ids.add(oid_el.text)
        party_map[tpa_id] = org_ids

    # 3. Build lot_map: LOT-ID → {title, start_date, end_date}
    lot_map: dict[str, dict] = {}
    for lot in root.findall(".//cac:ProcurementProjectLot", nsmap):
        lot_id_el = lot.find("cbc:ID", nsmap)
        if lot_id_el is None:
            continue
        lot_title_el = lot.find("cac:ProcurementProject/cbc:Name", nsmap)
        start_el = lot.find("cac:ProcurementProject/cac:PlannedPeriod/cbc:StartDate", nsmap)
        end_el = lot.find("cac:ProcurementProject/cac:PlannedPeriod/cbc:EndDate", nsmap)
        lot_map[lot_id_el.text] = {
            "title": lot_title_el.text if lot_title_el is not None else "",
            "start_date": (start_el.text.split("+")[0].split("Z")[0] if start_el is not None and start_el.text else ""),
            "end_date": (end_el.text.split("+")[0].split("Z")[0] if end_el is not None and end_el.text else ""),
        }

    # 4. Build tender_map from direct NoticeResult LotTender children
    tender_map: dict[str, dict] = {}
    for lt in nr.findall("efac:LotTender", nsmap):
        ten_id_el = lt.find("cbc:ID", nsmap)
        if ten_id_el is None:
            continue
        ten_id = ten_id_el.text
        lot_id_el = lt.find("efac:TenderLot/cbc:ID", nsmap)
        party_id_el = lt.find("efac:TenderingParty/cbc:ID", nsmap)
        amount_el = lt.find("cac:LegalMonetaryTotal/cbc:PayableAmount", nsmap)
        rank_el = lt.find("efac:TenderRank/cbc:RankCode", nsmap) or \
                  lt.find("efac:SubcontractingTerm/efbc:TermCode", nsmap)

        tender_map[ten_id] = {
            "lot_id": lot_id_el.text if lot_id_el is not None else "",
            "party_id": party_id_el.text if party_id_el is not None else "",
            "value": float(amount_el.text) if amount_el is not None and amount_el.text else None,
            "currency": amount_el.get("currencyID", "") if amount_el is not None else "",
        }

    # 5. Build lot_results: lot_id → {winner_tender_ids, result_code, fa_max}
    lot_results: dict[str, dict] = {}
    for lr in nr.findall("efac:LotResult", nsmap):
        lot_id_el = lr.find("efac:TenderLot/cbc:ID", nsmap)
        if lot_id_el is None:
            continue
        lot_id = lot_id_el.text
        result_code_el = lr.find("cbc:TenderResultCode", nsmap)
        winner_ids: set[str] = set()
        for wlt in lr.findall("efac:LotTender/cbc:ID", nsmap):
            winner_ids.add(wlt.text)
        # Framework agreement max value
        fa_max_el = lr.find(
            "efac:FrameworkAgreementValues/cbc:MaximumValueAmount", nsmap
        )
        entry = lot_results.setdefault(lot_id, {
            "winner_tender_ids": set(),
            "result_code": "",
            "fa_max_value": None,
            "fa_max_currency": "",
        })
        entry["winner_tender_ids"].update(winner_ids)
        if result_code_el is not None and result_code_el.text:
            entry["result_code"] = result_code_el.text
        if fa_max_el is not None and fa_max_el.text:
            entry["fa_max_value"] = float(fa_max_el.text)
            entry["fa_max_currency"] = fa_max_el.get("currencyID", "")

    # 6. Build contract_map: TEN-ID → contract title
    contract_map: dict[str, str] = {}
    for sc in nr.findall("efac:SettledContract", nsmap):
        title_el = sc.find("cbc:ID", nsmap)
        title_text_el = sc.find("cac:Contract/cbc:Title", nsmap)
        title_text = title_text_el.text if title_text_el is not None else ""
        for lt_ref in sc.findall("efac:LotTender/cbc:ID", nsmap):
            contract_map[lt_ref.text] = title_text or ""

    # 7. Count tenders per lot
    tenders_per_lot: dict[str, int] = {}
    for t_info in tender_map.values():
        lid = t_info["lot_id"]
        tenders_per_lot[lid] = tenders_per_lot.get(lid, 0) + 1

    # 8. Find our company's org IDs
    our_org_ids: set[str] = set()
    for org_id, name in org_map.items():
        name_l = name.lower()
        if company_lower in name_l or (core != company_lower and core in name_l):
            our_org_ids.add(org_id)

    if not our_org_ids:
        return []

    # 9. Find our party IDs
    our_party_ids: set[str] = set()
    for tpa_id, org_ids in party_map.items():
        if our_org_ids & org_ids:
            our_party_ids.add(tpa_id)

    # 10. Filter tenders for our company and count wins
    our_tenders: list[tuple[str, dict]] = []
    for ten_id, t in tender_map.items():
        if t["party_id"] in our_party_ids:
            our_tenders.append((ten_id, t))

    total_lots = len(lot_map)
    eql_lots_won = 0
    for ten_id, t in our_tenders:
        lot_id = t["lot_id"]
        lr = lot_results.get(lot_id, {})
        if ten_id in lr.get("winner_tender_ids", set()):
            eql_lots_won += 1

    results: list[dict] = []
    for ten_id, t in our_tenders:
        lot_id = t["lot_id"]
        lr = lot_results.get(lot_id, {})
        is_winner = ten_id in lr.get("winner_tender_ids", set())

        lot_info = lot_map.get(lot_id, {})
        start_date = lot_info.get("start_date", "") if isinstance(lot_info, dict) else ""
        end_date = lot_info.get("end_date", "") if isinstance(lot_info, dict) else ""
        lot_title = lot_info.get("title", "") if isinstance(lot_info, dict) else ""

        # Calculate duration in years
        duration_years = None
        if start_date and end_date:
            try:
                from datetime import date as _date
                sd = _date.fromisoformat(start_date)
                ed = _date.fromisoformat(end_date)
                duration_years = round((ed - sd).days / 365.25, 1)
            except ValueError:
                pass

        results.append({
            "lot_id": lot_id,
            "lot_title": lot_title,
            "tender_value": t["value"],
            "currency": t["currency"],
            "result": "Won" if is_winner else "Lost",
            "contract_title": contract_map.get(ten_id, ""),
            "num_tenders_on_lot": tenders_per_lot.get(lot_id, 0),
            "fa_max_value": lr.get("fa_max_value"),
            "fa_max_currency": lr.get("fa_max_currency", ""),
            "start_date": start_date,
            "end_date": end_date,
            "duration_years": duration_years,
            "total_lots": total_lots,
            "eql_lots_won": eql_lots_won,
            "proc_estimated_value": proc_estimated_value,
            "proc_estimated_currency": proc_estimated_currency,
            "total_awarded_value": total_awarded_value,
            "total_awarded_currency": total_awarded_currency,
        })

    return results


def fetch_lot_details(
    notices: list[dict],
    config: dict,
) -> pd.DataFrame:
    """For each notice, download XML and parse lot-level data.

    Returns a DataFrame with one row per lot tender where the company
    participated, enriched with notice-level metadata.
    """
    company = config["display_name"]
    all_rows: list[dict] = []
    total = len(notices)

    for i, n in enumerate(notices, 1):
        pub_num = _extract_text(n.get("publication-number"))
        if not pub_num:
            continue

        log.info("  XML %d/%d: %s", i, total, pub_num)
        xml_bytes = fetch_notice_xml(pub_num)
        if xml_bytes is None:
            continue

        lot_rows = parse_eforms_xml(xml_bytes, company)
        if not lot_rows:
            continue

        # Enrich with notice-level info
        pub_date = _extract_text(n.get("publication-date", "")).split("+")[0]
        buyer = _extract_text(n.get("buyer-name"))
        notice_type_raw = _extract_text(n.get("notice-type"))
        notice_title = (
            _extract_text(n.get("notice-title"))
            or _extract_text(n.get("title-proc"))
        )
        for row in lot_rows:
            row["publication_number"] = pub_num
            row["publication_date"] = pub_date
            row["buyer_name"] = buyer
            row["notice_type"] = NOTICE_TYPE_LABELS.get(notice_type_raw, notice_type_raw)
            row["notice_title"] = notice_title

        all_rows.extend(lot_rows)
        time.sleep(0.5)  # polite delay to avoid 429 rate-limiting

    if not all_rows:
        return pd.DataFrame()

    df = pd.DataFrame(all_rows)
    # Reorder columns for readability
    col_order = [
        "publication_date", "buyer_name", "lot_id", "lot_title",
        "result", "tender_value", "currency",
        "fa_max_value", "fa_max_currency",
        "proc_estimated_value", "proc_estimated_currency",
        "total_awarded_value", "total_awarded_currency",
        "total_lots", "eql_lots_won",
        "start_date", "end_date", "duration_years",
        "num_tenders_on_lot", "contract_title",
        "notice_type", "notice_title", "publication_number",
    ]
    col_order = [c for c in col_order if c in df.columns]
    df = df[col_order]
    df.sort_values("publication_date", ascending=True, inplace=True)
    df.reset_index(drop=True, inplace=True)
    return df


def _build_query(config: dict) -> str:
    """Build a TED expert-search query from *config*.

    Each search term becomes its own ``FT ~ "..."`` clause joined with OR.
    For a single term we also add a suffix-stripped variant inside parentheses
    (e.g. ``FT ~ ("Exsitec AB" OR "Exsitec")``).
    For multiple terms we chain separate FT clauses
    (e.g. ``FT ~ "Ambea" OR FT ~ "Nytida" OR ...``) because TED's parser
    does not reliably support large parenthesised OR lists inside a single
    FT ~ operator.

    Corporate suffixes are automatically stripped to broaden recall.
    """
    terms = config["search_terms"]

    if len(terms) == 1:
        # Single term: FT ~ ("Name" OR "CoreName") variant
        name = terms[0]
        core = re.sub(
            r"\s+(AB|AS|A/S|GmbH|Ltd|Inc|Oy|ApS|BV|NV|SA|SL|SRL|AG)\s*$",
            "", name, flags=re.IGNORECASE,
        ).strip()
        if core and core.lower() != name.lower():
            return f'FT ~ ("{name}" OR "{core}") AND PD >= {MIN_PUBLICATION_DATE}'
        return f'FT ~ "{name}" AND PD >= {MIN_PUBLICATION_DATE}'

    # Multiple terms: chain as separate FT clauses, wrapped in parens for the date AND
    clauses = [f'FT ~ "{t}"' for t in terms]
    ft_part = f'({" OR ".join(clauses)})'
    return f'{ft_part} AND PD >= {MIN_PUBLICATION_DATE}'


# ---------------------------------------------------------------------------
# API client
# ---------------------------------------------------------------------------

def search_notices(query: str, scope: str = "ALL") -> list[dict]:
    """Run a paginated search against the TED Search API.

    Parameters
    ----------
    query : str
        Expert-search query string.
    scope : str
        ``"ALL"`` for the full archive or ``"ACTIVE"`` for currently active.

    Returns
    -------
    list[dict]
        Raw notice dicts as returned by the API.
    """
    all_notices: list[dict] = []
    page = 1

    while page <= MAX_PAGES:
        payload = {
            "query": query,
            "page": page,
            "limit": PAGE_SIZE,
            "scope": scope,
            "fields": API_FIELDS,
            "onlyLatestVersions": True,
        }
        log.info("  Fetching page %d (scope=%s) ...", page, scope)
        for attempt in range(3):
            try:
                resp = requests.post(
                    API_URL,
                    json=payload,
                    headers={"Content-Type": "application/json"},
                    timeout=60,
                )
                resp.raise_for_status()
                break
            except requests.RequestException as exc:
                if attempt < 2:
                    log.warning("  Attempt %d failed (%s), retrying...", attempt + 1, exc)
                    time.sleep(2)
                else:
                    log.error("  API request failed after 3 attempts: %s", exc)
                    return all_notices

        data = resp.json()
        notices = data.get("notices", [])
        total = data.get("totalNoticeCount", 0)

        if page == 1:
            log.info("  Total matching notices: %s", total)

        if not notices:
            break

        all_notices.extend(notices)

        # Are we done?
        if isinstance(total, int) and len(all_notices) >= total:
            break

        page += 1
        time.sleep(REQUEST_DELAY)

    return all_notices


# ---------------------------------------------------------------------------
# Data transformation
# ---------------------------------------------------------------------------

NOTICE_TYPE_LABELS = {
    # Planning
    "pin-buyer": "Prior Information Notice - Buyer Profile",
    "pin-only": "Prior Information Notice",
    "pin-cfc-standard": "Prior Information - Call for Competition",
    "pin-cfc-social": "Prior Information - Call for Competition (Social)",
    "pin-tran": "Prior Information - Public Transport",
    # Competition
    "cn-standard": "Contract Notice",
    "cn-social": "Contract Notice - Social",
    "cn-desg": "Design Contest Notice",
    "qu-sy": "Qualification System Notice",
    "subco": "Subcontracting Notice",
    # Result
    "can-standard": "Contract Award Notice",
    "can-social": "Contract Award Notice - Social",
    "can-desg": "Design Contest Result",
    "can-tran": "Contract Award - Public Transport",
    "can-modif": "Contract Modification Notice",
    "veat": "Voluntary Ex-Ante Transparency Notice",
}

PROCEDURE_TYPE_LABELS = {
    "open": "Open",
    "restricted": "Restricted",
    "neg-w-call": "Negotiated with Call",
    "neg-wo-call": "Negotiated without Call",
    "comp-dial": "Competitive Dialogue",
    "comp-neg": "Competitive Negotiation",
    "innovation": "Innovation Partnership",
    "oth-single": "Other (single stage)",
    "oth-multi": "Other (multi stage)",
}

# Notice types that indicate an ongoing/open procurement opportunity
_ACTIVE_NOTICE_TYPES = {
    "pin-buyer", "pin-only", "pin-cfc-standard", "pin-cfc-social", "pin-tran",
    "cn-standard", "cn-social", "cn-desg", "qu-sy", "subco",
}


def notices_to_dataframe(
    notices: list[dict],
    config: dict,
    active_pub_nums: set[str],
) -> pd.DataFrame:
    """Convert raw TED notices into a flat DataFrame.

    Parameters
    ----------
    notices : list[dict]
        Raw notice dicts from the API (full / ALL scope).
    config : dict
        Normalised company config (display_name, search_terms, org_numbers).
    active_pub_nums : set[str]
        Publication numbers that appeared in the ACTIVE scope query.
    """
    company = config["display_name"]
    org_numbers = config["org_numbers"]   # normalised set
    rows: list[dict] = []

    for n in notices:
        pub_num = _extract_text(n.get("publication-number"))
        notice_type_raw = _extract_text(n.get("notice-type"))
        winner_raw = _extract_text(n.get("winner-name"))
        tenderer_raw = _extract_text(n.get("organisation-name-tenderer"))

        # Determine company role.
        # When org numbers are provided, identifier-field matching takes
        # precedence over (and is added to) name-based matching.
        is_winner = (
            _company_in_field(n.get("winner-name"), company)
            or _org_in_identifier_field(n.get("winner-identifier"), org_numbers)
        )
        is_tenderer = _company_in_field(n.get("organisation-name-tenderer"), company)
        is_buyer = (
            _company_in_field(n.get("buyer-name"), company)
            or _org_in_identifier_field(n.get("buyer-identifier"), org_numbers)
        )

        if is_winner:
            role = "Winner"
        elif is_tenderer:
            role = "Tenderer"
        elif is_buyer:
            role = "Buyer"
        else:
            role = "Mentioned"

        # Status: use ACTIVE-scope membership first, then fall back to notice type
        if pub_num in active_pub_nums:
            status = "Active"
        elif notice_type_raw in _ACTIVE_NOTICE_TYPES:
            status = "Active"
        else:
            status = "Historical"

        # Best available value
        est_val = n.get("estimated-value-lot")
        est_cur = _extract_text(n.get("estimated-value-cur-lot"))
        award_val = n.get("tender-value") or n.get("total-value")
        award_cur = _extract_text(
            n.get("tender-value-cur") or n.get("total-value-cur")
        )

        rows.append({
            "Status": status,
            "Publication Date": _extract_text(n.get("publication-date", "")).split("+")[0],
            "Won by Company": "Yes" if is_winner else ("No" if winner_raw else ""),
            "Awarded Value": _parse_numeric(award_val),
            "Buyer Name": _extract_text(n.get("buyer-name")),
            "Estimated Value": _parse_numeric(est_val),
            "Winner Name": winner_raw,
            "Notice Type": NOTICE_TYPE_LABELS.get(notice_type_raw, notice_type_raw),
            "Title": (
                _extract_text(n.get("notice-title"))
                or _extract_text(n.get("title-proc"))
            ),
            "Description": _extract_text(n.get("description-proc"))[:500],
            "Publication Number": pub_num,
            "Awarded Currency": award_cur,
            "Estimated Currency": est_cur,
            "Tenderers": tenderer_raw,
            "Procedure Type": PROCEDURE_TYPE_LABELS.get(
                _extract_text(n.get("procedure-type")),
                _extract_text(n.get("procedure-type")),
            ),
        })

    df = pd.DataFrame(rows)
    if df.empty:
        return df

    # For companies tracked via org numbers:
    # keep only rows where the group won the contract OR the procurement is
    # still active.  Historical losses (Tenderer / Mentioned) are dropped.
    if config["org_numbers"]:
        df = df[(df["Status"] == "Active") | (df["Won by Company"] == "Yes")].copy()

    # Sort: oldest first so new rows append at the bottom; Active at the end
    df["_sort_status"] = df["Status"].map({"Historical": 0, "Active": 1})
    df["_sort_date"] = pd.to_datetime(df["Publication Date"], errors="coerce")
    df.sort_values(
        ["_sort_status", "_sort_date"],
        ascending=[True, True],
        inplace=True,
    )
    df.drop(columns=["_sort_status", "_sort_date"], inplace=True)
    df.reset_index(drop=True, inplace=True)

    return df


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

def _sanitise_sheet_name(name: str) -> str:
    """Excel sheet names must be <= 31 chars and cannot contain certain chars."""
    name = re.sub(r"[\\/*?\[\]:]", " ", name).strip()
    return name[:31]


def write_excel(
    company_frames: dict[str, pd.DataFrame],
    detail_frames: dict[str, pd.DataFrame] | None = None,
) -> None:
    """Write one sheet per company to the output Excel file.

    If *detail_frames* is provided, an additional sheet per company is added
    with lot-level detail data from XML parsing.
    """
    if detail_frames is None:
        detail_frames = {}

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        for company, df in company_frames.items():
            sheet = _sanitise_sheet_name(company)
            if df.empty:
                # Write a placeholder so the sheet still exists
                pd.DataFrame(
                    {"Info": [f"No procurement notices found for {company}"]}
                ).to_excel(writer, sheet_name=sheet, index=False)
            else:
                # Identify won rows using the internal column BEFORE subsetting
                won_row_indices: set[int] = set()
                if "Won by Company" in df.columns:
                    won_row_indices = set(
                        df.index[df["Won by Company"] == "Yes"].tolist()
                    )

                # Write only the defined output columns
                out_cols = [c for c in OUTPUT_COLUMNS if c in df.columns]
                df_out = df[out_cols]
                df_out.to_excel(writer, sheet_name=sheet, index=False)

                ws = writer.sheets[sheet]

                # Apply thousand-separator number format to value columns
                for col_name in ("Awarded Value", "Estimated Value"):
                    if col_name in df_out.columns:
                        col_idx = list(df_out.columns).index(col_name) + 1  # 1-based
                        for row_idx in range(2, len(df_out) + 2):  # skip header
                            cell = ws.cell(row=row_idx, column=col_idx)
                            if cell.value is not None:
                                cell.number_format = '#,##0'

                # Highlight rows where the company won with dark green text
                green_font = Font(color="006100")
                num_cols = len(df_out.columns)
                for df_idx in won_row_indices:
                    row_idx = df_idx + 2  # +1 for 1-based index, +1 for header
                    for col_idx in range(1, num_cols + 1):
                        ws.cell(row=row_idx, column=col_idx).font = green_font

                # Auto-size columns (approximate)
                for col_idx, col_name in enumerate(df_out.columns, 1):
                    max_len = max(
                        len(str(col_name)),
                        df_out[col_name].astype(str).str.len().max(),
                    )
                    # Cap at 60 characters width
                    adjusted = min(int(max_len) + 2, 60)
                    col_letter = ws.cell(1, col_idx).column_letter
                    ws.column_dimensions[col_letter].width = adjusted

        # --- Detail sheets (lot-level XML data) ---
        for company, df_detail in detail_frames.items():
            if df_detail.empty:
                continue
            sheet = _sanitise_sheet_name(company + " Detail")
            df_detail.to_excel(writer, sheet_name=sheet, index=False)
            ws = writer.sheets[sheet]

            # Format value columns with thousand separators
            for col_name in ("tender_value", "fa_max_value", "proc_estimated_value", "total_awarded_value"):
                if col_name in df_detail.columns:
                    col_idx = list(df_detail.columns).index(col_name) + 1
                    for row_idx in range(2, len(df_detail) + 2):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if cell.value is not None:
                            cell.number_format = '#,##0'

            # Highlight won rows in green
            green_font = Font(color="006100")
            if "result" in df_detail.columns:
                res_col = list(df_detail.columns).index("result")
                num_cols = len(df_detail.columns)
                for row_idx in range(2, len(df_detail) + 2):
                    if ws.cell(row=row_idx, column=res_col + 1).value == "Won":
                        for col_idx in range(1, num_cols + 1):
                            ws.cell(row=row_idx, column=col_idx).font = green_font

            # Auto-size columns
            for col_idx, col_name in enumerate(df_detail.columns, 1):
                max_len = max(
                    len(str(col_name)),
                    df_detail[col_name].astype(str).str.len().max(),
                )
                adjusted = min(int(max_len) + 2, 60)
                col_letter = ws.cell(1, col_idx).column_letter
                ws.column_dimensions[col_letter].width = adjusted

    log.info("Wrote %s", OUTPUT_FILE)


def _build_winner_id_queries(org_numbers: set, batch_size: int = 25) -> list[str]:
    """Build one or more ``winner-identifier IN (...)`` queries for *org_numbers*.

    Splits into batches to stay well within TED's query-length limits.
    Each query is already suffixed with the MIN_PUBLICATION_DATE filter.
    """
    nums = list(org_numbers)
    queries = []
    for i in range(0, len(nums), batch_size):
        batch = nums[i : i + batch_size]
        id_list = " ".join(batch)
        queries.append(
            f"winner-identifier IN ({id_list}) AND PD >= {MIN_PUBLICATION_DATE}"
        )
    return queries


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    log.info("=" * 60)
    log.info("TED Procurement Data Fetcher")
    log.info("=" * 60)

    company_frames: dict[str, pd.DataFrame] = {}
    company_notices: dict[str, tuple[list[dict], dict]] = {}  # raw notices + config

    for raw_company in COMPANIES:
        config = _normalize_company_config(raw_company)
        display_name = config["display_name"]
        log.info("Processing: %s", display_name)

        if config["org_numbers"]:
            # --- Identifier-based strategy (companies with org numbers) ---
            # Historical: query ONLY for notices where a subsidiary actually won.
            # This avoids fetching tens of thousands of false-positive FT matches.
            log.info(
                "  Using %d org numbers; querying winner-identifier directly",
                len(config["org_numbers"]),
            )
            id_queries = _build_winner_id_queries(config["org_numbers"])
            won_notices: list[dict] = []
            for q in id_queries:
                log.info("  Winner-ID query: %s", q)
                won_notices.extend(search_notices(q, scope="ALL"))
            log.info("  Retrieved %d won notices", len(won_notices))

            # Active: use FT query so we catch open tenders before a winner
            # is assigned (winner-identifier not yet populated).
            ft_query = _build_query(config)
            log.info("  FT query (active scope): %s", ft_query)
            active_notices = search_notices(ft_query, scope="ACTIVE")
            active_pub_nums = {
                _extract_text(n.get("publication-number"))
                for n in active_notices
            }
            log.info("  Active procurements found: %d", len(active_pub_nums))

            # Merge: active notices + won notices (deduplicate by pub number)
            seen: set[str] = set()
            all_notices: list[dict] = []
            for n in active_notices + won_notices:
                pn = _extract_text(n.get("publication-number"))
                if pn not in seen:
                    seen.add(pn)
                    all_notices.append(n)

        else:
            # --- Name-based strategy (e.g. Exsitec AB) ---
            query = _build_query(config)
            log.info("  Expert query: %s", query)

            all_notices = search_notices(query, scope="ALL")
            log.info("  Retrieved %d notices total", len(all_notices))

            active_raw = search_notices(query, scope="ACTIVE")
            active_pub_nums = {
                _extract_text(n.get("publication-number"))
                for n in active_raw
            }
            log.info("  Of which %d are currently active", len(active_pub_nums))

        # Build DataFrame
        df = notices_to_dataframe(all_notices, config, active_pub_nums)
        company_frames[display_name] = df
        company_notices[display_name] = (all_notices, config)
        log.info("  Done - %d rows for sheet '%s'", len(df), display_name)

    # --- Lot-level detail via XML parsing (only for selected companies) ---
    DETAIL_COMPANIES = {"EQL Pharma AB"}
    detail_frames: dict[str, pd.DataFrame] = {}
    for display_name, (raw_notices, config) in company_notices.items():
        if not raw_notices or display_name not in DETAIL_COMPANIES:
            continue
        log.info("Fetching XML lot details for: %s", display_name)
        df_detail = fetch_lot_details(raw_notices, config)
        if not df_detail.empty:
            detail_frames[display_name] = df_detail
            log.info("  %d lot-level rows for '%s'", len(df_detail), display_name)
        else:
            log.info("  No lot-level data found for '%s'", display_name)

    write_excel(company_frames, detail_frames)
    log.info("All done!")


if __name__ == "__main__":
    main()
