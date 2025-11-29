import asyncio
import re
import os
from datetime import date
from pathlib import Path
from playwright.async_api import async_playwright

# Behövs för Excel
try:
    import openpyxl
except ImportError:
    print("Du måste installera openpyxl: pip install openpyxl")
    exit()

# ── KONFIGURATION ──────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).resolve().parent
DATA_DIR = (SCRIPT_DIR / ".." / "data").resolve()
DATA_DIR.mkdir(parents=True, exist_ok=True)
XLSX_PATH = str((DATA_DIR / "fractal_rank_tracking.xlsx").resolve())
SHEET_NAME = "Rankings_Pro"

TARGETS = [
    # --- USA (amazon.com) ---
    {
        "name": "Scape US",
        "asin": "B0D5HK6JRS", 
        "domain": "amazon.com",
        "locale": "en-US",
        "keywords": ["Best Sellers Rank"]
    },
    {
        "name": "Refine US",
        "asin": "B0CSYWWRSV",
        "domain": "amazon.com",
        "locale": "en-US",
        "keywords": ["Best Sellers Rank"]
    },
    # --- TYSKLAND (amazon.de) ---
    {
        "name": "Scape DE",
        "asin": "B0D5HK6JRS",
        "domain": "amazon.de",
        "locale": "de-DE",
        # Vi letar efter både tysk och engelsk rubrik
        "keywords": ["Bestseller-Rang", "Best Sellers Rank"] 
    },
    {
        "name": "Refine DE",
        "asin": "B0CSYWWRSV",
        "domain": "amazon.de",
        "locale": "de-DE",
        "keywords": ["Bestseller-Rang", "Best Sellers Rank"]
    },
]

HEADLESS = True 

# ───────────────────────────────────────────────────────────────────────────

def extract_rank_v6(text: str, keywords: list) -> int:
    """
    ULTRA-ROBUST LOGIK:
    1. Hitta sektionen.
    2. RADERA texten "Top 100" helt och hållet (för att undvika falska träffar).
    3. Leta efter siffror som antingen har prefix (Nr/#) eller suffix (in).
    4. Filtrera bort årtal.
    """
    found_keyword = False
    relevant_chunk = ""

    # 1. Hitta var sektionen börjar
    for kw in keywords:
        idx = text.lower().find(kw.lower())
        if idx != -1:
            start_pos = idx + len(kw)
            # Vi tar en rejäl bit text
            relevant_chunk = text[start_pos : start_pos + 1500]
            found_keyword = True
            break
    
    if not found_keyword:
        return 0

    # 2. STÄDA TEXTEN (Detta är nyckeln!)
    # Vi tar bort fraser som "Top 100", "Top 1000" etc. så att siffran 100 försvinner.
    # Vi tar också bort årtal för säkerhets skull.
    clean_chunk = re.sub(r"top\s*100", "", relevant_chunk, flags=re.IGNORECASE)
    clean_chunk = re.sub(r"202[0-9]", "", clean_chunk) # Tar bort 2024, 2025 etc.

    candidates = []

    # 3. REGEX SOM TÄCKER ALLA FALL
    # Grupp 1: Prefix-metoden (Nr. 31 eller #31)
    # Grupp 2: Suffix-metoden (31 in Computer...)
    
    regex_pattern = r"(?:(?:Nr\.?|#)\s*([0-9.,]+))|([0-9.,]+)\s+in\s+"
    
    matches = re.findall(regex_pattern, clean_chunk, re.IGNORECASE)

    for m in matches:
        # m är en tuple, t.ex ('31', '') eller ('', '5.233')
        # Vi tar den som inte är tom
        raw_num = m[0] if m[0] else m[1]
        
        # Rensa
        clean_str = raw_num.replace(",", "").replace(".", "")
        
        if clean_str.isdigit():
            val = int(clean_str)
            # Extra säkerhetsfilter
            if val > 0 and val < 10000000:
                candidates.append(val)

    if candidates:
        # Nu när vi har städat bort "Top 100" och årtal är det säkert att ta lägsta
        lowest = min(candidates)
        print(f"      🔍 Hittade kandidater: {candidates} -> Valde: {lowest}")
        return lowest
        
    return 0

async def handle_blockers(page):
    try:
        if await page.locator("#sp-cc-accept").is_visible(timeout=2000):
            await page.click("#sp-cc-accept")
            await page.wait_for_timeout(500)
    except:
        pass

    upsells = ['text="Continue shopping"', 'text="Weiter shoppen"', 'text="Weiter einkaufen"']
    for sel in upsells:
        try:
            if await page.locator(sel).is_visible(timeout=500):
                await page.click(sel)
                await page.wait_for_timeout(500)
        except:
            pass

async def get_rank(context, item) -> int:
    page = await context.new_page()
    url = f"https://www.{item['domain']}/dp/{item['asin']}?th=1&psc=1"
    print(f"  Fetching {item['name']} ({item['domain']})...")

    try:
        await page.goto(url, wait_until="domcontentloaded", timeout=30000)
        await handle_blockers(page)

        text_content = ""
        selectors = [
            "#prodDetails", 
            "#detailBullets_feature_div", 
            "#productDescription", 
            "table", 
            "ul",
            "div#centerCol", # Huvudkolumnen
            "div#productDetails_db_sections"
        ]

        for sel in selectors:
            if await page.locator(sel).count() > 0:
                elements = await page.locator(sel).all_inner_texts()
                for txt in elements:
                    text_content += txt + "\n"
        
        if len(text_content) < 500:
            text_content = await page.inner_text("body")

        rank = extract_rank_v6(text_content, item['keywords'])

        if rank > 0:
            print(f"    ✅ Rank: #{rank}")
        else:
            print(f"    ⚠️ Kunde inte hitta rank för {item['name']}")
            # Debug: Om det fortfarande strular, sparar vi texten för analys
            # with open(f"debug_{item['name']}.txt", "w", encoding="utf-8") as f:
            #     f.write(text_content)
        
        return rank

    except Exception as e:
        print(f"    ❌ Error: {e}")
        return 0
    finally:
        await page.close()

def append_to_excel(data_dict):
    file_exists = os.path.exists(XLSX_PATH)
    headers = ["Date"] + [t["name"] for t in TARGETS]

    if not file_exists:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(headers)
    else:
        wb = openpyxl.load_workbook(XLSX_PATH)
        if SHEET_NAME in wb.sheetnames:
            ws = wb[SHEET_NAME]
        else:
            ws = wb.create_sheet(SHEET_NAME)
            ws.append(headers)

    row = [data_dict["Date"]]
    for t in TARGETS:
        row.append(data_dict.get(t["name"], 0))
    
    ws.append(row)
    wb.save(XLSX_PATH)
    print(f"💾 Data saved to {XLSX_PATH}")

async def run():
    today = date.today().isoformat()
    results = {"Date": today}

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=HEADLESS)
        
        print(f"--- Starting Final V6 (Noise Canceller) ({today}) ---")

        for item in TARGETS:
            context = await browser.new_context(
                locale=item['locale'],
                viewport={"width": 1920, "height": 1080},
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
            )
            
            rank = await get_rank(context, item)
            results[item['name']] = rank
            await context.close()
        
        await browser.close()

    append_to_excel(results)

if __name__ == "__main__":
    asyncio.run(run())