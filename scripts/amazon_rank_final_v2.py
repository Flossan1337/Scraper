import asyncio
import re
import os
from datetime import date
from pathlib import Path
from playwright.async_api import async_playwright

# Behövs för Excel
try:
    import openpyxl
except ImportError:
    print("Du måste installera openpyxl: pip install openpyxl")
    exit()

# ── KONFIGURATION ──────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).resolve().parent
DATA_DIR = (SCRIPT_DIR / ".." / "data").resolve()
DATA_DIR.mkdir(parents=True, exist_ok=True)
XLSX_PATH = str((DATA_DIR / "fractal_rank_tracking.xlsx").resolve())
SHEET_NAME = "Rankings_Pro"

TARGETS = [
    # --- USA (amazon.com) ---
    {
        "name": "Scape US",
        "asin": "B0D5HK6JRS", 
        "domain": "amazon.com",
        "locale": "en-US",
        "keywords": ["Best Sellers Rank"]
    },
    {
        "name": "Refine US",
        "asin": "B0CSYWWRSV",
        "domain": "amazon.com",
        "locale": "en-US",
        "keywords": ["Best Sellers Rank"]
    },
    # --- TYSKLAND (amazon.de) ---
    {
        "name": "Scape DE",
        "asin": "B0D5HK6JRS",
        "domain": "amazon.de",
        "locale": "de-DE",
        # NYTT: Vi lägger till den engelska termen också, eftersom din bild visade engelska på DE-sidan!
        "keywords": ["Amazon Bestseller-Rang", "Bestseller-Rang", "Best Sellers Rank"] 
    },
    {
        "name": "Refine DE",
        "asin": "B0CSYWWRSV",
        "domain": "amazon.de",
        "locale": "de-DE",
        "keywords": ["Amazon Bestseller-Rang", "Bestseller-Rang", "Best Sellers Rank"]
    },
]

HEADLESS = True 

# ───────────────────────────────────────────────────────────────────────────

def extract_lowest_rank(text: str, keywords: list) -> int:
    """
    Hittar lägsta rankingen.
    Uppdaterad för att klara siffror UTAN prefix (t.ex "105,199" istället för "#105,199").
    """
    found_keyword = False
    relevant_chunk = ""

    # 1. Hitta var sektionen börjar (t.ex "Best Sellers Rank")
    for kw in keywords:
        idx = text.lower().find(kw.lower())
        if idx != -1:
            # Ta texten från nyckelordet och en bit framåt (500 tecken räcker oftast)
            # Vi plussar på längden av ordet så vi börjar söka EFTER "Best Sellers Rank"
            start_pos = idx + len(kw)
            relevant_chunk = text[start_pos : start_pos + 500]
            found_keyword = True
            break
    
    if not found_keyword:
        return 0

    # 2. Regex som hittar siffror
    # Förklaring:
    # (?:Nr\.?|#)?   -> Icke-fångande grupp. Matchar "Nr." eller "#". Frågetecknet på slutet gör hela denna VALFRI.
    # \s* -> Eventuella mellanslag
    # ([0-9]{1,3}(?:[.,][0-9]{3})*|[0-9]+) -> Matchar tal som 105,199 eller 105.199 eller 145
    matches = re.findall(r"(?:Nr\.?|#)?\s*([0-9]{1,3}(?:[.,][0-9]{3})*|[0-9]+)", relevant_chunk)
    
    clean_integers = []
    for m in matches:
        # Rensa bort punkt och komma
        clean_str = m.replace(",", "").replace(".", "")
        
        if clean_str.isdigit():
            val = int(clean_str)
            
            # 3. SÄKERHETSFILTER:
            # Om vi hittar en väldigt låg siffra (typ 4) utan prefix, är det risk att det är betyget (4.1 stjärnor)
            # och inte rankingen, eftersom vi nu tillåter siffror utan '#'.
            # Men eftersom vi söker EFTER "Best Sellers Rank" är risken liten.
            # Vi lägger ändå in en spärr: Vi antar att en rank sällan är exakt samma som ett typiskt betyg (1-5) 
            # om den kommer precis i början av strängen utan kontext.
            # Men för enkelhetens skull: Min() funktionen löser oftast detta då rank 145 > betyg 4.
            # Vänta... Min() väljer det LÄGSTA. Om den hittar "4" (från 4.1 stars) och "145" (rank), väljer den 4.
            # Det är FEL.
            
            # Lösning: "Best Sellers Rank" brukar inte följas av betyget. Betyget står ovanför.
            # Vi litar på att relevant_chunk inte innehåller betyget eftersom det står tidigare i HTML-koden.
            
            clean_integers.append(val)
    
    if clean_integers:
        # Välj lägsta siffran (Bästa ranking)
        lowest = min(clean_integers)
        print(f"      🔍 Hittade kandidater i texten: {clean_integers} -> Valde: {lowest}")
        return lowest
        
    return 0

async def handle_blockers(page):
    """Klickar bort Cookies & Upsells"""
    try:
        if await page.locator("#sp-cc-accept").is_visible(timeout=2000):
            await page.click("#sp-cc-accept")
            await page.wait_for_timeout(500)
    except:
        pass

    upsells = ['text="Continue shopping"', 'text="Weiter shoppen"', 'text="Weiter einkaufen"']
    for sel in upsells:
        try:
            if await page.locator(sel).is_visible(timeout=500):
                await page.click(sel)
                await page.wait_for_timeout(500)
        except:
            pass

async def get_rank(context, item) -> int:
    page = await context.new_page()
    url = f"https://www.{item['domain']}/dp/{item['asin']}?th=1&psc=1"
    print(f"  Fetching {item['name']} ({item['domain']})...")

    try:
        await page.goto(url, wait_until="domcontentloaded", timeout=30000)
        await handle_blockers(page)

        text_content = ""
        
        # Samla text från relevanta delar
        selectors = [
            "#prodDetails", 
            "#detailBullets_feature_div", 
            "#productDescription", 
            "table", 
            "ul"
        ]

        for sel in selectors:
            elements = await page.locator(sel).all_inner_texts()
            for txt in elements:
                text_content += txt + "\n"
        
        if len(text_content) < 100:
            text_content = await page.inner_text("body")

        rank = extract_lowest_rank(text_content, item['keywords'])

        if rank > 0:
            print(f"    ✅ Rank: #{rank}")
        else:
            print(f"    ⚠️ Kunde inte hitta rank för {item['name']}")
            # Debug: Om det fortfarande felar, avkommentera dessa rader:
            # print("    DEBUG: Letade efter:", item['keywords'])
            # await page.screenshot(path=f"debug_{item['name']}.png")
        
        return rank

    except Exception as e:
        print(f"    ❌ Error: {e}")
        return 0
    finally:
        await page.close()

def append_to_excel(data_dict):
    file_exists = os.path.exists(XLSX_PATH)
    headers = ["Date"] + [t["name"] for t in TARGETS]

    if not file_exists:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(headers)
    else:
        wb = openpyxl.load_workbook(XLSX_PATH)
        if SHEET_NAME in wb.sheetnames:
            ws = wb[SHEET_NAME]
        else:
            ws = wb.create_sheet(SHEET_NAME)
            ws.append(headers)

    row = [data_dict["Date"]]
    for t in TARGETS:
        row.append(data_dict.get(t["name"], 0))
    
    ws.append(row)
    wb.save(XLSX_PATH)
    print(f"💾 Data saved to {XLSX_PATH}")

async def run():
    today = date.today().isoformat()
    results = {"Date": today}

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=HEADLESS)
        
        print(f"--- Starting Final V3 Rank Scraping ({today}) ---")

        for item in TARGETS:
            context = await browser.new_context(
                locale=item['locale'],
                viewport={"width": 1920, "height": 1080},
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
            )
            
            rank = await get_rank(context, item)
            results[item['name']] = rank
            await context.close()
        
        await browser.close()

    append_to_excel(results)

if __name__ == "__main__":
    asyncio.run(run())