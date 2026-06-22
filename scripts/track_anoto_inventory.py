#!/usr/bin/env python3
"""
track_anoto_inventory.py

Tracks inq.shop (Anoto/Inq) AND shop.neosmartpen.com (Neo Smart Pen)
per-variant inventory and estimates daily sales via stock deltas.
Results are written to state JSON files and a shared Excel workbook.

──────────────────────────────────────────────────────────────────────────────
Anoto / inq.shop — inventory data
──────────────────────────────────────────────────────────────────────────────
inq.shop runs on Shopify.  Every product page contains a:

    <script type="application/json" data-section-type="product" …>
      {
        "product":         { …variants with price, SKU, title… },
        "variantInventory": {
          "<variant_id>": {"inventory_quantity": 444, …},
          …
        },
        …
      }
    </script>

This block exposes real inventory counts.  The script fetches each product page
as plain HTML, extracts this JSON block with a regex, and reads
inventory_quantity per variant.

──────────────────────────────────────────────────────────────────────────────
Neo Smart Pen — inventory data
──────────────────────────────────────────────────────────────────────────────
shop.neosmartpen.com is also on Shopify, but its product HTML pages redirect
to a Shopify checkout intermediary.  Instead, the script uses the public
Shopify product JSON endpoint:

    GET /products/<handle>.json

This returns full variant objects including inventory_quantity and
price_currency directly — no HTML parsing needed.

──────────────────────────────────────────────────────────────────────────────
Sales estimation methodology (applies to both stores)
──────────────────────────────────────────────────────────────────────────────
- Negative stock delta (curr < prev) → estimated units sold.
- Positive delta                     → restock / return (ignored for revenue).
- Day 1 has no prior snapshot → all deltas are zero (baseline only).
- est_revenue = est_sold_units × variant_price

State files :
  data/anoto_inventory_state.json
  data/neo_inventory_state.json
Excel output: data/anoto_inventory.xlsx  (all sheets for both stores)
"""

import json
import re
import time
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Configuration — Anoto / inq.shop ──────────────────────────────────────────
SHOP_BASE_URL   = "https://inq.shop"

# Add ?currency=USD to page fetches to force USD pricing regardless of IP geo.
# Set to "" to use whatever the server returns.
FORCE_CURRENCY  = "USD"

SCRIPT_DIR      = Path(__file__).resolve().parent
STATE_FILE      = (SCRIPT_DIR / ".." / "data" / "anoto_inventory_state.json").resolve()

# Products whose ALL variants match this SKU string are skipped entirely.
SKIP_SKU        = "ROUTEINS"
# Product titles containing any of these strings are also skipped.
SKIP_TITLES     = ["Shipping Protection"]

# ── Configuration — Neo Smart Pen ──────────────────────────────────────────────
NEO_SHOP_BASE_URL   = "https://shop.neosmartpen.com"
# Neo serves SEK prices to European IPs.  Leave empty to accept the server
# default, or set to e.g. "USD" if you want to force a specific currency.
NEO_FORCE_CURRENCY  = ""
NEO_STATE_FILE      = (SCRIPT_DIR / ".." / "data" / "neo_inventory_state.json").resolve()
# Neo product titles / SKU prefixes to skip (e.g. gift-card, shipping).
NEO_SKIP_TITLES     = []

# ── Shared configuration ────────────────────────────────────────────────────────
XLSX_PATH       = (SCRIPT_DIR / ".." / "data" / "anoto_inventory.xlsx").resolve()

# Pause between product-page fetches to avoid rate-limiting.
REQUEST_DELAY   = 1.5   # seconds

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/121.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}

# Compiled regex to extract the product section JSON from the Anoto page HTML.
_SECTION_RE = re.compile(
    r'<script[^>]+type=["\']application/json["\'][^>]+'
    r'data-section-type=["\']product["\'][^>]*>(.*?)</script>',
    re.DOTALL | re.IGNORECASE,
)


# ── State I/O — Anoto ─────────────────────────────────────────────────────────

def load_state() -> dict:
    if STATE_FILE.exists():
        return json.loads(STATE_FILE.read_text(encoding="utf-8-sig"))
    return {
        "daily_summary":   [],
        "last_snapshot":   {},   # {variant_id_str: stock_int}
        "product_catalog": {},   # {variant_id_str: {product_title, variant_title, sku, price, currency}}
    }


def save_state(state: dict) -> None:
    STATE_FILE.write_text(
        json.dumps(state, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


# ── State I/O — Neo Smart Pen ─────────────────────────────────────────────────

def load_neo_state() -> dict:
    if NEO_STATE_FILE.exists():
        return json.loads(NEO_STATE_FILE.read_text(encoding="utf-8-sig"))
    return {
        "daily_summary":   [],
        "last_snapshot":   {},
        "product_catalog": {},
    }


def save_neo_state(state: dict) -> None:
    NEO_STATE_FILE.write_text(
        json.dumps(state, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


# ── Product discovery — Anoto ─────────────────────────────────────────────────

def fetch_product_handles() -> list[dict]:
    """
    Fetch all product handles from /products.json, filtering out Route and
    any other non-trackable products.

    Returns list of {id, handle, title} dicts.
    """
    url  = f"{SHOP_BASE_URL}/products.json"
    resp = requests.get(url, params={"limit": 250}, headers=HEADERS, timeout=(10, 30))
    resp.raise_for_status()
    raw = resp.json().get("products", [])

    result: list[dict] = []
    for p in raw:
        title    = p.get("title", "")
        variants = p.get("variants", [])

        # Skip shipping protection by title
        if any(pat.lower() in title.lower() for pat in SKIP_TITLES):
            continue
        # Skip if every variant has the Route SKU
        if variants and all(SKIP_SKU in (v.get("sku") or "") for v in variants):
            continue

        result.append({"id": p["id"], "handle": p["handle"], "title": title})

    return result


# ── Per-product inventory fetch — Anoto ──────────────────────────────────────

def fetch_product_inventory(handle: str) -> tuple[dict[str, int], dict[str, dict]]:
    """
    Fetch a single product page and extract per-variant inventory + metadata.

    Returns
    -------
    inventory : {variant_id_str: inventory_quantity_int}
    catalog   : {variant_id_str: {product_title, variant_title, sku, price, currency}}
    """
    params = {}
    if FORCE_CURRENCY:
        params["currency"] = FORCE_CURRENCY

    url = f"{SHOP_BASE_URL}/products/{handle}"
    try:
        resp = requests.get(url, headers=HEADERS, params=params, timeout=(10, 30))
        resp.raise_for_status()
    except requests.RequestException as exc:
        print(f"    [WARN] {handle}: {exc}")
        return {}, {}

    m = _SECTION_RE.search(resp.text)
    if not m:
        print(f"    [WARN] {handle}: product JSON block not found")
        return {}, {}

    try:
        data = json.loads(m.group(1))
    except json.JSONDecodeError as exc:
        print(f"    [WARN] {handle}: JSON parse error — {exc}")
        return {}, {}

    product       = data.get("product") or {}
    variant_inv   = data.get("variantInventory") or {}
    product_title = product.get("title", handle)
    currency      = FORCE_CURRENCY or "?"

    # Build variant metadata map from the embedded product object.
    variant_meta: dict[str, dict] = {}
    for v in product.get("variants") or []:
        vid = str(v.get("id", ""))
        try:
            price = float(v.get("price") or 0)
        except (TypeError, ValueError):
            price = 0.0
        variant_meta[vid] = {
            "variant_title": v.get("title", ""),
            "sku":           v.get("sku", ""),
            "price":         price,
            "available":     bool(v.get("available")),
        }

    inventory: dict[str, int]  = {}
    catalog:   dict[str, dict] = {}

    for vid_str, inv_data in variant_inv.items():
        qty = inv_data.get("inventory_quantity")
        if qty is None:
            continue
        inventory[vid_str] = int(qty)

        meta = variant_meta.get(vid_str, {})
        catalog[vid_str] = {
            "product_title": product_title,
            "variant_title": meta.get("variant_title", ""),
            "sku":           meta.get("sku", ""),
            "price":         meta.get("price", 0.0),
            "currency":      currency,
            "handle":        handle,
        }

    return inventory, catalog


def fetch_all_inventory(
    handles: list[dict],
) -> tuple[dict[str, int], dict[str, dict]]:
    """Fetch inventory for every Anoto product, returning merged dicts."""
    all_inv:     dict[str, int]  = {}
    all_catalog: dict[str, dict] = {}

    for p in handles:
        handle = p["handle"]
        title  = p["title"]
        print(f"  [{title}]  ({handle})")
        inv, cat = fetch_product_inventory(handle)
        if inv:
            print(f"    → {len(inv)} variant(s) with inventory data")
        else:
            print("    → no inventory data found")
        all_inv.update(inv)
        all_catalog.update(cat)
        time.sleep(REQUEST_DELAY)

    return all_inv, all_catalog


# ── Product discovery — Neo Smart Pen ─────────────────────────────────────────

def fetch_neo_product_handles() -> list[dict]:
    """
    Fetch all product handles from Neo Smart Pen /products.json.
    Handles pagination automatically (Shopify limit=250 per page).
    Returns list of {id, handle, title} dicts.
    """
    result: list[dict] = []
    page = 1
    while True:
        url  = f"{NEO_SHOP_BASE_URL}/products.json"
        resp = requests.get(
            url,
            params={"limit": 250, "page": page},
            headers=HEADERS,
            timeout=(10, 30),
        )
        resp.raise_for_status()
        products = resp.json().get("products", [])
        if not products:
            break

        for p in products:
            title = p.get("title", "")
            if any(pat.lower() in title.lower() for pat in NEO_SKIP_TITLES):
                continue
            result.append({"id": p["id"], "handle": p["handle"], "title": title})

        if len(products) < 250:
            break
        page += 1
        time.sleep(REQUEST_DELAY)

    return result


# ── Per-product inventory fetch — Neo Smart Pen ───────────────────────────────

def fetch_neo_product_inventory(handle: str) -> tuple[dict[str, int], dict[str, dict]]:
    """
    Fetch a single Neo Smart Pen product via /products/<handle>.json and
    extract per-variant inventory + metadata.

    The individual product JSON endpoint exposes inventory_quantity and
    price_currency directly in each variant object — no HTML scraping needed.

    Returns
    -------
    inventory : {variant_id_str: inventory_quantity_int}
    catalog   : {variant_id_str: {product_title, variant_title, sku, price, currency}}
    """
    url = f"{NEO_SHOP_BASE_URL}/products/{handle}.json"
    try:
        resp = requests.get(url, headers=HEADERS, timeout=(10, 30))
        resp.raise_for_status()
    except requests.RequestException as exc:
        print(f"    [WARN] neo/{handle}: {exc}")
        return {}, {}

    try:
        data = resp.json()
    except ValueError as exc:
        print(f"    [WARN] neo/{handle}: JSON parse error — {exc}")
        return {}, {}

    product       = data.get("product") or {}
    product_title = product.get("title", handle)
    variants      = product.get("variants") or []

    inventory: dict[str, int]  = {}
    catalog:   dict[str, dict] = {}

    for v in variants:
        # Skip variants where inventory is not managed by Shopify
        if v.get("inventory_management") != "shopify":
            continue
        qty = v.get("inventory_quantity")
        if qty is None:
            continue

        vid = str(v["id"])
        try:
            price = float(v.get("price") or 0)
        except (TypeError, ValueError):
            price = 0.0

        # Use price_currency from variant if available; fall back to config.
        currency = (
            NEO_FORCE_CURRENCY
            or v.get("price_currency")
            or "?"
        )

        inventory[vid] = int(qty)
        catalog[vid] = {
            "product_title": product_title,
            "variant_title": v.get("title", ""),
            "sku":           v.get("sku", ""),
            "price":         price,
            "currency":      currency,
            "handle":        handle,
        }

    return inventory, catalog


def fetch_all_neo_inventory(
    handles: list[dict],
) -> tuple[dict[str, int], dict[str, dict]]:
    """Fetch inventory for every Neo Smart Pen product, returning merged dicts."""
    all_inv:     dict[str, int]  = {}
    all_catalog: dict[str, dict] = {}

    for p in handles:
        handle = p["handle"]
        title  = p["title"]
        print(f"  [{title}]  ({handle})")
        inv, cat = fetch_neo_product_inventory(handle)
        if inv:
            print(f"    → {len(inv)} variant(s) with inventory data")
        else:
            print("    → no inventory data found (unmanaged or zero)")
        all_inv.update(inv)
        all_catalog.update(cat)
        time.sleep(REQUEST_DELAY)

    return all_inv, all_catalog


# ── Delta computation ──────────────────────────────────────────────────────────

def compute_summary(
    curr_inv:      dict[str, int],
    last_snapshot: dict[str, int],
    catalog:       dict[str, dict],
) -> tuple[dict, list[dict]]:
    """
    Compare current inventory to the previous snapshot and produce a summary.

    Returns
    -------
    summary     : aggregated metrics for this run
    detail_rows : one dict per variant, sorted by product title / variant title
    """
    is_first_run  = not last_snapshot
    detail_rows:  list[dict] = []
    by_product:   dict[str, dict] = {}

    # Derive currency label from catalog (first entry wins, fallback to "?")
    currency_label = "?"
    if catalog:
        currency_label = next(iter(catalog.values()), {}).get("currency", "?")

    total_est_units = 0
    total_est_rev   = 0.0
    total_restocks  = 0

    for vid, curr_stock in curr_inv.items():
        meta   = catalog.get(vid, {})
        ptitle = meta.get("product_title", "Unknown")
        vtitle = meta.get("variant_title", "")
        sku    = meta.get("sku", "")
        price  = meta.get("price", 0.0)

        prev_stock = last_snapshot.get(vid)
        if prev_stock is None or is_first_run:
            delta    = None   # no comparison available
            est_sold = 0
        else:
            delta    = curr_stock - prev_stock
            est_sold = max(0, -delta)   # negative delta = sold units

        est_rev    = round(est_sold * price, 2)
        is_restock = delta is not None and delta > 0

        total_est_units += est_sold
        total_est_rev   += est_rev
        if is_restock:
            total_restocks += 1

        if ptitle not in by_product:
            by_product[ptitle] = {
                "est_sold_units": 0,
                "est_rev":        0.0,
                "restocks":       0,
            }
        by_product[ptitle]["est_sold_units"] += est_sold
        by_product[ptitle]["est_rev"]        += est_rev
        if is_restock:
            by_product[ptitle]["restocks"]   += 1

        detail_rows.append({
            "variant_id":    vid,
            "product_title": ptitle,
            "variant_title": vtitle,
            "sku":           sku,
            "price":         price,
            "currency":      meta.get("currency", "?"),
            "stock_prev":    "" if (prev_stock is None or is_first_run) else prev_stock,
            "stock_curr":    curr_stock,
            "delta":         "" if (delta is None)                       else delta,
            "est_sold":      est_sold,
            "est_rev":       est_rev,
            "is_restock":    is_restock,
        })

    # Round the summary revenue
    by_product = {k: {**v, "est_rev": round(v["est_rev"], 2)} for k, v in by_product.items()}

    summary = {
        "total_variants": len(curr_inv),
        "est_sold_units": total_est_units,
        "est_revenue":    round(total_est_rev, 2),
        "currency":       currency_label,
        "restocks":       total_restocks,
        "by_product":     by_product,
    }

    detail_rows.sort(key=lambda r: (r["product_title"], r["variant_title"]))
    return summary, detail_rows


# ── Excel writer ───────────────────────────────────────────────────────────────

_HDR_FILL = PatternFill("solid", fgColor="1F497D")
_HDR_FONT = Font(bold=True, color="FFFFFF")


def _write_headers(ws, headers: list[str]) -> None:
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = Alignment(horizontal="center")


def _autofit(ws) -> None:
    for col in ws.columns:
        max_len = max(
            (len(str(c.value)) for c in col if c.value is not None), default=8
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(
            max_len + 4, 55
        )


def write_excel(anoto_state: dict, neo_state: dict) -> None:
    if XLSX_PATH.exists():
        wb = load_workbook(XLSX_PATH)
    else:
        wb = Workbook()
        for name in list(wb.sheetnames):
            del wb[name]

    # ─────────────────────────────────────────────────────────────────────────
    # Helpers to (re-)create sheets for a given state dict and sheet-name prefix
    # ─────────────────────────────────────────────────────────────────────────
    def _curr_label(state: dict, fallback: str = "USD") -> str:
        cat = state.get("product_catalog") or {}
        if cat:
            return next(iter(cat.values()), {}).get("currency", fallback)
        return fallback

    def _write_store_sheets(
        state: dict,
        prefix: str,
        sheet_index_start: int,
        curr_label: str,
    ) -> None:
        """Write the four standard sheets for one store."""
        all_entries = state.get("daily_summary", [])

        # Sheet: Daily Summary
        ds = f"{prefix}Daily Summary"
        if ds in wb.sheetnames:
            del wb[ds]
        ws_ds = wb.create_sheet(ds, sheet_index_start)
        _write_headers(ws_ds, [
            "Date",
            "Est. Sold (units)",
            f"Est. Revenue ({curr_label})",
            "Restocks",
        ])
        for entry in all_entries:
            s = entry.get("summary", {})
            ws_ds.append([
                entry.get("date", ""),
                s.get("est_sold_units", 0),
                s.get("est_revenue", 0.0),
                s.get("restocks", 0),
            ])
        _autofit(ws_ds)

        # Sheet: By Product
        pn = f"{prefix}By Product"
        if pn in wb.sheetnames:
            del wb[pn]
        ws_p = wb.create_sheet(pn, sheet_index_start + 1)

        all_prods: set[str] = set()
        for entry in all_entries:
            all_prods.update(entry.get("summary", {}).get("by_product", {}).keys())

        latest_bp = (
            all_entries[-1].get("summary", {}).get("by_product", {})
            if all_entries else {}
        )
        sorted_prods = sorted(
            all_prods, key=lambda p: -latest_bp.get(p, {}).get("est_rev", 0)
        )
        prod_cols = (
            [f"{p} (Units)" for p in sorted_prods]
            + [f"{p} (Rev {curr_label})" for p in sorted_prods]
        )
        _write_headers(ws_p, ["Date"] + prod_cols)
        for entry in all_entries:
            by_p = entry.get("summary", {}).get("by_product", {})
            row: list = [entry.get("date", "")]
            for p in sorted_prods:
                row.append(by_p.get(p, {}).get("est_sold_units", 0))
            for p in sorted_prods:
                row.append(by_p.get(p, {}).get("est_rev", 0.0))
            ws_p.append(row)
        _autofit(ws_p)

        # Sheet: Latest Snapshot
        snap = f"{prefix}Latest Snapshot"
        if snap in wb.sheetnames:
            del wb[snap]
        ws_s = wb.create_sheet(snap, sheet_index_start + 2)
        _write_headers(ws_s, [
            "Product", "Variant", "SKU",
            f"Price ({curr_label})", "Stock Today", "Stock Yesterday",
            "Delta", "Est. Sold", f"Est. Revenue ({curr_label})",
        ])
        if all_entries:
            for row_d in all_entries[-1].get("detail_rows", []):
                ws_s.append([
                    row_d.get("product_title", ""),
                    row_d.get("variant_title", ""),
                    row_d.get("sku", ""),
                    row_d.get("price", 0.0),
                    row_d.get("stock_curr", ""),
                    row_d.get("stock_prev", ""),
                    row_d.get("delta", ""),
                    row_d.get("est_sold", 0),
                    row_d.get("est_rev", 0.0),
                ])
        _autofit(ws_s)

        # Sheet: History Detail
        hist = f"{prefix}History Detail"
        if hist in wb.sheetnames:
            del wb[hist]
        ws_h = wb.create_sheet(hist, sheet_index_start + 3)
        _write_headers(ws_h, [
            "Date", "Product", "Variant", "SKU",
            f"Price ({curr_label})", "Stock", "Delta",
            "Est. Sold", f"Est. Revenue ({curr_label})",
        ])
        for entry in all_entries:
            d = entry.get("date", "")
            for row_d in entry.get("detail_rows", []):
                ws_h.append([
                    d,
                    row_d.get("product_title", ""),
                    row_d.get("variant_title", ""),
                    row_d.get("sku", ""),
                    row_d.get("price", 0.0),
                    row_d.get("stock_curr", ""),
                    row_d.get("delta", ""),
                    row_d.get("est_sold", 0),
                    row_d.get("est_rev", 0.0),
                ])
        _autofit(ws_h)

    # ── Anoto sheets (indices 0-3) ────────────────────────────────────────────
    _write_store_sheets(
        anoto_state,
        prefix="",                  # no prefix keeps original sheet names
        sheet_index_start=0,
        curr_label=_curr_label(anoto_state, "USD"),
    )

    # ── Neo Smart Pen sheets (indices 4-7) ────────────────────────────────────
    _write_store_sheets(
        neo_state,
        prefix="Neo - ",
        sheet_index_start=4,
        curr_label=_curr_label(neo_state, "SEK"),
    )

    wb.save(XLSX_PATH)
    print(f"  Saved -> {XLSX_PATH}")


# ── Main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    today = date.today().isoformat()
    now   = datetime.now().isoformat(timespec="seconds")

    print(f"[{now}] Inventory tracker — Anoto/inq.shop + Neo Smart Pen")
    print("=" * 60)

    # ══════════════════════════════════════════════════════════════════════════
    # PART 1 — Anoto / inq.shop
    # ══════════════════════════════════════════════════════════════════════════
    print("\n\u2500\u2500 Anoto / inq.shop \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500")
    anoto_state = load_state()
    anoto_skip  = False

    if anoto_state.get("daily_summary") and anoto_state["daily_summary"][-1]["date"] == today:
        print(
            f"  Already ran today ({today}). Delete the last entry in daily_summary "
            f"from {STATE_FILE.name} to re-run."
        )
        anoto_skip = True
    else:
        print("\nDiscovering products ...")
        handles = fetch_product_handles()
        if not handles:
            print("  No trackable products found — check shop URL.")
            anoto_skip = True
        else:
            print(f"  {len(handles)} product(s) to track:")
            for p in handles:
                print(f"    {p['handle']}  (id={p['id']})")

            print(f"\nFetching inventory (currency={FORCE_CURRENCY or 'geo-default'}) ...")
            curr_inv, catalog = fetch_all_inventory(handles)
            print(f"\n  Total variants with inventory data: {len(curr_inv)}")

            if not curr_inv:
                print("  No inventory data fetched — skipping Anoto.")
                anoto_skip = True
            else:
                print("\nComputing deltas ...")
                last_snapshot = anoto_state.get("last_snapshot") or {}
                is_first_run  = not last_snapshot
                summary, detail_rows = compute_summary(curr_inv, last_snapshot, catalog)

                if is_first_run:
                    print("  (First run — all deltas are zero / baseline only.)")

                curr_label = summary.get("currency", "USD")
                print(f"  Variants tracked         : {summary['total_variants']}")
                print(f"  Est. sold today (units)  : {summary['est_sold_units']}")
                print(f"  Est. revenue today       : {summary['est_revenue']:,.2f} {curr_label}")
                print(f"  Restocks detected        : {summary['restocks']}")

                print(f"\n  By product (est. revenue {curr_label}):")
                for ptitle, pdata in sorted(
                    summary["by_product"].items(), key=lambda x: -x[1]["est_rev"]
                ):
                    print(
                        f"    [{ptitle}]  {pdata['est_rev']:,.2f} {curr_label}"
                        f"  ({pdata['est_sold_units']} units)"
                        f"  restocks={pdata['restocks']}"
                    )

                if not isinstance(anoto_state.get("daily_summary"), list):
                    anoto_state["daily_summary"] = []

                anoto_state["daily_summary"].append({
                    "date":        today,
                    "timestamp":   now,
                    "summary":     summary,
                    "detail_rows": detail_rows,
                })
                anoto_state["last_snapshot"]   = curr_inv
                anoto_state["product_catalog"] = {
                    **anoto_state.get("product_catalog", {}), **catalog
                }

                save_state(anoto_state)
                print(f"\n  State saved -> {STATE_FILE.name}")

    # ══════════════════════════════════════════════════════════════════════════
    # PART 2 — Neo Smart Pen
    # ══════════════════════════════════════════════════════════════════════════
    print("\n\u2500\u2500 Neo Smart Pen \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500")
    neo_state = load_neo_state()
    neo_skip  = False

    if neo_state.get("daily_summary") and neo_state["daily_summary"][-1]["date"] == today:
        print(
            f"  Already ran today ({today}). Delete the last entry in daily_summary "
            f"from {NEO_STATE_FILE.name} to re-run."
        )
        neo_skip = True
    else:
        print("\nDiscovering Neo Smart Pen products ...")
        neo_handles = fetch_neo_product_handles()
        if not neo_handles:
            print("  No trackable products found — check Neo shop URL.")
            neo_skip = True
        else:
            print(f"  {len(neo_handles)} product(s) to track:")
            for p in neo_handles:
                print(f"    {p['handle']}  (id={p['id']})")

            currency_info = f"currency={NEO_FORCE_CURRENCY}" if NEO_FORCE_CURRENCY else "geo-default currency"
            print(f"\nFetching Neo inventory ({currency_info}) ...")
            neo_curr_inv, neo_catalog = fetch_all_neo_inventory(neo_handles)
            print(f"\n  Total variants with inventory data: {len(neo_curr_inv)}")

            if not neo_curr_inv:
                print("  No inventory data fetched — skipping Neo.")
                neo_skip = True
            else:
                print("\nComputing Neo deltas ...")
                neo_last_snapshot = neo_state.get("last_snapshot") or {}
                neo_is_first_run  = not neo_last_snapshot
                neo_summary, neo_detail_rows = compute_summary(
                    neo_curr_inv, neo_last_snapshot, neo_catalog
                )

                if neo_is_first_run:
                    print("  (First run — all deltas are zero / baseline only.)")

                neo_curr_label = neo_summary.get("currency", "SEK")
                print(f"  Variants tracked         : {neo_summary['total_variants']}")
                print(f"  Est. sold today (units)  : {neo_summary['est_sold_units']}")
                print(f"  Est. revenue today       : {neo_summary['est_revenue']:,.2f} {neo_curr_label}")
                print(f"  Restocks detected        : {neo_summary['restocks']}")

                print(f"\n  By product (est. revenue {neo_curr_label}):")
                for ptitle, pdata in sorted(
                    neo_summary["by_product"].items(), key=lambda x: -x[1]["est_rev"]
                ):
                    print(
                        f"    [{ptitle}]  {pdata['est_rev']:,.2f} {neo_curr_label}"
                        f"  ({pdata['est_sold_units']} units)"
                        f"  restocks={pdata['restocks']}"
                    )

                if not isinstance(neo_state.get("daily_summary"), list):
                    neo_state["daily_summary"] = []

                neo_state["daily_summary"].append({
                    "date":        today,
                    "timestamp":   now,
                    "summary":     neo_summary,
                    "detail_rows": neo_detail_rows,
                })
                neo_state["last_snapshot"]   = neo_curr_inv
                neo_state["product_catalog"] = {
                    **neo_state.get("product_catalog", {}), **neo_catalog
                }

                save_neo_state(neo_state)
                print(f"\n  State saved -> {NEO_STATE_FILE.name}")

    # ══════════════════════════════════════════════════════════════════════════
    # Write combined Excel (always, even if one store was skipped today)
    # ══════════════════════════════════════════════════════════════════════════
    if not (anoto_skip and neo_skip):
        print("\nWriting combined Excel ...")
        write_excel(anoto_state, neo_state)
    else:
        print("\nBoth stores already ran today — skipping Excel update.")

    print("\nDone.")


if __name__ == "__main__":
    main()
