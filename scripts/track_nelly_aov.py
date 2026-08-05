#!/usr/bin/env python3

import os
import csv
import re
import time
import statistics
from datetime import date
from bs4 import BeautifulSoup

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager

# NEW: Excel
from openpyxl import Workbook, load_workbook
from pathlib import Path

from core.db import safe_insert

# ──────────────────────────────────────────────────────────────────────────────
URLS = [
    "https://nelly.com/se/topplistan/?page=1",
    "https://nelly.com/se/topplistan/?page=2",
    "https://nelly.com/se/topplistan/?page=3",
    "https://nelly.com/se/topplistan/?page=4",
    "https://nelly.com/se/topplistan/?page=5",
    "https://nelly.com/se/topplistan/?page=6",
    "https://nelly.com/se/topplistan/?page=7",
    "https://nelly.com/se/topplistan/?page=8",
    "https://nelly.com/se/topplistan/?page=9",
    "https://nelly.com/se/topplistan/?page=10",
]

# Each product card on the listing page shows exactly one price (already the
# effective/selling price - there's no separate struck-through "original price"
# element on this page anymore, verified 2026-08-05 against the current site).
# Was `span.text-sm.text-darkGrey` with a separate `ins` discount selector
# until Nelly's frontend rebuild renamed the class (see KNOWN_ISSUES.md #6).
PRICE_SELECTOR = "span.text-subhead.leading-none.text-darkGrey"

# OLD: CSV_FILENAME = "nelly_aov.csv"
# NEW: write to ../data/nelly_aov.xlsx (repo-root/data)
SCRIPT_DIR = Path(__file__).resolve().parent
XLSX_PATH = (SCRIPT_DIR / ".." / "data" / "nelly_aov.xlsx").resolve()
# ──────────────────────────────────────────────────────────────────────────────

def create_driver():
    opts = Options()
    opts.add_argument("--headless")
    opts.add_argument("--window-size=1920,1080")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--no-sandbox")
    opts.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/115.0.0.0 Safari/537.36"
    )
    service = Service(ChromeDriverManager().install())
    return webdriver.Chrome(service=service, options=opts)

def _to_number(text):
    # Keep digits, comma, dot; normalize comma to dot; strip NBSPs.
    num = re.sub(r"[^\d,\.]", "", text.replace("\xa0", " "))
    num = num.replace(",", ".")
    try:
        return float(num)
    except ValueError:
        return None

def fetch_prices(driver, url):
    driver.get(url)
    time.sleep(3)
    # Trigger lazy loading
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(2)

    soup = BeautifulSoup(driver.page_source, "html.parser")

    prices = []
    for tag in soup.select(PRICE_SELECTOR):
        val = _to_number(tag.get_text(strip=True))
        if val is not None:
            prices.append(val)

    return prices

# ──────────────────────────────────────────────────────────────────────────────
# Excel handling (replaces CSV)

def ensure_header_xlsx():
    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    if XLSX_PATH.exists():
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["date", "median_price", "average_price"])
    wb.save(XLSX_PATH)

def append_to_xlsx(median_price, average_price):
    today = date.today().isoformat()
    if not XLSX_PATH.exists():
        ensure_header_xlsx()
    wb = load_workbook(XLSX_PATH)
    ws = wb.active
    ws.append([today, round(median_price, 2), round(average_price, 2)])
    wb.save(XLSX_PATH)
    print(f"✓ Appended {today}: median={median_price:.2f}, avg={average_price:.2f} → {XLSX_PATH}")

    # Best-effort: also write to Postgres (must not break the script)
    db_rows_written, db_error = safe_insert(
        table="nelly_aov",
        columns=["snapshot_date", "median_price", "average_price"],
        rows=[(today, round(median_price, 2), round(average_price, 2))],
        conflict_columns=["snapshot_date"],
    )
    if db_error is not None:
        print(f"Databas: MISSLYCKADES – {db_error}")
    else:
        print(f"Databas: {db_rows_written} rader skrivna")

# ──────────────────────────────────────────────────────────────────────────────

def main():
    ensure_header_xlsx()
    driver = create_driver()
    all_prices = []

    for url in URLS:
        ps = fetch_prices(driver, url)
        print(f"  • Nelly Topplistan – found {len(ps)} prices on {url}")
        all_prices.extend(ps)

    driver.quit()

    if not all_prices:
        print("❌ No prices found; check selectors or if the site changed its HTML.")
        return

    med = statistics.median(all_prices)
    avg = statistics.mean(all_prices)
    append_to_xlsx(med, avg)

if __name__ == "__main__":
    main()
