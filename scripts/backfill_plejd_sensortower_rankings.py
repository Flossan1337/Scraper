# backfill_plejd_sensortower_rankings.py
#
# ONE-TIME backfill: fetches 90 days of Plejd category ranking history
# from the Sensor Tower internal API (no auth required) and writes each
# date as a row into data/plejd_sensortower_rankings.xlsx.
# Skips dates that are already present in the sheet.

import sys
from datetime import date, timedelta, datetime, timezone
from pathlib import Path

import requests
from openpyxl import load_workbook

from excel_utils import append_row

# ── CONFIG ─────────────────────────────────────────────────────────────────────
APP_ID      = "1032689423"
CATEGORY    = "6012"        # Lifestyle
CHART_TYPE  = "topfreeapplications"

COUNTRIES = ["SE", "NO", "FI", "NL", "DE", "DK", "ES"]

BACKFILL_DAYS = 90  # Sensor Tower free tier gives 90 days

REPO_ROOT  = Path(__file__).resolve().parent.parent
DATA_DIR   = REPO_ROOT / "data"
DATA_DIR.mkdir(exist_ok=True)
XLSX_PATH  = str(DATA_DIR / "plejd_sensortower_rankings.xlsx")
SHEET_NAME = "category_rankings"

API_URL = "https://app.sensortower.com/api/ios/category/category_history"
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Referer": "https://app.sensortower.com/app-analysis/category-rankings",
}

# ── HELPERS ────────────────────────────────────────────────────────────────────

def existing_dates() -> set[str]:
    """Return the set of date strings already written to the Excel sheet."""
    path = Path(XLSX_PATH)
    if not path.exists():
        return set()
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        if SHEET_NAME not in wb.sheetnames:
            return set()
        ws = wb[SHEET_NAME]
        return {
            str(row[0])[:10]
            for row in ws.iter_rows(min_row=2, max_col=1, values_only=True)
            if row[0] is not None
        }
    except Exception:
        return set()


def fetch_history(start: str, end: str) -> dict[str, dict[str, int | None]]:
    """
    Call the Sensor Tower category_history API for all countries at once.
    Returns: {date_str: {country: rank_or_None}}
    """
    # Build query params — multiple values for the same key
    params = [
        ("app_ids[]", APP_ID),
        ("categories[]", CATEGORY),
        ("chart_type_ids[]", CHART_TYPE),
        ("start_date", start),
        ("end_date", end),
        ("is_hourly", "false"),
    ]
    for cc in COUNTRIES:
        params.append(("countries[]", cc))

    resp = requests.get(API_URL, params=params, headers=HEADERS, timeout=30)
    resp.raise_for_status()
    data = resp.json()

    # data structure: {app_id: {country: {category: {chart_type: {graphData: [[ts, rank, _], ...]}}}}}
    # Convert to {date_str: {country: rank}}
    by_date: dict[str, dict[str, int | None]] = {}
    app_data = data.get(int(APP_ID), data.get(str(APP_ID), {}))

    for cc in COUNTRIES:
        country_data = app_data.get(cc, {})
        cat_data = country_data.get(str(CATEGORY), country_data.get(int(CATEGORY), {}))
        chart_data = cat_data.get(CHART_TYPE, {})
        graph = chart_data.get("graphData", [])

        for point in graph:
            ts, rank = point[0], point[1]
            dt_str = datetime.fromtimestamp(ts, tz=timezone.utc).strftime("%Y-%m-%d")
            if dt_str not in by_date:
                by_date[dt_str] = {c: None for c in COUNTRIES}
            by_date[dt_str][cc] = rank  # rank is already an int

    return by_date


# ── MAIN ───────────────────────────────────────────────────────────────────────

def main():
    end_date   = date.today()
    start_date = end_date - timedelta(days=BACKFILL_DAYS - 1)

    print(f"Fetching history {start_date} → {end_date} for: {', '.join(COUNTRIES)}")
    history = fetch_history(str(start_date), str(end_date))
    print(f"  API returned data for {len(history)} dates.")

    already = existing_dates()
    print(f"  Dates already in Excel: {len(already)}")

    new_dates = sorted(dt for dt in history if dt not in already)
    print(f"  New dates to write: {len(new_dates)}")

    if not new_dates:
        print("Nothing to backfill — all dates already present.")
        sys.exit(0)

    for dt_str in new_dates:
        row = {"Date": dt_str} | history[dt_str]
        append_row(XLSX_PATH, SHEET_NAME, row)
        ranks_str = "  ".join(
            f"{cc}={history[dt_str][cc] if history[dt_str][cc] is not None else '-'}"
            for cc in COUNTRIES
        )
        print(f"  {dt_str}  {ranks_str}")

    print(f"\nDone. Wrote {len(new_dates)} rows to {XLSX_PATH}")


if __name__ == "__main__":
    main()
