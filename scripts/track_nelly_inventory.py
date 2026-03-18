#!/usr/bin/env python3
"""
track_nelly_inventory.py

Tracks Nelly.com inventory at product-colour level across multiple markets
(SE, NO, DK, FI) using the Voyado Elevate API directly.

Methodology
-----------
Each daily run:
  1. On first run, uses Playwright to load nelly.com and intercept Elevate API
     XHR requests, extracting the cluster ID and working pageReferences.
     Saves them to the state file for all future runs.
  2. Queries the Voyado Elevate storefront API for every top-level category
     in each market, paginating with limit=600 until all products are fetched.
  3. Extracts per-variant stockNumber, sellingPrice, listPrice, brand, title,
     category, and inStock status.
  4. Groups variants by product-colour key (e.g. "262438-6915") and sums
     stockNumber across all sizes to get total stock per product-colour.
  5. Compares to the previous day's snapshot to compute stock deltas
     (negative delta = estimated sales proxy, with caveats below).

Why stock-delta instead of a sales counter
-------------------------------------------
Unlike RVRC's Elevate instance which exposes `variant.custom.sale_last_week`,
Nelly's Elevate instance does not expose a direct sales counter.  Stock deltas
are used as a proxy with important caveats:
  - Restocks inflate POSITIVE deltas (ignore these for sales estimation)
  - Negative deltas represent stock decreases, which are mostly sales but can
    also include: returns processed as fresh picks, inventory write-offs or
    adjustments, and ATP reservation changes.
  - Day 1 has no prior snapshot — deltas start from Day 2 onwards.

Shared warehouse, per-market listing availability
--------------------------------------------------
All markets for a given brand (Nelly or NlyMan) share a SINGLE physical stock
pool — the stockNumber returned by the API is identical across all markets for
the same variant.  Stock totals and sales-delta estimates therefore use ONLY the
SE market (W_SE → Nelly, M_SE → NlyMan) as the primary source of truth.
All other markets (NO, DK, FI, NL, DE, BE, PL, FR, AT) are tracked for product
LISTING AVAILABILITY only — showing whether each product is currently offered
in that country's storefront.

How the data is fetched
-----------------------
Nelly uses Voyado Elevate (Apptus) as its product catalogue/search backend.
The Elevate cloud API endpoint is accessible directly without authentication;
customerKey and sessionKey are random UUIDs generated fresh each run.

Endpoint: GET https://{cluster_id}.api.esales.apptus.cloud/api/storefront/v3/queries/landing-page
Params: market, locale, pageReference, limit (≤600), skip (pagination offset),
        customerKey, sessionKey, touchpoint

The cluster ID is discovered automatically on first run using Playwright, which
loads a Nelly category page and intercepts the outgoing Elevate XHR request URL.
Once discovered, the cluster ID is saved to the state file.

State file  : data/nelly_inventory_state.json
Excel output: data/nelly_inventory.xlsx
"""

import json
import re
import time
import uuid
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Elevate API configuration ──────────────────────────────────────────────────
ELEVATE_BASE_URL_TEMPLATE = "https://{cluster}.api.esales.apptus.cloud"
ELEVATE_ENDPOINT          = "/api/storefront/v3/queries/landing-page"
ELEVATE_LIMIT             = 600      # max allowed by Elevate API
ELEVATE_CLUSTER_ID        = "w67a8630f"

# Stock increases below this threshold are treated as customer returns (negative sales).
# Increases >= this value are treated as true warehouse restocks.
RESTOCK_MIN_UNITS         = 4

# Extra Elevate params to retrieve price tiers and all custom fields.
NELLY_PRESENT_PRICES = "member"
NELLY_PRESENT_CUSTOM = (
    "filterColor|swatch|inOutlet|isSellable|campaignIds|filterColorHex|"
    "categoryNode|onlyInStock|contentfulLabelPrimary|contentfulLabelSecondary|"
    "variant.historic_lowest_selling_price|variant.historic_lowest_member_price"
)

# ── Market configuration ───────────────────────────────────────────────────────
# Keys are "<SITE>_<COUNTRY>" where SITE is "W" (Nelly women) or "M" (NlyMan).
# Norway has no single top-level clothing page for either site, so sub-categories
# are fetched and deduplicated by product key within the market.
MARKETS: dict[str, dict] = {
    # ── Nelly (women) ───────────────────────────────────────────────────────
    "W_SE": {
        "site": "Nelly", "country": "SE",
        "elevate_market": "Nelly_Sweden",
        "locale": "sv-SE", "currency": "SEK",
        "primary": True,   # source of truth for Nelly stock & sales delta
        "categories": ["/klader/", "/skor/", "/accessoarer/"],
    },
    "W_NO": {
        "site": "Nelly", "country": "NO",
        "elevate_market": "Nelly_Norway",
        "locale": "nb-NO", "currency": "NOK",
        "categories": [
            "/kjoler/", "/bukser/", "/gensere/", "/jakker/",
            "/bluser/", "/skjorter/", "/cardigans/", "/skjort/",
            "/festkjoler/", "/jeans/", "/shorts/",
            "/sko/", "/sandaler/", "/sneakers/",
            "/accessories/", "/smykker/", "/vesker/",
            "/bh/", "/sport/",
        ],
    },
    "W_DK": {
        "site": "Nelly", "country": "DK",
        "elevate_market": "Nelly_Denmark",
        "locale": "da-DK", "currency": "DKK",
        "categories": ["/toj/", "/sko/", "/accessories/"],
    },
    "W_FI": {
        "site": "Nelly", "country": "FI",
        "elevate_market": "Nelly_Finland",
        "locale": "fi-FI", "currency": "EUR",
        "categories": ["/vaatteet/", "/kengat/", "/asusteet/"],
    },
    "W_NL": {
        "site": "Nelly", "country": "NL",
        "elevate_market": "Nelly_Netherlands",
        "locale": "en-US", "currency": "EUR",
        "categories": ["/clothes/", "/shoes/", "/accessories/"],
    },
    "W_DE": {
        "site": "Nelly", "country": "DE",
        "elevate_market": "Nelly_Germany",
        "locale": "en-US", "currency": "EUR",
        "categories": ["/clothes/", "/shoes/", "/accessories/"],
    },
    "W_BE": {
        "site": "Nelly", "country": "BE",
        "elevate_market": "Nelly_Belgium",
        "locale": "en-US", "currency": "EUR",
        "categories": ["/clothes/", "/shoes/", "/accessories/"],
    },
    "W_PL": {
        "site": "Nelly", "country": "PL",
        "elevate_market": "Nelly_Poland",
        "locale": "en-US", "currency": "EUR",
        "categories": ["/clothes/", "/shoes/", "/accessories/"],
    },
    "W_FR": {
        "site": "Nelly", "country": "FR",
        "elevate_market": "Nelly_France",
        "locale": "en-US", "currency": "EUR",
        "categories": ["/clothes/", "/shoes/", "/accessories/"],
    },
    "W_AT": {
        "site": "Nelly", "country": "AT",
        "elevate_market": "Nelly_Austria",
        "locale": "en-US", "currency": "EUR",
        "categories": ["/clothes/", "/shoes/", "/accessories/"],
    },
    # ── NlyMan (men) ──────────────────────────────────────────────────────
    "M_SE": {
        "site": "NlyMan", "country": "SE",
        "elevate_market": "NlyMan_Sweden",
        "locale": "sv-SE", "currency": "SEK",
        "primary": True,   # source of truth for NlyMan stock & sales delta
        "categories": ["/klader/", "/skor/", "/accessoarer/"],
    },
    "M_NO": {
        "site": "NlyMan", "country": "NO",
        "elevate_market": "NlyMan_Norway",
        "locale": "nb-NO", "currency": "NOK",
        "categories": [
            "/bukser/", "/gensere/", "/jakker/", "/jeans/",
            "/shorts/", "/hoodies/", "/cardigans/", "/vester/",
            "/sko/", "/sandaler/", "/sneakers/", "/stovler/",
            "/accessories/", "/vesker/", "/solbriller/",
            "/undertoy/", "/sport/",
        ],
    },
    "M_DK": {
        "site": "NlyMan", "country": "DK",
        "elevate_market": "NlyMan_Denmark",
        "locale": "da-DK", "currency": "DKK",
        "categories": ["/toj/", "/sko/", "/accessories/"],
    },
    "M_FI": {
        "site": "NlyMan", "country": "FI",
        "elevate_market": "NlyMan_Finland",
        "locale": "fi-FI", "currency": "EUR",
        "categories": ["/vaatteet/", "/kengat/", "/asusteet/"],
    },
    "M_NL": {
        "site": "NlyMan", "country": "NL",
        "elevate_market": "NlyMan_Netherlands",
        "locale": "en-US", "currency": "EUR",
        "categories": ["/clothes/", "/shoes/", "/accessories/"],
    },
    "M_DE": {
        "site": "NlyMan", "country": "DE",
        "elevate_market": "NlyMan_Germany",
        "locale": "en-US", "currency": "EUR",
        "categories": ["/clothes/", "/shoes/", "/accessories/"],
    },
    "M_BE": {
        "site": "NlyMan", "country": "BE",
        "elevate_market": "NlyMan_Belgium",
        "locale": "en-US", "currency": "EUR",
        "categories": ["/clothes/", "/shoes/", "/accessories/"],
    },
    "M_PL": {
        "site": "NlyMan", "country": "PL",
        "elevate_market": "NlyMan_Poland",
        "locale": "en-US", "currency": "EUR",
        "categories": ["/clothes/", "/shoes/", "/accessories/"],
    },
    "M_FR": {
        "site": "NlyMan", "country": "FR",
        "elevate_market": "NlyMan_France",
        "locale": "en-US", "currency": "EUR",
        "categories": ["/clothes/", "/shoes/", "/accessories/"],
    },
    "M_AT": {
        "site": "NlyMan", "country": "AT",
        "elevate_market": "NlyMan_Austria",
        "locale": "en-US", "currency": "EUR",
        "categories": ["/clothes/", "/shoes/", "/accessories/"],
    },
}

# Per-site primary market: the single source of truth for stock levels and
# sales-delta estimates.  All other markets share the same physical stock pool
# and are tracked for product listing AVAILABILITY only (0 = not listed, 1 = listed).
PRIMARY_MARKET_FOR_SITE: dict[str, str] = {"Nelly": "W_SE", "NlyMan": "M_SE"}

SCRIPT_DIR  = Path(__file__).resolve().parent
STATE_FILE  = (SCRIPT_DIR / ".." / "data" / "nelly_inventory_state.json").resolve()
XLSX_PATH   = (SCRIPT_DIR / ".." / "data" / "nelly_inventory.xlsx").resolve()

# Nelly category pages used for Playwright-based cluster discovery (fallback only).
NELLY_DISCOVERY_URLS = [
    "https://nelly.com/se/klader/jeans/",
    "https://nelly.com/se/klader/",
]


# ── State I/O ──────────────────────────────────────────────────────────────────

def load_state() -> dict:
    if STATE_FILE.exists():
        raw = json.loads(STATE_FILE.read_text(encoding="utf-8-sig"))
        return raw
    return {
        "cluster_id":    "",
        "daily_summary": [],
        "last_snapshot": {},   # {"<market_key>/<product_colour_key>": stock}
    }


def save_state(state: dict) -> None:
    STATE_FILE.write_text(
        json.dumps(state, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


# ── Cluster ID discovery via Playwright ───────────────────────────────────────

def discover_cluster_id() -> tuple[str, list[str]]:
    """
    Use Playwright to load a Nelly category page and intercept the outgoing
    Elevate API XHR request.  Extracts the cluster ID from the request URL and
    the pageReference from the query parameters.

    Returns (cluster_id, [pageReference, ...]) or ("", []) if discovery fails.

    Requires: pip install playwright && playwright install chromium
    """
    try:
        from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
    except ImportError:
        print("  [WARN] Playwright not installed — cannot auto-discover cluster ID.")
        print("  Run: pip install playwright && playwright install chromium")
        return "", []

    cluster_id  = ""
    page_refs:  list[str] = []
    intercepted: list[dict] = []

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=True)
        context = browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/121.0.0.0 Safari/537.36"
            ),
            locale="sv-SE",
        )
        page = context.new_page()

        def on_request(request):
            if "api.esales.apptus.cloud" in request.url:
                intercepted.append({"url": request.url})

        page.on("request", on_request)

        for url in NELLY_DISCOVERY_URLS:
            print(f"  [Discovery] Loading {url} ...")
            try:
                page.goto(url, wait_until="domcontentloaded", timeout=30_000)
            except PWTimeout:
                print(f"  [Discovery] Page-load timeout on {url}, trying next ...")
                continue
            # Wait up to 15 s for at least one Elevate XHR to fire
            deadline = time.time() + 15
            while time.time() < deadline and not intercepted:
                time.sleep(0.5)
            if intercepted:
                break

        browser.close()

    for item in intercepted:
        req_url = item["url"]
        m = re.search(r"https://([^.]+)\.api\.esales\.apptus\.cloud", req_url)
        if m and not cluster_id:
            cluster_id = m.group(1)
        pr = re.search(r"pageReference=([^&]+)", req_url)
        if pr:
            ref = pr.group(1)
            if ref not in page_refs:
                page_refs.append(ref)

    if cluster_id:
        print(f"  [Discovery] Cluster ID: {cluster_id}")
        print(f"  [Discovery] Page refs captured: {page_refs}")
    else:
        print("  [Discovery] Could not capture cluster ID from intercepted requests.")

    return cluster_id, page_refs


# ── Elevate API helpers ────────────────────────────────────────────────────────

def _get_price(price_raw) -> float:
    """Parse a price that may be a float/int or {'min': x, 'max': x} dict."""
    if isinstance(price_raw, dict):
        for key in ("sellingPrice", "listPrice", "min", "value"):
            if price_raw.get(key) is not None:
                try:
                    return float(price_raw[key])
                except (TypeError, ValueError):
                    pass
        return 0.0
    try:
        return float(price_raw) if price_raw is not None else 0.0
    except (TypeError, ValueError):
        return 0.0


def _parse_category(product: dict) -> str:
    """
    Derive a clean category label from the Elevate categoryNode custom field.
    Returns the deepest meaningful category node label, e.g. "Jeans".
    Falls back to "Other".
    """
    nodes = (product.get("custom") or {}).get("categoryNode") or []
    # Nodes are sorted shallowest → deepest; take the deepest one with a label
    if nodes:
        # Skip the root "Kläder" / "Skor" nodes if there are deeper ones
        labels = [n.get("label", "") for n in nodes if n.get("label")]
        if labels:
            return labels[-1]   # deepest category node
    return "Other"


def _group_category(full_cat: str) -> str:
    """Truncate a full '>'-separated path to at most 2 levels.
    E.g. 'Kläder>Jeans>Low waist jeans' → 'Kläder>Jeans'.
    E.g. 'Accessoarer>Väskor' → 'Accessoarer>Väskor' (unchanged).
    """
    parts = [p.strip() for p in full_cat.split(">")]
    return ">".join(parts[:2]) if len(parts) >= 2 else full_cat


def fetch_elevate_page(
    cluster_id: str,
    elevate_market: str,
    locale: str,
    page_ref: str,
    skip: int = 0,
    customer_key: str = "",
    session_key: str = "",
) -> Optional[dict]:
    """Fetch one page from the Elevate landing-page API."""
    url = ELEVATE_BASE_URL_TEMPLATE.format(cluster=cluster_id) + ELEVATE_ENDPOINT
    params = {
        "market":        elevate_market,
        "locale":        locale,
        "customerKey":   customer_key,
        "sessionKey":    session_key,
        "touchpoint":    "desktop",
        "pageReference": page_ref,
        "limit":         ELEVATE_LIMIT,
        "skip":          skip,
        "presentPrices": NELLY_PRESENT_PRICES,
        "presentCustom": NELLY_PRESENT_CUSTOM,
    }
    try:
        resp = requests.get(url, params=params, timeout=(10, 60))
        resp.raise_for_status()
        return resp.json()
    except requests.RequestException as exc:
        print(f"    [WARN] Elevate {page_ref} skip={skip}: {exc}")
        return None


def extract_products_from_page(data: dict, page_ref: str = "") -> tuple[dict[str, dict], int, int]:
    """
    Extract product-colour level data from an Elevate landing-page response.

    Returns
    -------
    (products_dict, total_hits, group_count)

    products_dict keyed by product_colour_key (e.g. "262438-6915"):
        {
            "stock":         int,   total stock across all sizes in this market
            "sizes_in_stock": int,  number of sizes with stockNumber > 0
            "sell_price":    float, representative selling price
            "list_price":    float, representative list price
            "discount_pct":  float, discount as % of list price (0 if no discount)
            "brand":         str,
            "title":         str,
            "category":      str,
            "in_stock":      bool,
        }
    """
    products: dict[str, dict] = {}
    pl = data.get("primaryList") or {}
    total_hits = int(pl.get("totalHits") or 0)
    groups = pl.get("productGroups") or []

    for group in groups:
        for product in group.get("products") or []:
            product_key = str(product.get("key") or "")
            if not product_key:
                continue

            brand    = str(product.get("brand") or "")
            title    = str(product.get("title") or product.get("name") or "")
            in_stock = bool(product.get("inStock"))
            category = _parse_category(product)

            # Badges: check for discount/sale badges
            badges_all = (
                (product.get("badges") or {}).get("primary", []) +
                (product.get("badges") or {}).get("secondary", [])
            )
            has_discount_badge = any(
                (b.get("theme") or "").upper() == "DISCOUNT"
                for b in badges_all
            )
            is_new = any(
                (b.get("theme") or "").upper() == "NEW"
                for b in badges_all
            )

            # Aggregate variant-level stock and price
            total_stock    = 0
            sizes_in_stock = 0
            sell_prices:   list[float] = []
            list_prices:   list[float] = []

            hist_low_sell: list[float] = []

            for variant in product.get("variants") or []:
                try:
                    stock = int(variant.get("stockNumber") or 0)
                except (TypeError, ValueError):
                    stock = 0
                total_stock += stock
                if stock > 0:
                    sizes_in_stock += 1

                sp = _get_price(variant.get("sellingPrice"))
                lp = _get_price(variant.get("listPrice")) or sp
                if sp > 0:
                    sell_prices.append(sp)
                if lp > 0:
                    list_prices.append(lp)

                # Historic lowest price (from presentCustom)
                v_custom = variant.get("custom") or {}
                hlsp = v_custom.get("historic_lowest_selling_price")
                if hlsp:
                    try:
                        hist_low_sell.append(float(hlsp))
                    except (TypeError, ValueError):
                        pass

            # Use median prices to handle weird outliers
            def _median_inner(lst: list) -> float:
                if not lst:
                    return 0.0
                s = sorted(lst)
                m = len(s) // 2
                return s[m] if len(s) % 2 else (s[m - 1] + s[m]) / 2.0

            sell_price = _median_inner(sell_prices) or _get_price(product.get("sellingPrice"))
            list_price = _median_inner(list_prices) or _get_price(product.get("listPrice")) or sell_price
            discount_pct = (
                round(100.0 * (1.0 - sell_price / list_price), 1)
                if list_price > 0 and sell_price < list_price else 0.0
            )
            historic_low = _median_inner(hist_low_sell) if hist_low_sell else 0.0

            products[product_key] = {
                "stock":          total_stock,
                "sizes_in_stock": sizes_in_stock,
                "sell_price":     round(sell_price, 2),
                "list_price":     round(list_price, 2),
                "discount_pct":   discount_pct,
                "historic_low":   round(historic_low, 2),
                "brand":          brand,
                "title":          title,
                "category":       category,
                "in_stock":       in_stock,
                "has_discount":   has_discount_badge,
                "is_new":         is_new,
            }

    return products, total_hits, len(groups)


def fetch_all_by_market(cluster_id: str) -> dict[str, dict[str, dict]]:
    """
    Fetch all product-colour level data for every market/site combination.
    Uses per-market category lists from the MARKETS config.

    Returns
    -------
    {market_key: {product_colour_key: {...}}}
    e.g. {"W_SE": {"262438-6915": {...}}, "M_SE": {...}, ...}
    """
    result: dict[str, dict[str, dict]] = {}

    for market_code, cfg in MARKETS.items():
        elevate_market = cfg["elevate_market"]
        locale         = cfg["locale"]
        categories     = cfg["categories"]
        customer_key   = str(uuid.uuid4())
        session_key    = str(uuid.uuid4())
        site_label     = cfg["site"]
        print(f"\n[{market_code}] {site_label} Elevate API (market={elevate_market})")

        market_products: dict[str, dict] = {}

        for cat in categories:
            skip, page = 0, 1
            cat_total  = None

            while True:
                data = fetch_elevate_page(
                    cluster_id, elevate_market, locale, cat,
                    skip, customer_key, session_key,
                )
                if data is None:
                    break

                new_p, total_hits, group_count = extract_products_from_page(data, cat)

                if cat_total is None:
                    cat_total = total_hits

                market_products.update(new_p)
                print(
                    f"  {cat} p{page}: +{len(new_p)} products "
                    f"[{skip}–{skip + group_count}/{total_hits}]"
                )

                if total_hits == 0:
                    print(f"  [WARN] {cat}: 0 hits — check pageReference")
                    break

                skip += group_count
                if skip >= total_hits or group_count == 0:
                    break
                page += 1

        site_tag = f"[{site_label}/{cfg['country']}]"
        print(f"  {site_tag} {len(market_products):,} unique product-colours fetched")
        result[market_code] = market_products

    return result


# ── Stock-delta analysis ───────────────────────────────────────────────────────

def compute_snapshot_summary(
    curr_by_market: dict[str, dict[str, dict]],
    last_snapshot:  dict[str, int],    # "{site}/{key}": primary_stock_int
) -> tuple[dict, list[dict], dict]:
    """
    Compute a daily summary from the current multi-market snapshot.

    Stock methodology (single shared warehouse pool)
    -------------------------------------------------
    All markets for a given brand share ONE stock pool.  Stock totals,
    pricing, and sales-delta estimates use ONLY the primary market per site
    (W_SE → Nelly, M_SE → NlyMan).  Prices are natively in SEK.
    All other markets contribute product listing AVAILABILITY flags (0/1) only.

    Parameters
    ----------
    curr_by_market : {market_code: {product_key: product_data}}
    last_snapshot  : {"{site}/{key}": int}  — previous run's primary stock

    Returns
    -------
    summary      : aggregated metrics for today
    detail_rows  : per-product-colour dicts for the Excel detail sheet
    new_snapshot : {"{site}/{key}": int}  — updated snapshot for state file
    """
    # Migrate from old snapshot format where values were dicts like {"W_SE": int, ...}
    # instead of plain ints.  Treat old format as an empty snapshot (first run).
    if last_snapshot and isinstance(next(iter(last_snapshot.values())), dict):
        print("  [INFO] Old snapshot format detected — treating as first run (no delta).")
        last_snapshot = {}

    # Group market codes by site; identify primary vs availability markets.
    site_avail_markets: dict[str, list[str]] = {}   # non-primary markets per site
    primary_mkt_for:    dict[str, str]       = {}   # site → primary market code
    for mc, cfg in MARKETS.items():
        site = cfg["site"]
        if cfg.get("primary"):
            primary_mkt_for[site] = mc
        else:
            site_avail_markets.setdefault(site, []).append(mc)

    detail_rows:      list[dict] = []
    new_snapshot:     dict[str, int]  = {}
    by_category:      dict[str, dict] = {}
    by_brand:         dict[str, dict] = {}

    total_products        = 0
    est_sold_sek         = 0.0
    est_sold_list_sek    = 0.0
    restocks             = 0
    returns_count        = 0
    restock_events_list:  list[dict] = []
    return_events_list:   list[dict] = []

    # Iterate over each site's primary market as the authoritative product list.
    for site, primary_mc in primary_mkt_for.items():
        avail_mkt_codes = site_avail_markets.get(site, [])

        for key, pd in curr_by_market.get(primary_mc, {}).items():
            primary_stock = pd["stock"]
            snap_key      = f"{site}/{key}"

            # Sales-delta:
            #   stock decreased            → units sold (positive est_sold)
            #   increased by < RESTOCK_MIN_UNITS → customer return (negative est_sold)
            #   increased by >= RESTOCK_MIN_UNITS → true warehouse restock (est_sold = 0)
            prev_stock  = last_snapshot.get(snap_key)
            stock_delta = (primary_stock - prev_stock) if prev_stock is not None else 0
            new_snapshot[snap_key] = primary_stock

            if prev_stock is None or stock_delta >= RESTOCK_MIN_UNITS:
                est_sold = 0            # no prior data, or true restock
            elif stock_delta > 0:
                est_sold = -stock_delta  # small increase = customer return
                returns_count += 1
                return_events_list.append({
                    "key":            key,
                    "site":           site,
                    "brand":          pd["brand"],
                    "title":          pd["title"],
                    "category":       pd["category"],
                    "stock_before":   prev_stock,
                    "stock_after":    primary_stock,
                    "delta":          stock_delta,
                    "sell_price_sek": pd["sell_price"],
                })
            else:
                est_sold = -stock_delta  # stock dropped = units sold (positive)

            # Availability flags: is this product listed in each non-primary market?
            avail: dict[str, int] = {
                mc: (1 if key in curr_by_market.get(mc, {}) else 0)
                for mc in avail_mkt_codes
            }
            listed_count = 1 + sum(avail.values())  # primary counts as 1

            # Pricing — natively in SEK (primary market is SE).
            sell_sek     = pd["sell_price"]
            list_sek     = pd["list_price"]
            rev_sek      = est_sold * sell_sek
            list_rev_sek = est_sold * list_sek
            hist_low_sek = pd.get("historic_low", 0.0)

            # Restock detection: large stock increase = true warehouse restock.
            if prev_stock is not None and stock_delta >= RESTOCK_MIN_UNITS:
                restocks += 1
                restock_events_list.append({
                    "key":            key,
                    "site":           site,
                    "brand":          pd["brand"],
                    "title":          pd["title"],
                    "category":       pd["category"],
                    "stock_before":   prev_stock,
                    "stock_after":    primary_stock,
                    "delta":          stock_delta,
                    "sell_price_sek": sell_sek,
                })

            # Global rollups.
            total_products    += 1
            est_sold_sek      += rev_sek
            est_sold_list_sek += list_rev_sek

            # Category rollup — grouped to 2 levels (e.g. 'Kläder>Jeans').
            grouped_cat = _group_category(pd["category"])
            cat_data = by_category.setdefault(grouped_cat, {
                "sell_rev_sek": 0.0, "list_rev_sek": 0.0,
            })
            cat_data["sell_rev_sek"] += rev_sek
            cat_data["list_rev_sek"] += list_rev_sek

            # Brand rollup.
            brand   = pd["brand"]
            br_data = by_brand.setdefault(brand, {
                "sell_rev_sek": 0.0, "list_rev_sek": 0.0,
            })
            br_data["sell_rev_sek"] += rev_sek
            br_data["list_rev_sek"] += list_rev_sek

            # Detail row: primary stock + availability flags per non-primary market.
            detail_rows.append({
                "key":              key,
                "site":             site,
                "brand":            brand,
                "title":            pd["title"],
                "category":         pd["category"],
                "sell_price_sek":   round(sell_sek, 0),
                "list_price_sek":   round(list_sek, 0),
                "historic_low_sek": hist_low_sek,
                "discount_pct":     pd["discount_pct"],
                "is_new":           pd["is_new"],
                "est_sold_today":   est_sold,
                f"stk_{primary_mc}": primary_stock,
                **{f"avl_{mc}": avail[mc] for mc in avail_mkt_codes},
                "primary_stock":    primary_stock,
                "listed_count":     listed_count,
            })

    summary = {
        "total_products":          total_products,
        "est_sold_today_units":    sum(r["est_sold_today"] for r in detail_rows),
        "est_sold_today_sek":      round(est_sold_sek, 0),
        "est_sold_today_list_sek": round(est_sold_list_sek, 0),
        "restocks":                restocks,
        "returns":                 returns_count,
        "restock_events":          restock_events_list,
        "return_events":           return_events_list,
        "by_category":             by_category,
        "by_brand":                by_brand,
    }

    return summary, detail_rows, new_snapshot


# ── Excel writer ──────────────────────────────────────────────────────────────

_HDR_FILL = PatternFill("solid", fgColor="1F497D")
_HDR_FONT = Font(bold=True, color="FFFFFF")


def _write_headers(ws, headers: list[str]) -> None:
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = Alignment(horizontal="center")


def _autofit(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 55)


def write_excel(state: dict, detail_rows: list[dict]) -> None:
    if XLSX_PATH.exists():
        wb = load_workbook(XLSX_PATH)
    else:
        wb = Workbook()
        for name in list(wb.sheetnames):
            del wb[name]

    # Remove obsolete sheets if present.
    for _obs in ("By Market", "By Site"):
        if _obs in wb.sheetnames:
            del wb[_obs]

    all_entries = state.get("daily_summary", [])

    # ── Sheet 1: Daily Summary ────────────────────────────────────────────────
    ds_name = "Daily Summary"
    if ds_name in wb.sheetnames:
        del wb[ds_name]
    ws = wb.create_sheet(ds_name, 0)
    _write_headers(ws, [
        "Date",
        "Est. Sales (sell price SEK)",
        "Est. Sales (list price SEK)",
        "Est. Returns (units)",
        "Restocks",
    ])
    for entry in all_entries:
        s = entry.get("summary", {})
        ws.append([
            entry.get("date", ""),
            s.get("est_sold_today_sek",    0),
            s.get("est_sold_today_list_sek", 0),
            s.get("returns",               0),
            s.get("restocks",              0),
        ])
    _autofit(ws)

    # ── Sheet 2: By Category (wide, rebuilt each run) ────────────────────────
    cat_name = "By Category"
    if cat_name in wb.sheetnames:
        del wb[cat_name]
    ws_c = wb.create_sheet(cat_name)

    # Collect all category names across all history.
    all_cats: set[str] = set()
    for entry in all_entries:
        all_cats.update(entry.get("summary", {}).get("by_category", {}).keys())

    # Sort categories left-to-right by the most recent day's sell revenue (desc).
    latest_by_cat = all_entries[-1].get("summary", {}).get("by_category", {}) if all_entries else {}
    sorted_cats = sorted(
        all_cats,
        key=lambda c: -latest_by_cat.get(c, {}).get("sell_rev_sek", 0),
    )

    # Columns: all categories (Sell) sorted by latest sell revenue desc,
    # then all categories (List) in the same category order.
    cat_cols = [f"{c} (Sell)" for c in sorted_cats] + [f"{c} (List)" for c in sorted_cats]
    _write_headers(ws_c, ["Date"] + cat_cols)
    for entry in all_entries:
        by_cat = entry.get("summary", {}).get("by_category", {})
        row: list = [entry.get("date", "")]
        for c in sorted_cats:
            row.append(round(by_cat.get(c, {}).get("sell_rev_sek", 0), 0))
        for c in sorted_cats:
            row.append(round(by_cat.get(c, {}).get("list_rev_sek", 0), 0))
        ws_c.append(row)
    _autofit(ws_c)

    # ── Sheet 3: By Brand (wide, rebuilt each run) ────────────────────────────
    brand_name = "By Brand"
    if brand_name in wb.sheetnames:
        del wb[brand_name]
    ws_b = wb.create_sheet(brand_name)

    # Collect all brand names across all history.
    all_brands: set[str] = set()
    for entry in all_entries:
        all_brands.update(entry.get("summary", {}).get("by_brand", {}).keys())

    # Sort brands left-to-right by the most recent day's sell revenue (desc).
    # Column order updates automatically each run as relative brand sizes shift.
    latest_by_br = all_entries[-1].get("summary", {}).get("by_brand", {}) if all_entries else {}
    sorted_brands = sorted(
        all_brands,
        key=lambda b: -latest_by_br.get(b, {}).get("sell_rev_sek", 0),
    )

    # Columns: all brands (Sell) sorted by latest sell revenue desc,
    # then all brands (List) in the same brand order.
    brand_cols = [f"{b} (Sell)" for b in sorted_brands] + [f"{b} (List)" for b in sorted_brands]
    _write_headers(ws_b, ["Date"] + brand_cols)

    for entry in all_entries:
        by_br = entry.get("summary", {}).get("by_brand", {})
        row: list = [entry.get("date", "")]
        for b in sorted_brands:
            row.append(round(by_br.get(b, {}).get("sell_rev_sek", 0), 0))
        for b in sorted_brands:
            row.append(round(by_br.get(b, {}).get("list_rev_sek", 0), 0))
        ws_b.append(row)
    _autofit(ws_b)

    # ── Sheet 4: Restocks (rebuilt each run, one row per event) ───────────────
    restock_sheet = "Restocks"
    if restock_sheet in wb.sheetnames:
        del wb[restock_sheet]
    ws_r = wb.create_sheet(restock_sheet)
    _write_headers(ws_r, [
        "Date", "Site", "Product Key", "Brand", "Title", "Category",
        "Stock Before", "Stock After", "Delta", "Sell Price (SEK)",
        "Est. Restock Value (SEK)",
    ])
    for entry in all_entries:
        d = entry.get("date", "")
        for ev in entry.get("summary", {}).get("restock_events", []):
            delta     = ev.get("delta", 0)
            sell_price = ev.get("sell_price_sek", 0)
            ws_r.append([
                d,
                ev.get("site", ""),
                ev.get("key", ""),
                ev.get("brand", ""),
                ev.get("title", ""),
                ev.get("category", ""),
                ev.get("stock_before", 0),
                ev.get("stock_after", 0),
                delta,
                round(sell_price, 0),
                round(delta * sell_price, 0),
            ])
    _autofit(ws_r)

    # ── Sheet 5: Returns Detail (rebuilt each run, one row per return event) ──
    ret_sheet = "Returns Detail"
    if ret_sheet in wb.sheetnames:
        del wb[ret_sheet]
    ws_ret = wb.create_sheet(ret_sheet)
    _write_headers(ws_ret, [
        "Date", "Site", "Product Key", "Brand", "Title", "Category",
        "Stock Before", "Stock After", "Units Returned", "Sell Price (SEK)",
        "Est. Return Value (SEK)",
    ])
    for entry in all_entries:
        d = entry.get("date", "")
        for ev in entry.get("summary", {}).get("return_events", []):
            delta      = ev.get("delta", 0)
            sell_price = ev.get("sell_price_sek", 0)
            ws_ret.append([
                d,
                ev.get("site", ""),
                ev.get("key", ""),
                ev.get("brand", ""),
                ev.get("title", ""),
                ev.get("category", ""),
                ev.get("stock_before", 0),
                ev.get("stock_after", 0),
                delta,
                round(sell_price, 0),
                round(delta * sell_price, 0),
            ])
    _autofit(ws_ret)

    # ── Sheet 6: Latest Detail (replaced each run) ────────────────────────────
    det_sheet = "Latest Detail"
    if det_sheet in wb.sheetnames:
        del wb[det_sheet]
    ws_d = wb.create_sheet(det_sheet)
    _write_headers(ws_d, [
        "Product-Colour Key", "Site", "Brand", "Title", "Category",
        "Sell Price (SEK)", "List Price (SEK)", "Historic Low (SEK)",
        "Discount %", "New?", "Est. Sold Today",
        # Primary-market stock (single shared pool per site)
        "Stock W_SE", "Stock M_SE",
        # Nelly availability markets (1 = listed, 0 = not listed, blank = N/A)
        "Avail W_NO", "Avail W_DK", "Avail W_FI", "Avail W_NL",
        "Avail W_DE", "Avail W_BE", "Avail W_PL", "Avail W_FR", "Avail W_AT",
        # NlyMan availability markets
        "Avail M_NO", "Avail M_DK", "Avail M_FI", "Avail M_NL",
        "Avail M_DE", "Avail M_BE", "Avail M_PL", "Avail M_FR", "Avail M_AT",
        "Total Stock", "Markets Listed",
    ])
    for row in sorted(detail_rows, key=lambda x: -x["primary_stock"]):
        ws_d.append([
            row["key"], row["site"], row["brand"], row["title"], row["category"],
            row["sell_price_sek"], row["list_price_sek"],
            row.get("historic_low_sek", 0),
            row["discount_pct"], "Yes" if row["is_new"] else "",
            row["est_sold_today"],
            row.get("stk_W_SE", ""), row.get("stk_M_SE", ""),
            # Nelly avail
            row.get("avl_W_NO", ""), row.get("avl_W_DK", ""),
            row.get("avl_W_FI", ""), row.get("avl_W_NL", ""),
            row.get("avl_W_DE", ""), row.get("avl_W_BE", ""),
            row.get("avl_W_PL", ""), row.get("avl_W_FR", ""),
            row.get("avl_W_AT", ""),
            # NlyMan avail
            row.get("avl_M_NO", ""), row.get("avl_M_DK", ""),
            row.get("avl_M_FI", ""), row.get("avl_M_NL", ""),
            row.get("avl_M_DE", ""), row.get("avl_M_BE", ""),
            row.get("avl_M_PL", ""), row.get("avl_M_FR", ""),
            row.get("avl_M_AT", ""),
            row["primary_stock"], row["listed_count"],
        ])
    _autofit(ws_d)

    wb.save(XLSX_PATH)
    print(f"  Saved -> {XLSX_PATH}")


# ── Main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    today = date.today().isoformat()
    now   = datetime.now().isoformat(timespec="seconds")

    print(f"[{now}] Nelly inventory tracker  ·  stock-delta methodology")
    print("=" * 60)

    state = load_state()

    # Guard: skip if already ran today
    if state.get("daily_summary") and state["daily_summary"][-1]["date"] == today:
        print(f"Already ran today ({today}).  Delete the last entry in daily_summary "
              f"from {STATE_FILE.name} to re-run.")
        return

    # ── Step 1: Use hardcoded cluster ID ─────────────────────────────────────
    cluster_id = ELEVATE_CLUSTER_ID
    state["cluster_id"] = cluster_id

    print(f"\nCluster ID : {cluster_id}")
    for mkt, cfg in MARKETS.items():
        print(f"  [{mkt}] market={cfg['elevate_market']} locale={cfg['locale']} "
              f"categories={len(cfg['categories'])}")

    # ── Step 2: Fetch inventory from all markets ─────────────────────────────
    print("\nFetching inventory across all markets and categories ...")
    curr_by_market = fetch_all_by_market(cluster_id)

    total_raw     = sum(len(v) for v in curr_by_market.values())
    unique_keys   = len({k for mv in curr_by_market.values() for k in mv})
    print(f"\nTotal fetched: {total_raw:,} raw  |  {unique_keys:,} unique product-colours")

    if total_raw == 0:
        print("No products fetched — check cluster ID and category page references.")
        return

    # ── Step 4: Compute summary and delta estimates ──────────────────────────
    print("\nComputing inventory summary and stock-delta estimates ...")
    last_snapshot = state.get("last_snapshot") or {}
    is_first_run  = not last_snapshot

    summary, detail_rows, new_snapshot = compute_snapshot_summary(
        curr_by_market, last_snapshot
    )

    if is_first_run:
        print("  (First run — no prior snapshot, stock-delta estimates are zero.)")

    print(f"  Unique product-colours         : {summary['total_products']:,}")
    print(f"  Est. sold today (units)        : {summary['est_sold_today_units']:,}")
    print(f"  Est. sold today (SEK)          : {summary['est_sold_today_sek']:,.0f}")
    print(f"  Est. sold at list price (SEK)  : {summary['est_sold_today_list_sek']:,.0f}")
    print(f"  Est. customer returns (units)  : {summary.get('returns', 0):,}")
    print(f"  Restocks (≥{RESTOCK_MIN_UNITS} units)              : {summary['restocks']:,}")

    print("\n  Top 10 categories by est. revenue (SEK):")
    top_cats = sorted(
        summary.get("by_category", {}).items(),
        key=lambda x: -x[1].get("sell_rev_sek", 0)
    )[:10]
    for cat_name, cdata in top_cats:
        print(f"    [{cat_name}] {cdata.get('sell_rev_sek', 0):,.0f} SEK")

    # ── Step 5: Update state ─────────────────────────────────────────────────
    if not isinstance(state.get("daily_summary"), list):
        state["daily_summary"] = []

    state["daily_summary"].append({
        "date":      today,
        "timestamp": now,
        "summary":   summary,
    })
    state["last_snapshot"] = new_snapshot

    save_state(state)
    print(f"\n  State saved -> {STATE_FILE.name}")

    # ── Step 6: Write Excel ─────────────────────────────────────────────────
    print("Writing Excel ...")
    write_excel(state, detail_rows)

    print("\nDone.")


if __name__ == "__main__":
    main()
