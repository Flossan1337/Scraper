# fetch_fractal_trends.py
# Skriver ALLTID om hela historiken i Excel (ingen append).

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

# Inte pytrends: se TrendReq-kommentaren i core/trends.py. pytrends bygger en
# ny session per request, och Google svarar 429 på en sessions första request,
# så under pytrends misslyckas varje anrop. Retrier och paus ligger numera i
# core.trends - inga sleeps behövs här.
from core.trends import TrendReq, fetch_series, write_trends_to_db
from excel_utils import append_df

# ── DEFINE YOUR GROUPS ──
# "Fractal North" upprepas i varje grupp som ankare: alla grupper hämtas i
# var sitt anrop, och den gemensamma termen låter dem skalas mot varandra.
GROUPS = [
    ["Fractal North", "Fractal Define", "Fractal Core", "Fractal Node", "Fractal Meshify"],
    ["Fractal North", "Fractal Focus", "Fractal Vector", "Fractal Era", "Fractal Torrent"],
    ["Fractal North", "Fractal Pop", "Fractal Ridge", "Fractal Terra", "Fractal Mood"],
    ["Fractal North", "Fractal Epoch", "Fractal Refine", "Fractal Scape"],
]

# ── OUTPUT ──
# Se till att alltid skriva till Scraper/data (en nivå upp från denna fil)
REPO_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = REPO_ROOT / "data"
DATA_DIR.mkdir(exist_ok=True)

XLSX_PATH = DATA_DIR / "fractal_trends_monthly.xlsx"
SHEET_NAME = "fractal_trends_monthly"


def main():
    # En klient för hela körningen - den delade uppvärmda sessionen är poängen.
    py = TrendReq(hl="en-US", tz=120)
    tf = f"2016-01-01 {datetime.now():%Y-%m-%d}"

    master = None
    for i, grp in enumerate(GROUPS, start=1):
        print(f"Kör grupp {i}/{len(GROUPS)}: {grp}")
        df = fetch_series(grp, timeframe=tf, geo="", client=py)
        df = df.drop(columns=["isPartial"], errors="ignore").resample("ME").mean()

        if master is None:
            master = df
        else:
            # Ta bort "Fractal North" så den inte dupliceras vid join
            df = df.drop(columns=["Fractal North"], errors="ignore")
            master = master.join(df, how="outer")

    # Sortera datum och gör om indexet till kolumn "Date"
    out = master.sort_index().reset_index().rename(columns={"date": "Date"})
    out["Date"] = pd.to_datetime(out["Date"])

    # Ta bort gammal flik (om finns) så vi skriver om HELA historiken
    if XLSX_PATH.exists():
        try:
            wb = load_workbook(XLSX_PATH)
            if SHEET_NAME in wb.sheetnames:
                ws = wb[SHEET_NAME]
                wb.remove(ws)
                wb.save(XLSX_PATH)
        except Exception as e:
            print(f"Varning: kunde inte ta bort gammal flik ({SHEET_NAME}): {e}")

    # Skriv hela datasetet på nytt
    append_df(str(XLSX_PATH), SHEET_NAME, out)
    print(f"Skrev om hela historiken ({len(out)} rader) till {XLSX_PATH} [{SHEET_NAME}].")

    db_rows, db_error = write_trends_to_db("fractal", {"default": out})
    if db_error is not None:
        print(f"Databas: MISSLYCKADES - {db_error}")
    else:
        print(f"Databas: {db_rows} rader uppserta")


if __name__ == "__main__":
    main()
