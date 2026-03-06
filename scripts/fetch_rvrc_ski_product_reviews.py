#!/usr/bin/env python3
"""
fetch_rvrc_ski_product_reviews.py

Fetches monthly review counts for RVRC's ski / winter products (Jan 2025 – Mar 2026)
and computes their share of the total PANTS and JACKETS reviews per month.

No sub-category exists in the API; we identify seasonal products by matching
"Ski" as a whole word in the product display-name (English), then exclude
skirt / legging items from the pants group.

Products tracked
----------------
  Ski Pants       : PANTS category, name has word "Ski", not a skirt/legging
  Ski Jackets     : JACKETS category, name has word "Ski"
  (treated as one combined Ski/Winter-jacket group as agreed)

Methodology
-----------
1. Resolve product universe via facets + displayName name-matching.
2. For each target product, paginate reviews newest-first and stop once
   we pass the start of the observation window.  These products have at most
   a few hundred reviews each, so fetching is fast.
3. Bucket by YYYY-MM.
4. Load monthly PANTS / JACKETS totals from the existing history JSON
   (produced by backfill_revolutionrace_history.py).
5. Compute share (%) and write Excel.

Output
------
  data/rvrc_ski_products_monthly.xlsx
    Sheet "SkiProducts_Monthly"  – one row per month
    Sheet "ProductReference"     – which base-product IDs map to each group
"""

import ast
import json
import re
import time
from collections import defaultdict
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Config ─────────────────────────────────────────────────────────────────────
GRAPHQL_URL = "https://reviews.revolutionrace.com/revolutionrace/graphql"
GQL_HEADERS = {
    "Content-Type": "application/json",
    "User-Agent":   "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
}

SCRIPT_DIR   = Path(__file__).resolve().parent
HISTORY_FILE = (SCRIPT_DIR / ".." / "data" / "revolutionrace_monthly_history.json").resolve()
XLSX_PATH    = (SCRIPT_DIR / ".." / "data" / "rvrc_ski_products_monthly.xlsx").resolve()

START_MONTH = "2025-01"
END_MONTH   = "2026-03"

TAKE       = 5000   # max page size the server accepts
DELAY_S    = 0.15   # polite pause between requests
BATCH_SIZE = 50     # products per alias round-trip


# ── GraphQL helper ─────────────────────────────────────────────────────────────

def gql(query: str, retries: int = 3) -> dict:
    """POST a GraphQL query; retry with exponential back-off on failure."""
    for attempt in range(retries):
        try:
            resp = requests.post(
                GRAPHQL_URL, json={"query": query},
                headers=GQL_HEADERS, timeout=30,
            )
            resp.raise_for_status()
            data = resp.json()
            if "errors" in data:
                raise ValueError(data["errors"][0]["message"])
            return data
        except Exception:
            if attempt == retries - 1:
                raise
            time.sleep(2 ** attempt)


# ── Product discovery ──────────────────────────────────────────────────────────

def _resolve_display_names(base_products: list[str]) -> dict[str, str | None]:
    """
    Batch-resolve English display names for a list of base-product IDs.
    Returns {base_product_id: name_or_None}.
    Prefers en-US, then en-GB, then sv-SE, then any locale.
    """
    results: dict[str, str | None] = {}
    for i in range(0, len(base_products), BATCH_SIZE):
        chunk = base_products[i : i + BATCH_SIZE]
        aliases = "\n".join(
            f'  p{bp}: publicReviews(filter: {{ item_baseProduct: ["{bp}"] }}, take: 1) {{'
            f'    hits {{ item {{ displayName }} }}'
            f'  }}'
            for bp in chunk
        )
        data = gql(f"{{ {aliases} }}")
        for bp in chunk:
            hits = data["data"].get(f"p{bp}", {}).get("hits", [])
            if not hits:
                results[bp] = None
                continue
            raw = (hits[0].get("item") or {}).get("displayName") or ""
            if not raw:
                results[bp] = None
                continue
            try:
                cleaned = raw.replace(": null", ": None").replace(":null", ": None")
                name_dict = ast.literal_eval(cleaned)
                name = (
                    name_dict.get("en-US")
                    or name_dict.get("en-GB")
                    or name_dict.get("sv-SE")
                    or next(iter(name_dict.values()), None)
                )
            except Exception:
                name = raw
            results[bp] = name
        time.sleep(DELAY_S)
    return results


def _is_ski_word(name: str) -> bool:
    """True if "Ski" appears as a whole word (avoids matching "Skirt", "Skiing" etc.)."""
    return bool(re.search(r"\bski\b", name, re.IGNORECASE))


def _is_skirt_or_legging(name: str) -> bool:
    """Exclude non-pants items that may contain 'ski' as a substring."""
    nl = name.lower()
    return "skirt" in nl or "legging" in nl or "skort" in nl


def discover_target_products() -> tuple[dict[str, str], dict[str, str]]:
    """
    Returns:
      ski_pants_products  : {base_product_id: product_name}
      ski_jacket_products : {base_product_id: product_name}
    """
    bp_to_category: dict[str, str] = {}
    for category in ["PANTS", "JACKETS"]:
        q = f"""
        {{
          publicReviews(
            take: 0,
            filter: {{ item_parentItemCategory: ["{category}"] }},
            facets: {{ item_baseProduct: [] }}
          ) {{
            facets {{ item_baseProduct {{ value }} }}
          }}
        }}
        """
        data = gql(q)
        for item in data["data"]["publicReviews"]["facets"]["item_baseProduct"]:
            bp_to_category[item["value"]] = category
        print(f"  {category}: {len([k for k,v in bp_to_category.items() if v == category])} base products")
        time.sleep(DELAY_S)

    all_bps = list(bp_to_category.keys())
    print(f"  Resolving names for {len(all_bps)} products (this takes ~{len(all_bps)//BATCH_SIZE + 1} requests)...")
    names = _resolve_display_names(all_bps)

    ski_pants: dict[str, str]   = {}
    ski_jackets: dict[str, str] = {}

    for bp, category in bp_to_category.items():
        name = names.get(bp) or ""
        if not name or not _is_ski_word(name):
            continue
        if category == "PANTS":
            if not _is_skirt_or_legging(name):
                ski_pants[bp] = name
        elif category == "JACKETS":
            ski_jackets[bp] = name

    return ski_pants, ski_jackets


# ── Review fetching ────────────────────────────────────────────────────────────

def fetch_months_in_window(bp: str) -> list[str]:
    """
    Paginate reviews for a single base product (newest first).
    Returns a list of YYYY-MM strings, one entry per review published in
    [START_MONTH, END_MONTH].  Stops paging as soon as the oldest review on a
    page predates START_MONTH (no need to go back further).
    """
    months: list[str] = []
    skip = 0
    while True:
        q = f"""
        {{
          publicReviews(
            take: {TAKE}, skip: {skip},
            filter: {{ item_baseProduct: ["{bp}"] }},
            sort: {{ order: [publishedAt_desc] }}
          ) {{ hits {{ publishedAt }} }}
        }}
        """
        data = gql(q)
        hits = data["data"]["publicReviews"]["hits"]
        if not hits:
            break

        oldest_on_page: str | None = None
        for h in hits:
            month = h["publishedAt"][:7]  # YYYY-MM
            oldest_on_page = month        # desc order → last hit is oldest
            if START_MONTH <= month <= END_MONTH:
                months.append(month)

        # Early exit: all remaining reviews are older than our window
        if oldest_on_page and oldest_on_page < START_MONTH:
            break
        skip += TAKE
        time.sleep(DELAY_S)

    return months


# ── Excel helpers ──────────────────────────────────────────────────────────────

_HDR_FILL = PatternFill("solid", fgColor="1F4E79")
_HDR_FONT = Font(color="FFFFFF", bold=True)

def _style_header_row(ws) -> None:
    for cell in ws[1]:
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = Alignment(horizontal="center")

def _autofit(ws) -> None:
    for col in ws.columns:
        length = max(len(str(cell.value or "")) for cell in col) + 2
        ws.column_dimensions[get_column_letter(col[0].column)].width = length


# ── Main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    print("RVRC Ski Product Reviews – Monthly Share Analysis")
    print(f"Period: {START_MONTH} → {END_MONTH}")
    print("=" * 60)

    # 1. Discover target products ───────────────────────────────────────────────
    print("\n[1/4]  Discovering Ski Pants and Ski Jacket products...")
    ski_pants, ski_jackets = discover_target_products()

    print(f"\n  Ski Pants   ({len(ski_pants)} base products):")
    for bp, name in sorted(ski_pants.items(), key=lambda x: x[1]):
        print(f"    {bp:8s}  {name}")

    print(f"\n  Ski/Winter Jackets  ({len(ski_jackets)} base products):")
    for bp, name in sorted(ski_jackets.items(), key=lambda x: x[1]):
        print(f"    {bp:8s}  {name}")

    if not ski_pants and not ski_jackets:
        print("\n  No ski products found – exiting.")
        return

    # 2. Fetch reviews in window ────────────────────────────────────────────────
    print("\n[2/4]  Fetching reviews for Ski Pants products...")
    pants_monthly: dict[str, int] = defaultdict(int)
    for bp, name in ski_pants.items():
        months = fetch_months_in_window(bp)
        for m in months:
            pants_monthly[m] += 1
        print(f"    {bp:8s}  {name:<50s}  {len(months)} reviews in window")

    print("\n        Fetching reviews for Ski Jacket products...")
    jacket_monthly: dict[str, int] = defaultdict(int)
    for bp, name in ski_jackets.items():
        months = fetch_months_in_window(bp)
        for m in months:
            jacket_monthly[m] += 1
        print(f"    {bp:8s}  {name:<50s}  {len(months)} reviews in window")

    # 3. Load category totals from history ─────────────────────────────────────
    print("\n[3/4]  Loading monthly PANTS / JACKETS totals from history JSON...")
    history = json.loads(HISTORY_FILE.read_text(encoding="utf-8"))
    cat_totals: dict[str, dict[str, int]] = {}
    for month, d in history.items():
        if START_MONTH <= month <= END_MONTH:
            cats = d.get("by_category", {})
            cat_totals[month] = {
                "PANTS":   cats.get("PANTS", 0),
                "JACKETS": cats.get("JACKETS", 0),
            }
    print(f"    Loaded {len(cat_totals)} months from {HISTORY_FILE.name}")

    # 4. Build summary and write Excel ─────────────────────────────────────────
    print("\n[4/4]  Building summary table and writing Excel...")

    all_months = sorted(
        set(cat_totals.keys())
        | set(pants_monthly.keys())
        | set(jacket_monthly.keys())
    )

    print(f"\n  {'Month':<8}  {'SkiPnts':>8}  {'TotPnts':>8}  {'Pnts%':>6}  "
          f"{'SkiJckt':>8}  {'TotJckt':>8}  {'Jckt%':>6}")
    print("  " + "-" * 63)

    rows = []
    for month in all_months:
        ski_p   = pants_monthly.get(month, 0)
        tot_p   = cat_totals.get(month, {}).get("PANTS", 0)
        ski_j   = jacket_monthly.get(month, 0)
        tot_j   = cat_totals.get(month, {}).get("JACKETS", 0)
        share_p = round(ski_p / tot_p * 100, 2) if tot_p else None
        share_j = round(ski_j / tot_j * 100, 2) if tot_j else None
        rows.append((month, ski_p, tot_p, share_p, ski_j, tot_j, share_j))

        sp_str = f"{share_p:.2f}%" if share_p is not None else "n/a"
        sj_str = f"{share_j:.2f}%" if share_j is not None else "n/a"
        print(f"  {month}  {ski_p:>8,}  {tot_p:>8,}  {sp_str:>6}  "
              f"{ski_j:>8,}  {tot_j:>8,}  {sj_str:>6}")

    # Write Excel ───────────────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "SkiProducts_Monthly"

    ws.append([
        "Month",
        "Ski Pants Reviews",
        "Total Pants Reviews",
        "Ski Pants Share (%)",
        "Ski/Winter Jacket Reviews",
        "Total Jacket Reviews",
        "Ski/Winter Jacket Share (%)",
    ])
    for row in rows:
        ws.append(list(row))

    _style_header_row(ws)
    _autofit(ws)

    # Product reference sheet
    ws_ref = wb.create_sheet("ProductReference")
    ws_ref.append(["Group", "Base Product ID", "Product Name"])
    for bp, name in sorted(ski_pants.items()):
        ws_ref.append(["Ski Pants", bp, name])
    for bp, name in sorted(ski_jackets.items()):
        ws_ref.append(["Ski/Winter Jacket", bp, name])
    _style_header_row(ws_ref)
    _autofit(ws_ref)

    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(XLSX_PATH)

    print(f"\n  Wrote {len(rows)} months to {XLSX_PATH}")
    print(f"  Columns: {[c.value for c in ws[1]]}")
    print("\nDone.")


if __name__ == "__main__":
    main()
