#!/usr/bin/env python3
"""
backfill_revolutionrace_history.py  (one-time run)

Fetches individual review records from the RVRC GraphQL API to build a
month-by-month histogram of review publication dates, going back as far
as the API allows.

Strategy
--------
- Products with ≤9,000 reviews:  paginate ascending (oldest first) → full history.
- Products with  >9,000 reviews:  paginate descending (newest first) up to 9,000
                                   reviews → recent history (most useful for
                                   understanding current run-rate).

Fields fetched per review: publishedAt, channelId, item.parentItemCategory,
item.gender  — so the histogram can be sliced by country/category/gender.

Output
------
  data/revolutionrace_monthly_history.json  – raw monthly buckets saved to disk
  data/revolutionrace_reviews.xlsx          – new sheet "MonthlyHistory" appended

Runtime: ~6–10 min on a single machine (1,055 requests, throttled to 3 req/s
         to avoid hammering the server).
"""

import json
import time
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import requests
from openpyxl import load_workbook, Workbook

# ── Config ─────────────────────────────────────────────────────────────────────
GRAPHQL_URL = "https://reviews.revolutionrace.com/revolutionrace/graphql"
GQL_HEADERS = {
    "Content-Type": "application/json",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
}
MAX_SKIP       = 9000     # hard server limit
TAKE           = 5000     # max page size that works
WORKERS        = 4        # parallel product fetches (be gentle)
DELAY_S        = 0.1      # per-request delay inside each worker

CHANNEL_COUNTRIES = {
    "76302142-cd49-4c57-a48e-9217cf41c8b5": "DE",
    "3963b20d-4d89-4ddb-92dc-d0c897dc149a": "SE",
    "4b897b07-7369-4665-abf3-ca895d31ecf3": "FI",
    "89853c2f-41a4-46db-b1ae-8b9977f1e4ec": "DK",
    "206916a9-9467-4fda-a02e-19a9307aec04": "PL",
    "977bb77c-8cb2-443c-bbd0-fad62f3c46b5": "NL",
    "65963f84-c285-4df7-9f6a-5876288db01c": "AT",
    "d0b1658e-ba32-4283-847f-aa8accd33fbf": "UK",
    "0f5ea666-e2bf-4c57-9ff4-6b849652f42d": "CZ",
    "719beded-b644-46be-98cc-4f35b90c0263": "NO",
    "03bbe5f0-b12f-406b-8817-eedb455cbfa2": "FR",
    "30bd1080-c713-49ac-8762-993cb31fae61": "CH",
    "c3b6bb57-4b88-4a2c-977d-29aa97a0d89d": "IT",
    "9c9dfccb-2138-4357-99da-2cec70102091": "IE",
}

SCRIPT_DIR    = Path(__file__).resolve().parent
STATE_FILE    = (SCRIPT_DIR / ".." / "data" / "revolutionrace_state.json").resolve()
HISTORY_FILE  = (SCRIPT_DIR / ".." / "data" / "revolutionrace_monthly_history.json").resolve()
XLSX_PATH     = (SCRIPT_DIR / ".." / "data" / "revolutionrace_reviews.xlsx").resolve()


# ── GraphQL helper ─────────────────────────────────────────────────────────────

def gql(query: str, retries: int = 3) -> dict:
    for attempt in range(retries):
        try:
            resp = requests.post(
                GRAPHQL_URL, json={"query": query},
                headers=GQL_HEADERS, timeout=30,
            )
            resp.raise_for_status()
            data = resp.json()
            if "errors" in data:
                raise ValueError(data["errors"][0]["message"])
            return data
        except Exception as e:
            if attempt == retries - 1:
                raise
            time.sleep(2 ** attempt)


# ── Per-product fetcher ────────────────────────────────────────────────────────

REVIEW_FIELDS = "publishedAt channelId item { parentItemCategory gender }"

def fetch_reviews_for_product(base_product: str, total_count: int) -> list[dict]:
    """
    Returns a list of review dicts {publishedAt, channelId, category, gender}.
    For products > MAX_SKIP reviews, fetches the NEWEST reviews (descending).
    For smaller products, fetches ALL reviews (ascending).
    """
    big = total_count > MAX_SKIP
    sort_dir = "publishedAt_desc" if big else "publishedAt_asc"
    max_to_fetch = MAX_SKIP if big else total_count

    reviews = []
    skip = 0
    while skip < max_to_fetch:
        take = min(TAKE, max_to_fetch - skip)
        q = f"""{{
          publicReviews(
            take: {take}, skip: {skip},
            filter: {{ item_baseProduct: ["{base_product}"] }},
            sort: {{ order: [{sort_dir}] }}
          ) {{ hits {{ {REVIEW_FIELDS} }} }}
        }}"""
        data = gql(q)
        hits = data["data"]["publicReviews"]["hits"]
        if not hits:
            break
        for h in hits:
            reviews.append({
                "month":    h["publishedAt"][:7],   # YYYY-MM
                "channel":  h.get("channelId", ""),
                "category": (h.get("item") or {}).get("parentItemCategory", ""),
                "gender":   (h.get("item") or {}).get("gender", ""),
                "is_big":   big,
            })
        skip += take
        time.sleep(DELAY_S)

    return reviews


# ── Aggregation ────────────────────────────────────────────────────────────────

def aggregate(all_reviews: list[dict]) -> dict:
    """
    Returns {YYYY-MM: {total, by_country, by_category, by_gender}}.
    """
    monthly: dict[str, dict] = defaultdict(lambda: {
        "total": 0,
        "total_excl_big": 0,
        "by_country":  defaultdict(int),
        "by_category": defaultdict(int),
        "by_gender":   defaultdict(int),
    })
    for r in all_reviews:
        m = r["month"]
        monthly[m]["total"] += 1
        if not r.get("is_big"):
            monthly[m]["total_excl_big"] += 1
        country = CHANNEL_COUNTRIES.get(r["channel"], "other")
        monthly[m]["by_country"][country]  += 1
        monthly[m]["by_category"][r["category"] or "unknown"] += 1
        monthly[m]["by_gender"][r["gender"]   or "unknown"] += 1

    # Convert defaultdicts to plain dicts for JSON serialisation
    return {
        month: {
            "total":          d["total"],
            "total_excl_big": d["total_excl_big"],
            "by_country":     dict(d["by_country"]),
            "by_category":    dict(d["by_category"]),
            "by_gender":      dict(d["by_gender"]),
        }
        for month, d in sorted(monthly.items())
    }


# ── Excel writer ───────────────────────────────────────────────────────────────

TRACKED_COUNTRIES  = ["DE","SE","FI","DK","PL","NL","AT","UK","CZ","NO","FR","CH","IT","IE","other"]
TRACKED_CATEGORIES = ["PANTS","TOPS","JACKETS","ACCS","FOOTWEAR","BASELAYERS","BAGS"]
TRACKED_GENDERS    = ["MEN","WOMEN","UNISEX"]

def write_monthly_sheet(monthly: dict) -> None:
    """Write / replace the 'MonthlyHistory' sheet in the existing Excel file."""
    if XLSX_PATH.exists():
        wb = load_workbook(XLSX_PATH)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    if "MonthlyHistory" in wb.sheetnames:
        del wb["MonthlyHistory"]
    ws = wb.create_sheet("MonthlyHistory")

    # Build column list
    cols = ["month", "total", "total_excl_big"]
    for c in TRACKED_COUNTRIES:
        cols.append(f"country_{c}")
    for cat in TRACKED_CATEGORIES:
        cols.append(f"cat_{cat}")
    cols.append("cat_other")
    for g in TRACKED_GENDERS:
        cols.append(f"gender_{g}")
    cols.append("gender_other")

    ws.append(cols)
    for month, d in sorted(monthly.items()):
        row = [month, d["total"], d.get("total_excl_big", 0)]
        for c in TRACKED_COUNTRIES:
            row.append(d["by_country"].get(c, 0))
        for cat in TRACKED_CATEGORIES:
            row.append(d["by_category"].get(cat, 0))
        other_cat = sum(v for k, v in d["by_category"].items() if k not in TRACKED_CATEGORIES)
        row.append(other_cat)
        for g in TRACKED_GENDERS:
            row.append(d["by_gender"].get(g, 0))
        other_gen = sum(v for k, v in d["by_gender"].items() if k not in TRACKED_GENDERS)
        row.append(other_gen)
        ws.append(row)

    wb.save(XLSX_PATH)
    print(f"  MonthlyHistory sheet written: {len(monthly)} months, {sum(d['total'] for d in monthly.values()):,} reviews")


# ── Main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    print("Revolution Race – historical review backfill")

    state    = json.loads(STATE_FILE.read_text(encoding="utf-8"))
    products = state["products"]

    # Get latest review count per product
    latest_count: dict[str, int] = {}
    for bp, p in products.items():
        if p["counts"]:
            latest_count[bp] = max(e["count"] for e in p["counts"])

    small = {bp: c for bp, c in latest_count.items() if c <= MAX_SKIP}
    big   = {bp: c for bp, c in latest_count.items() if c >  MAX_SKIP}

    total_products = len(small) + len(big)
    print(f"  {len(small)} products <={MAX_SKIP} reviews (full history)")
    print(f"  {len(big)} products >{MAX_SKIP} reviews (newest {MAX_SKIP} only)")
    print(f"  Total: {total_products} products to fetch")
    print()

    all_reviews: list[dict] = []
    done = 0
    errors = 0

    # Process in parallel with a thread pool
    tasks = list(latest_count.items())
    with ThreadPoolExecutor(max_workers=WORKERS) as pool:
        futures = {
            pool.submit(fetch_reviews_for_product, bp, count): (bp, count)
            for bp, count in tasks
        }
        for future in as_completed(futures):
            bp, count = futures[future]
            done += 1
            try:
                reviews = future.result()
                all_reviews.extend(reviews)
                strategy = "newest" if count > MAX_SKIP else "all"
                if done % 50 == 0 or done == total_products:
                    print(f"  [{done:>4}/{total_products}]  {len(all_reviews):>8,} reviews collected so far")
            except Exception as e:
                errors += 1
                print(f"  [{done:>4}/{total_products}]  ERROR for {bp}: {e}")

    print(f"\n  Fetch complete: {len(all_reviews):,} reviews, {errors} errors")

    # Aggregate
    print("  Aggregating by month...")
    monthly = aggregate(all_reviews)

    # Show summary
    print(f"  Monthly buckets: {min(monthly)} -> {max(monthly)}")
    recent = sorted(monthly.items())[-12:]
    print(f"  Last 12 months:")
    for month, d in recent:
        bar = "#" * min(40, d["total"] // 20)
        print(f"    {month}  {d['total']:>5,}  {bar}")

    # Save JSON
    HISTORY_FILE.parent.mkdir(parents=True, exist_ok=True)
    HISTORY_FILE.write_text(
        json.dumps(monthly, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    print(f"\n  Saved: {HISTORY_FILE}")

    # Write Excel sheet
    write_monthly_sheet(monthly)
    print(f"  Saved: {XLSX_PATH}  (MonthlyHistory sheet)")


if __name__ == "__main__":
    main()
