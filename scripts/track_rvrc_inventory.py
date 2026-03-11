#!/usr/bin/env python3
"""
track_rvrc_inventory.py

Tracks RevolutionRace inventory at product-colour level across multiple markets
(SE, DE, NO, UK, COM) using the Voyado Elevate API directly.

Methodology
-----------
Each daily run:
  1. Queries the Voyado Elevate storefront API for every top-level product
     category (clothing, accessories, shoes) in each market, paginating with
     limit=600 until all products are retrieved.
  2. Passes the `presentCustom` parameter (the full RVRC E1 attribute list)
     to unlock the `sale_last_week` and `sale_last_days` custom fields that
     the RVRC website uses for its own ranking/merchandising.
  3. Extracts per-variant `sale_last_week`, sellingPrice, and listPrice.
  4. Groups variants by product-colour key (e.g. "10004_2001") and reads
     `sale_last_week` — a global 7-day sales counter maintained by RVRC's
     OMS and exposed through Elevate (same value in every market, every size).
  5. Estimated daily units per product-colour = sale_last_week / 7.
     Revenue is denominated in EUR using DE market prices (EUR directly).

Why sale_last_week instead of stock-delta
-----------------------------------------
The `stockNumber` field in the Elevate API is an Available-to-Promise (ATP)
metric that changes due to warehouse transfers, reservation changes, batch
inventory adjustments, and actual customer sales.  Day-over-day deltas are
dominated by these non-sale events and grossly over-estimate actual orders
(observed ~7–20× inflation vs. RVRC's financial scale).

The `sale_last_week` field is RVRC's own sales counter — accurate, stable,
and globally consistent across all markets.  It requires the `presentCustom`
URL parameter (discovered via the RVRC website's JS bundle).

How the data is fetched
-----------------------
The RVRC site uses Voyado Elevate (Apptus) as its product search/catalog backend.
The storefront API is publicly accessible without authentication — customerKey
and sessionKey are random UUIDs generated fresh each run.

Endpoint: GET https://{cluster}.api.esales.apptus.cloud/api/storefront/v3/queries/landing-page
Required params: pageReference, presentCustom (pipe-separated E1 attribute list)
Optional params: limit (max 600), skip (pagination offset), market, locale

The response includes variant.custom.sale_last_week = total units sold globally
across all markets in the trailing 7 days (per product-colour combination).

Pagination uses skip (offset by productGroup count) until skip >= totalHits.
Typically 5–15 total requests cover all 5 markets, completing in ~30 seconds.

State file  : data/rvrc_inventory_state.json
Excel output: data/rvrc_inventory.xlsx
"""

import json
import time
import uuid
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Elevate API configuration ──────────────────────────────────────────────────
ELEVATE_CLUSTER_ID = "wA4BFC9F5"
ELEVATE_BASE_URL   = f"https://{ELEVATE_CLUSTER_ID}.api.esales.apptus.cloud"
ELEVATE_ENDPOINT   = "/api/storefront/v3/queries/landing-page"
ELEVATE_LIMIT      = 600  # max allowed by API (1000 returns 400 error)

# Custom attributes to request from Elevate (mirrors the E1 list in RVRC's
# website JS bundle — BMqHAaDt.js, imported as `c0` / `xo` and passed as
# `presentCustom` in the landingPage query in DXwsVHXO.js).
# Passing this param unlocks variant.custom.sale_last_week and
# variant.custom.sale_last_days — RVRC's own OMS sales counters.
ELEVATE_PRESENT_CUSTOM = (
    "allcategories|color_name|features_name|fit_name|gender_name|isdead|"
    "labeltext_name|googleanalyticsname|ratingcount|totalcolorcount|blobtext_name|"
    "sustainability|categorybreadcrumb|color|features|fit|gender|labeltext|blobtext|"
    "variant.campaign_id|variant.campaign_type|variant.campaign_has_conditions|"
    "variant.sale_last_days|variant.sale_last_week|variant.length_variant|variant.new_in"
)

# Top-level Elevate category pageReference values covering all RVRC products.
# clothing ~1,700 products, accessories ~180, shoes ~70 (SE market; similar for others).
ELEVATE_CATEGORIES = ["clothing", "accessories", "shoes"]

# ── Market configuration ───────────────────────────────────────────────────────
# Each market is fetched to maximise product-colour coverage and obtain local
# pricing.  NOTE: sale_last_week is a GLOBAL counter — the same value is
# returned in every market for the same product-colour.  compute_daily_from_slw()
# deduplicates by base_key with MARKET_PRIORITY (DE first so EUR prices are used
# by default for the majority of product-colour entries).
MARKETS: dict[str, dict] = {
    "SE":  {"elevate_market": "SE",  "locale": "sv-SE",  "currency": "SEK"},
    "DE":  {"elevate_market": "DE",  "locale": "de-DE",  "currency": "EUR"},
    "NO":  {"elevate_market": "NO",  "locale": "nb-NO",  "currency": "NOK"},
    "UK":  {"elevate_market": "UK",  "locale": "en-GB",  "currency": "GBP"},
    "COM": {"elevate_market": "EU",  "locale": "en-001", "currency": "EUR"},
}

# FX fallback rates (local currency → SEK) used if the Frankfurter API is down
FX_FALLBACKS: dict[str, float] = {
    "SEK": 1.0,
    "EUR": 11.5,
    "NOK": 0.97,
    "GBP": 13.0,
}

SCRIPT_DIR = Path(__file__).resolve().parent
STATE_FILE = (SCRIPT_DIR / ".." / "data" / "rvrc_inventory_state.json").resolve()
XLSX_PATH  = (SCRIPT_DIR / ".." / "data" / "rvrc_inventory.xlsx").resolve()

# ── State I/O ──────────────────────────────────────────────────────────────────

def load_state() -> dict:
    if STATE_FILE.exists():
        raw = json.loads(STATE_FILE.read_text(encoding="utf-8-sig"))
        raw.pop("last_snapshot", None)  # No longer used — drop to reduce file size
        return raw
    return {"daily_sales": []}


def save_state(state: dict) -> None:
    STATE_FILE.write_text(
        json.dumps(state, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


# ── FX rates ───────────────────────────────────────────────────────────────────

def fetch_fx_rates() -> dict[str, float]:
    """
    Fetch current exchange rates to SEK for all non-SEK market currencies via
    the Frankfurter API (free, ECB-sourced).  Falls back to FX_FALLBACKS.

    Returns {currency_code: sek_per_unit}, e.g. {"SEK": 1.0, "EUR": 11.5, ...}
    """
    non_sek = sorted({cfg["currency"] for cfg in MARKETS.values() if cfg["currency"] != "SEK"})
    rates: dict[str, float] = {"SEK": 1.0}
    if not non_sek:
        return rates
    try:
        resp = requests.get(
            f"https://api.frankfurter.app/latest?from=EUR&to=SEK,{','.join(non_sek)}",
            timeout=10,
        )
        resp.raise_for_status()
        data = resp.json()["rates"]
        eur_sek = float(data["SEK"])
        rates["EUR"] = eur_sek
        for ccy in non_sek:
            if ccy == "EUR":
                continue
            if ccy in data:
                # Cross-rate: 1 ccy = (EUR/SEK) / (EUR/ccy) SEK
                rates[ccy] = eur_sek / float(data[ccy])
            else:
                rates[ccy] = FX_FALLBACKS.get(ccy, 1.0)
                print(f"  [FX] {ccy} not in API response — using fallback {rates[ccy]}")
        print(f"  [FX] {rates}")
    except Exception as e:
        print(f"  [FX] API error ({e}) — using fallbacks")
        for ccy in non_sek:
            rates[ccy] = FX_FALLBACKS.get(ccy, 1.0)
    return rates


# ── Elevate API helpers ────────────────────────────────────────────────────────

def _get_price(price_raw) -> float:
    """Parse a price that may be a float or {'min': x, 'max': x} dict."""
    if isinstance(price_raw, dict):
        try:
            return float(price_raw.get("min") or 0.0)
        except (TypeError, ValueError):
            return 0.0
    try:
        return float(price_raw) if price_raw is not None else 0.0
    except (TypeError, ValueError):
        return 0.0


# Ordered keyword → display-name mapping for _parse_category().
# Matches against the second segment of the Elevate categorybreadcrumb id,
# e.g. "Men>Jackets>Waterproof Jackets" → "Jackets" → "Jackets & Vests".
_CATEGORY_KEYWORDS: list[tuple[str, str]] = [
    ("jacket",      "Jackets & Vests"),
    ("vest",        "Jackets & Vests"),
    ("trouser",     "Trousers & Pants"),
    ("pant",        "Trousers & Pants"),
    ("tight",       "Trousers & Pants"),
    ("fleece",      "Fleece & Midlayer"),
    ("midlayer",    "Fleece & Midlayer"),
    ("sweater",     "Fleece & Midlayer"),
    ("base layer",  "Base Layer"),
    ("baselayer",   "Base Layer"),
    ("t-shirt",     "T-Shirts & Tops"),
    ("shirt",       "T-Shirts & Tops"),
    ("top",         "T-Shirts & Tops"),
    ("hoodie",      "Hoodies & Sweatshirts"),
    ("sweatshirt",  "Hoodies & Sweatshirts"),
    ("shoe",        "Shoes"),
    ("boot",        "Shoes"),
    ("sock",        "Accessories"),
    ("cap",         "Accessories"),
    ("hat",         "Accessories"),
    ("glove",       "Accessories"),
    ("bag",         "Accessories"),
    ("backpack",    "Accessories"),
    ("accessori",   "Accessories"),
]


def _parse_category(breadcrumb_id: str, page_ref: str) -> str:
    """
    Derive a clean product category label from an Elevate categorybreadcrumb id
    (e.g. "Men>Jackets>Waterproof Jackets") or fall back to the top-level page
    reference ("clothing", "accessories", "shoes").
    """
    if breadcrumb_id:
        parts = [p.strip() for p in breadcrumb_id.split(">")]
        raw   = parts[1] if len(parts) >= 2 else parts[0]
        lower = raw.lower()
        for kw, label in _CATEGORY_KEYWORDS:
            if kw in lower:
                return label
        return raw  # preserve unrecognised breadcrumb segment as-is
    return {"clothing": "Clothing", "accessories": "Accessories", "shoes": "Shoes"}.get(
        page_ref, page_ref.capitalize()
    )


def fetch_elevate_page(
    elevate_market: str, locale: str, page_ref: str,
    skip: int = 0, customer_key: str = "", session_key: str = "",
) -> Optional[dict]:
    """Fetch one page from the Elevate landing-page API."""
    params = {
        "market":        elevate_market,
        "locale":        locale,
        "customerKey":   customer_key,
        "sessionKey":    session_key,
        "touchpoint":    "desktop",
        "pageReference": page_ref,
        "limit":         ELEVATE_LIMIT,
        "skip":          skip,
        "presentCustom": ELEVATE_PRESENT_CUSTOM,
    }
    try:
        resp = requests.get(
            ELEVATE_BASE_URL + ELEVATE_ENDPOINT,
            params=params,
            timeout=(10, 60),
        )
        resp.raise_for_status()
        return resp.json()
    except requests.RequestException as exc:
        print(f"    [WARN] Elevate {page_ref} skip={skip}: {exc}")
        return None


def extract_variants_from_elevate(data: dict, page_ref: str = "") -> tuple[dict[str, dict], int, int]:
    """
    Extract variant data from an Elevate landing-page response.

    Returns
    -------
    (variants_dict, total_hits, group_count)
    variants_dict: {variant_key: {"stock", "sale_last_week", "sale_last_days",
                                   "sell_price", "list_price", "title", "size", "category"}}
    total_hits:    totalHits from the API response (used for pagination)
    group_count:   number of productGroups in this page (pagination offset increment)
    """
    variants: dict[str, dict] = {}
    pl = data.get("primaryList") or {}
    total_hits = int(pl.get("totalHits") or 0)
    groups = pl.get("productGroups") or []
    for group in groups:
        for product in group.get("products") or []:
            title  = str(product.get("title") or product.get("name") or "")
            p_sell = _get_price(product.get("sellingPrice"))
            p_list = _get_price(product.get("listPrice")) or p_sell
            # Derive product category from Elevate categorybreadcrumb attribute
            custom      = product.get("custom") or {}
            breadcrumbs = custom.get("categorybreadcrumb") or []
            bc_id       = breadcrumbs[0].get("id", "") if breadcrumbs else ""
            category    = _parse_category(bc_id, page_ref)
            for variant in product.get("variants") or []:
                key = variant.get("key")
                if not key or not isinstance(key, str):
                    continue
                try:
                    stock = int(variant.get("stockNumber") or 0)
                except (TypeError, ValueError):
                    stock = 0
                sell_price = _get_price(variant.get("sellingPrice")) or p_sell
                list_price = _get_price(variant.get("listPrice")) or p_list or sell_price
                size = str(variant.get("size") or variant.get("label") or "")
                v_custom = variant.get("custom") or {}

                slw_list = v_custom.get("sale_last_week") or []
                sale_last_week = 0
                try:
                    if slw_list:
                        sale_last_week = int(slw_list[0].get("label") or 0)
                except (TypeError, ValueError):
                    pass

                sld_list = v_custom.get("sale_last_days") or []
                sale_last_days = 0
                try:
                    if sld_list:
                        sale_last_days = int(sld_list[0].get("label") or 0)
                except (TypeError, ValueError):
                    pass

                variants[key] = {
                    "stock":          stock,
                    "sale_last_week": sale_last_week,
                    "sale_last_days": sale_last_days,
                    "sell_price":     sell_price,
                    "list_price":     list_price,
                    "title":          title,
                    "size":           size,
                    "category":       category,
                }
    return variants, total_hits, len(groups)


# Minimum expected unique variants per market.  If a market returns fewer than
# this after fetching all categories, a loud warning is printed.  The Elevate
# API typically returns ~13,000–16,000 variants per market; 10,000 is a safe
# lower threshold that catches silent failures.
MIN_EXPECTED_VARIANTS: dict[str, int] = {
    "SE":  10000,
    "DE":  10000,
    "NO":  10000,
    "UK":  10000,
    "COM": 10000,
}


def fetch_all_by_market() -> dict[str, dict[str, dict]]:
    """
    Fetch variants for every market using the Voyado Elevate API directly.
    ~5–15 total requests across all markets (<30s) vs old ~600 requests (~40min).

    Returns
    -------
    {market_code: {variant_key: {"stock", "sell_price", "list_price", "title", "size"}}}
    """
    result: dict[str, dict[str, dict]] = {}

    for market_code, cfg in MARKETS.items():
        elevate_market = cfg["elevate_market"]
        locale         = cfg["locale"]
        customer_key   = str(uuid.uuid4())
        session_key    = str(uuid.uuid4())
        print(f"\n[{market_code}] Elevate API (market={elevate_market}, locale={locale})")

        market_variants: dict[str, dict] = {}

        for cat in ELEVATE_CATEGORIES:
            skip, page = 0, 1
            while True:
                data = fetch_elevate_page(
                    elevate_market, locale, cat, skip, customer_key, session_key
                )
                if data is None:
                    break
                new_v, total_hits, pg_count = extract_variants_from_elevate(data, cat)
                market_variants.update(new_v)
                print(
                    f"  {cat} p{page}: +{len(new_v)} variants "
                    f"[{skip}–{skip + pg_count}/{total_hits}]"
                )
                skip += pg_count
                if skip >= total_hits or pg_count == 0:
                    break
                page += 1

        variant_count = len(market_variants)
        min_expected  = MIN_EXPECTED_VARIANTS.get(market_code, 5000)
        status = "OK" if variant_count >= min_expected else "LOW"
        print(f"  [{market_code}] {variant_count:,} unique variants | status: {status}")
        if status == "LOW":
            print(
                f"  *** WARNING [{market_code}]: Only {variant_count:,} variants "
                f"(expected >= {min_expected:,}). Possible API change or fetch failure. ***"
            )

        result[market_code] = market_variants

    return result


# ── Sales estimation via sale_last_week ───────────────────────────────────────

MARKET_PRIORITY = ["DE", "SE", "NO", "UK", "COM"]


def compute_daily_from_slw(
    curr_by_market: dict[str, dict[str, dict]],
    fx_rates: dict[str, float],
) -> tuple[list[dict], dict, float, float, float, int]:
    """
    Estimate daily sales using the sale_last_week custom field from Elevate.

    sale_last_week is a global 7-day unit sales counter maintained by RVRC's
    OMS.  It is identical in every market request for the same product-colour,
    and shared across all size variants of that colour.

    Algorithm
    ---------
    - Group variants by base_key = product_id + "_" + colour_id (key without
      the trailing "-SIZE" suffix, e.g. "10004_2001" from "10004_2001-M").
    - For each unique base_key: daily_estimate = max(sale_last_week) / 7  (max
      across sizes is a safety measure; values should be identical).
    - Revenue = daily_estimate × price_eur, where price_eur comes from the DE
      market if available (EUR directly), otherwise converted from local currency.
    - Each base_key is counted exactly once (MARKET_PRIORITY = DE first).

    Returns
    -------
    per_product_color : list of dicts, one per active product-colour
    by_category       : {category: {units, revenue_sell_eur, revenue_list_eur, product_colors}}
    total_units       : total estimated daily units (float)
    total_rev_sell    : total estimated daily sell revenue in EUR
    total_rev_list    : total estimated daily list revenue in EUR
    n_product_colors  : count of product-colours with sale_last_week > 0
    """
    eur_sek = fx_rates.get("EUR", 11.5)
    counted_bases: set[str] = set()
    per_product_color: list[dict] = []
    by_category: dict[str, dict] = {}

    for market_code in MARKET_PRIORITY:
        if market_code not in curr_by_market:
            continue
        curr_variants = curr_by_market[market_code]
        currency  = MARKETS[market_code]["currency"]
        fx_to_eur = fx_rates.get(currency, 1.0) / eur_sek

        # Group variant keys by base_key (product_id_colourId, no size suffix)
        base_groups: dict[str, list] = {}
        for key, v in curr_variants.items():
            base = key.rsplit("-", 1)[0]
            base_groups.setdefault(base, []).append((key, v))

        for base_key, items in base_groups.items():
            if base_key in counted_bases:
                continue
            counted_bases.add(base_key)

            # All sizes share the same sale_last_week; take max for safety
            slw = max((v.get("sale_last_week", 0) for _, v in items), default=0)
            if slw <= 0:
                continue

            daily_units = slw / 7.0

            # Representative variant: prefer one with a positive sell_price
            rep_v = next((v for _, v in items if v.get("sell_price", 0) > 0), items[0][1])
            sell_price_eur = rep_v["sell_price"] * fx_to_eur
            list_price_raw = rep_v.get("list_price") or rep_v["sell_price"]
            list_price_eur = list_price_raw * fx_to_eur
            discount_pct = (
                round(100.0 * (1.0 - rep_v["sell_price"] / list_price_raw), 1)
                if list_price_raw > 0 and rep_v["sell_price"] < list_price_raw else 0.0
            )
            category = rep_v.get("category", "Other")
            title    = rep_v.get("title", "")
            rev_sell = daily_units * sell_price_eur
            rev_list = daily_units * list_price_eur

            per_product_color.append({
                "key":              base_key,
                "category":         category,
                "title":            title,
                "sale_last_week":   slw,
                "daily_estimate":   round(daily_units, 1),
                "sell_price_eur":   round(sell_price_eur, 2),
                "list_price_eur":   round(list_price_eur, 2),
                "sell_revenue_eur": round(rev_sell, 2),
                "list_revenue_eur": round(rev_list, 2),
                "discount_pct":     discount_pct,
            })

            cat = by_category.setdefault(category, {
                "units": 0.0, "revenue_sell_eur": 0.0,
                "revenue_list_eur": 0.0, "product_colors": 0,
            })
            cat["units"]            += daily_units
            cat["revenue_sell_eur"] += rev_sell
            cat["revenue_list_eur"] += rev_list
            cat["product_colors"]   += 1

    for c in by_category.values():
        c["units"]            = round(c["units"], 1)
        c["revenue_sell_eur"] = round(c["revenue_sell_eur"], 2)
        c["revenue_list_eur"] = round(c["revenue_list_eur"], 2)

    total_units    = round(sum(v["daily_estimate"] for v in per_product_color), 1)
    total_rev_sell = sum(v["sell_revenue_eur"] for v in per_product_color)
    total_rev_list = sum(v["list_revenue_eur"] for v in per_product_color)
    return (
        per_product_color, by_category,
        total_units, total_rev_sell, total_rev_list,
        len(per_product_color),
    )




_HDR_FILL = PatternFill("solid", fgColor="1F497D")
_HDR_FONT = Font(bold=True, color="FFFFFF")


def _write_headers(ws, headers: list[str]) -> None:
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = Alignment(horizontal="center")


def _autofit(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 55)


def write_excel(state: dict, per_product_color_today: list[dict]) -> None:
    if XLSX_PATH.exists():
        wb = load_workbook(XLSX_PATH)
    else:
        wb = Workbook()
        for name in list(wb.sheetnames):
            del wb[name]

    today_row   = state["daily_sales"][-1] if state["daily_sales"] else {}
    fx_snapshot = today_row.get("fx_rates", {})

    # ── Sheet 1: Daily Summary (one row appended per run) ────────────────────
    summary_name = "Daily Summary"
    if summary_name in wb.sheetnames:
        ws_sum = wb[summary_name]
    else:
        ws_sum = wb.create_sheet(summary_name)
        _write_headers(ws_sum, [
            "Date",
            "Est. Daily Units (slw/7)",
            "Est. Daily Rev Sell (EUR)",
            "Est. Daily Rev List (EUR)",
            "Avg Discount %",
            "Product-Colors Active",
            "EUR/SEK",
        ])

    rev_sell = today_row.get("estimated_revenue_sell_eur", 0.0)
    rev_list = today_row.get("estimated_revenue_list_eur", 0.0)
    avg_disc = (
        round(100.0 * (1.0 - rev_sell / rev_list), 1)
        if rev_list > 0 else 0.0
    )
    ws_sum.append([
        today_row.get("date", ""),
        today_row.get("estimated_units_daily", 0),
        round(rev_sell, 0),
        round(rev_list, 0),
        avg_disc,
        today_row.get("product_colors_active", 0),
        fx_snapshot.get("EUR", ""),
    ])
    _autofit(ws_sum)

    # ── Sheet 2: By Category (one row per category per run, appended) ────────
    by_cat_name = "By Category"
    if by_cat_name in wb.sheetnames:
        ws_cat = wb[by_cat_name]
    else:
        ws_cat = wb.create_sheet(by_cat_name)
        _write_headers(ws_cat, [
            "Date",
            "Category",
            "Est. Daily Units",
            "Daily Rev Sell (EUR)",
            "Daily Rev List (EUR)",
            "Product-Colors",
        ])

    for cat_name, cdata in sorted(today_row.get("by_category", {}).items()):
        ws_cat.append([
            today_row.get("date", ""),
            cat_name,
            cdata.get("units", 0),
            round(cdata.get("revenue_sell_eur", 0.0), 0),
            round(cdata.get("revenue_list_eur", 0.0), 0),
            cdata.get("product_colors", 0),
        ])
    _autofit(ws_cat)

    # ── Sheet 3: Latest Detail (replaced each run) ────────────────────────────
    detail_name = "Latest Detail"
    if detail_name in wb.sheetnames:
        del wb[detail_name]
    ws_det = wb.create_sheet(detail_name)
    _write_headers(ws_det, [
        "Product-Color Key",
        "Product Title",
        "Category",
        "Sale Last Week (global)",
        "Daily Estimate (slw/7)",
        "Sell Price (EUR)",
        "List Price (EUR)",
        "Daily Rev Sell (EUR)",
        "Daily Rev List (EUR)",
        "Discount %",
    ])
    for row in sorted(per_product_color_today, key=lambda x: -x["sell_revenue_eur"]):
        ws_det.append([
            row["key"],
            row["title"],
            row["category"],
            row["sale_last_week"],
            row["daily_estimate"],
            row["sell_price_eur"],
            row["list_price_eur"],
            round(row["sell_revenue_eur"], 0),
            round(row["list_revenue_eur"], 0),
            row["discount_pct"],
        ])
    _autofit(ws_det)

    wb.save(XLSX_PATH)
    print(f"  Saved -> {XLSX_PATH}")


# ── Main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    today = date.today().isoformat()
    now   = datetime.now().isoformat(timespec="seconds")

    print(f"[{now}] RevolutionRace inventory tracker  ·  sale_last_week methodology")
    print("=" * 60)

    state = load_state()

    # Guard: skip if already ran today
    if state.get("daily_sales") and state["daily_sales"][-1]["date"] == today:
        print(f"Already ran today ({today}). Delete the last entry in daily_sales "
              f"from {STATE_FILE.name} to re-run.")
        return

    print("\nFetching live FX rates ...")
    fx_rates = fetch_fx_rates()

    print("\nFetching inventory across all markets and categories ...")
    curr_by_market = fetch_all_by_market()

    total_variants  = sum(len(v) for v in curr_by_market.values())
    unique_variants = len({k for m in curr_by_market.values() for k in m})
    print(f"\nTotal variants fetched across all markets: {total_variants:,} raw ({unique_variants:,} unique)")

    if total_variants == 0:
        print("No variants fetched — check URLs and network connectivity.")
        return

    print("\nComputing daily sales estimates from sale_last_week ...")
    per_product_color, by_category, total_units, total_rev_sell, total_rev_list, n_pc = (
        compute_daily_from_slw(curr_by_market, fx_rates)
    )
    print(f"  Est. daily units      : {total_units:,.1f}")
    print(f"  Est. sell revenue/day : {total_rev_sell:,.0f} EUR")
    print(f"  Est. list revenue/day : {total_rev_list:,.0f} EUR")
    print(f"  EUR/SEK               : {fx_rates.get('EUR', 'n/a')}")
    print(f"  Product-colours active: {n_pc:,}")
    for cat_name, cdata in sorted(by_category.items(), key=lambda x: -x[1]["revenue_sell_eur"]):
        print(f"  [{cat_name}] {cdata['units']:,.1f} units/day | "
              f"{cdata['revenue_sell_eur']:,.0f} EUR sell | "
              f"{cdata['revenue_list_eur']:,.0f} EUR list")

    if not state.get("daily_sales"):
        state["daily_sales"] = []
    state["daily_sales"].append({
        "date":                       today,
        "timestamp":                  now,
        "method":                     "sale_last_week",
        "estimated_units_daily":      round(total_units, 1),
        "estimated_revenue_sell_eur": round(total_rev_sell, 2),
        "estimated_revenue_list_eur": round(total_rev_list, 2),
        "product_colors_active":      n_pc,
        "fx_rates":                   {k: round(v, 4) for k, v in fx_rates.items()},
        "by_category":                by_category,
    })

    save_state(state)
    print(f"\n  State saved -> {STATE_FILE.name}")

    print("Writing Excel ...")
    write_excel(state, per_product_color)

    print("\nDone.")


if __name__ == "__main__":
    main()
