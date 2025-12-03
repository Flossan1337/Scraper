# track_fractal_rankings_playwright.py
from datetime import datetime
import re
import time
import random
from pathlib import Path
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeoutError

# NEW: Excel
from openpyxl import Workbook, load_workbook

# Notera: Newegg ändrar ofta URL-strukturen. Om scriptet slutar fungera, kontrollera dessa.
HEADSET_URL = "https://www.newegg.com/Gaming-Headsets/SubCategory/ID-3767?Order=3&View=96"
CHAIR_URL   = "https://www.newegg.com/Gaming-Chairs/SubCategory/ID-3628?Order=3&View=96"

HEADSET_PRODUCTS = {
    "Fractal Design Scape Dark RGB Wireless Gaming Headset": [
        "fractal design scape dark",
        "fd-hs-sca1-01"
    ],
    "Fractal Design Scape Light RGB Wireless Gaming Headset": [
        "fractal design scape light",
        "fd-hs-sca1-02"
    ],
}

CHAIR_PRODUCTS = {
    "Fractal Design Refine Gaming Chair (Fabric Dark)": [
        "fractal design refine fabric dark",
        "refine fabric dark"
    ],
    "Fractal Design Refine Gaming Chair (Mesh Dark)": [
        "fractal design refine mesh dark",
        "refine mesh dark"
    ],
    "Fractal Design Refine Gaming Chair (Fabric Light)": [
        "fractal design refine fabric light",
        "refine fabric light"
    ],
    "Fractal Design Refine Gaming Chair (Mesh Light)": [
        "fractal design refine mesh light",
        "refine mesh light"
    ],
}

SCRIPT_DIR = Path(__file__).resolve().parent
XLSX_PATH = (SCRIPT_DIR / ".." / "data" / "fractal_rankings.xlsx").resolve()

def canon(s: str) -> str:
    if not s: return ""
    s = s.lower()
    s = re.sub(r"[\s\-\(\)\[\],.:/®™]+", " ", s)
    return " ".join(s.split())

def wait_page_ready(page):
    """Väntar på att listan ska laddas."""
    try:
        page.wait_for_load_state("domcontentloaded", timeout=15000)
    except PWTimeoutError:
        pass
    
    # Vänta specifikt på att produktcellerna ska synas
    try:
        page.wait_for_selector(".item-cell", timeout=15000)
    except PWTimeoutError:
        print("[WARN] Could not find .item-cell selector immediately.")

def human_scroll(page, steps=5):
    """
    Scrollar mjukare för att trigga lazy loading och undvika bot-detektion.
    """
    for _ in range(steps):
        # Scrolla en slumpmässig mängd pixlar
        scroll_y = random.randint(400, 800)
        page.mouse.wheel(0, scroll_y)
        time.sleep(random.uniform(0.5, 1.5))

def get_items_from_cells(page):
    """
    Hämtar (href, title) genom att iterera över .item-cell.
    Detta är mer robust för att bestämma faktisk rankingposition.
    """
    items = []
    # Hämta alla celler (både list view och grid view använder item-cell oftast)
    cells = page.locator(".item-cell")
    count = cells.count()
    
    for i in range(count):
        cell = cells.nth(i)
        
        # Försök hitta titeln inuti cellen
        title_el = cell.locator("a.item-title")
        
        # Om ingen titel finns i cellen (kanske en annons eller tom plats), hoppa över
        if not title_el.is_visible():
            continue
            
        href = title_el.get_attribute("href") or ""
        text = title_el.inner_text() or ""
        
        if "/p/" in href: # Se till att det är en produktsida
            items.append((href.strip(), text.strip()))
            
    return items

def paginate_and_rank(page, url, targets_aliases, max_pages=3, debug_name=""):
    alias_map = {k: [canon(k)] + [canon(a) for a in v] for k, v in targets_aliases.items()}
    out = {k: "NA" for k in alias_map.keys()}

    print(f"--- Processing {debug_name} ---")
    page.goto(url)
    wait_page_ready(page)

    global_rank = 0
    page_idx = 0
    seen_hrefs_global = set()
    
    while page_idx < max_pages:
        page_idx += 1
        print(f"Scanning page {page_idx}...")

        # Scrolla för att ladda in items
        human_scroll(page, steps=8)

        # Hämta items baserat på faktisk cell-position
        items = get_items_from_cells(page)
        
        # Filtrera bort dubbletter som Playwright kanske ser om DOMen uppdateras konstigt,
        # men behåll ordningen för rankingens skull.
        new_items = []
        for h, t in items:
            if h not in seen_hrefs_global:
                seen_hrefs_global.add(h)
                new_items.append((h, t))
        
        if not new_items:
            print("[WARN] No new items found on this page.")
        
        for href, title in new_items:
            global_rank += 1
            ct = canon(title)
            
            # Check matches
            for canonical_name, aliases in alias_map.items():
                # Om vi redan hittat en rank, hoppa över (vi vill ha den högsta/första ranken)
                if out[canonical_name] != "NA" and isinstance(out[canonical_name], int):
                    continue
                
                # Jämför canonical text
                if any(a in ct for a in aliases):
                    print(f"   MATCH! Rank {global_rank}: {title[:30]}...")
                    out[canonical_name] = global_rank

        # Om vi hittat allt, bryt
        if all(isinstance(v, int) for v in out.values()):
            print("Found all targets.")
            break

        # Pagination logic
        next_btn = page.locator("button[aria-label='Next']").first
        # Fallback för andra typer av knappar
        if not next_btn.is_visible():
            next_btn = page.locator("a[aria-label='Next']").first
            
        if next_btn.is_visible() and next_btn.is_enabled():
            try:
                next_btn.click()
                time.sleep(3) # Vänta lite extra vid sidbyte
                wait_page_ready(page)
            except Exception as e:
                print(f"Error clicking next: {e}")
                break
        else:
            print("No next button found or reached end.")
            break

    return out

# ──────────────────────────────────────
# Excel helpers

def ensure_header_xlsx():
    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    if XLSX_PATH.exists():
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    header = ["Date"] + list(HEADSET_PRODUCTS.keys()) + list(CHAIR_PRODUCTS.keys())
    ws.append(header)
    wb.save(XLSX_PATH)

def append_row(all_ranks):
    header = ["Date"] + list(HEADSET_PRODUCTS.keys()) + list(CHAIR_PRODUCTS.keys())
    row = [datetime.now().strftime("%Y-%m-%d")] + [all_ranks.get(k, "NA") for k in header[1:]]
    
    if not XLSX_PATH.exists():
        ensure_header_xlsx()
        
    wb = load_workbook(XLSX_PATH)
    ws = wb.active
    ws.append(row)
    wb.save(XLSX_PATH)
    print(f"✓ Appended to {XLSX_PATH}")

def main():
    ensure_header_xlsx()
    with sync_playwright() as p:
        # VIKTIGT: Arguments för att undvika bot-detektion
        browser = p.chromium.launch(
            headless=False,  # Ändra till True när du verifierat att det funkar
            args=[
                "--disable-blink-features=AutomationControlled",
                "--start-maximized"
            ]
        )
        
        context = browser.new_context(
            viewport={"width": 1920, "height": 1080},
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        )
        
        # Scripting removal kan hjälpa mot vissa detection scripts
        context.add_init_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        
        page = context.new_page()

        headsets = paginate_and_rank(page, HEADSET_URL, HEADSET_PRODUCTS, max_pages=3, debug_name="Headsets")
        chairs   = paginate_and_rank(page, CHAIR_URL,   CHAIR_PRODUCTS,   max_pages=6, debug_name="Chairs")

        all_ranks = {}
        all_ranks.update(headsets)
        all_ranks.update(chairs)

        append_row(all_ranks)
        print("Final Rankings:", all_ranks)

        context.close()
        browser.close()

if __name__ == "__main__":
    main()