import asyncio
import re
import os
from datetime import date
from pathlib import Path
from playwright.async_api import async_playwright, TimeoutError as PWTimeout
import openpyxl

# ── KONFIGURATION ──────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).resolve().parent
DATA_DIR = (SCRIPT_DIR / ".." / "data").resolve()
DATA_DIR.mkdir(parents=True, exist_ok=True)
XLSX_PATH = str((DATA_DIR / "fractal_rank_tracking.xlsx").resolve())
SHEET_NAME = "Rankings_Simple"

# Här definierar vi exakt de 4 punkter du vill hämta
TARGETS = [
    # --- USA ---
    {
        "excel_name": "Scape US",
        "asin": "B0D5HK6JRS",       # Scape Light
        "domain": "amazon.com",
        "locale": "en-US",
        "rank_label": "Best Sellers Rank"
    },
    {
        "excel_name": "Refine US",
        "asin": "B0CSYWWRSV",       # Refine Fabric Dark
        "domain": "amazon.com",
        "locale": "en-US",
        "rank_label": "Best Sellers Rank"
    },
    # --- TYSKLAND (DE) ---
    {
        "excel_name": "Scape DE",
        "asin": "B0D5HK6JRS",       # Scape Light (Samma ASIN brukar funka)
        "domain": "amazon.de",
        "locale": "de-DE",          # Viktigt för att tolka siffror (punkt vs komma)
        "rank_label": "Bestseller-Rang" # Vad det heter på tyska
    },
    {
        "excel_name": "Refine DE",
        "asin": "B0CSYWWRSV",       # Refine Fabric Dark
        "domain": "amazon.de",
        "locale": "de-DE",
        "rank_label": "Bestseller-Rang"
    },
]

HEADLESS = True 

# ───────────────────────────────────────────────────────────────────────────

def parse_rank_number(text: str) -> int:
    """
    Hittar första heltalet som följer efter ett '#' tecken.
    Hanterar både "1,500" (US) och "1.500" (DE).
    """
    # Regex: Hitta '#' följt av siffror, punkter eller kommatecken
    match = re.search(r"#([0-9.,]+)", text)
    if match:
        raw_num = match.group(1)
        # Ta bort både punkter och kommatecken för att få en ren integer
        # (Detta funkar så länge rankingen är ett heltal, vilket den alltid är)
        clean_num = raw_num.replace(",", "").replace(".", "")
        return int(clean_num)
    return 0

async def handle_blockers(page, domain):
    """Klickar bort Cookies (DE) och Upsells (US/DE)"""
    # 1. Cookies (Viktigast för Tyskland)
    try:
        if await page.locator("#sp-cc-accept").is_visible(timeout=2000):
            await page.click("#sp-cc-accept")
            await page.wait_for_timeout(500)
    except:
        pass

    # 2. "Fortsätt handla" popups
    upsell_texts = ['text="Continue shopping"', 'text="Weiter shoppen"', 'text="Weiter einkaufen"']
    for txt in upsell_texts:
        try:
            if await page.locator(txt).is_visible(timeout=1000):
                await page.click(txt)
                await page.wait_for_timeout(1000)
                return
        except:
            continue

async def get_rank(context, item) -> int:
    page = await context.new_page()
    url = f"https://www.{item['domain']}/dp/{item['asin']}?th=1&psc=1"
    print(f"  Fetching {item['excel_name']} ({item['domain']})...")

    try:
        await page.goto(url, wait_until="domcontentloaded", timeout=30000)
        await handle_blockers(page, item['domain'])

        # Vi letar efter tabellraden som innehåller rätt etikett (Label)
        # T.ex. "Best Sellers Rank" eller "Bestseller-Rang"
        label = item['rank_label']
        
        # Sökstrategi 1: Tabell (Standard på Desktop)
        # Vi letar efter en rad (tr) som innehåller labeln
        rank_text = ""
        
        # Locator: Hitta en rad som innehåller vår label
        row_locator = page.locator(f"tr:has-text('{label}')").first
        
        if await row_locator.count() > 0:
            rank_text = await row_locator.inner_text()
        else:
            # Sökstrategi 2: Punktlista (Alternativ layout / Mobil)
            li_locator = page.locator(f"li:has-text('{label}')").first
            if await li_locator.count() > 0:
                rank_text = await li_locator.inner_text()

        # Extrahera siffran
        rank = parse_rank_number(rank_text)

        if rank > 0:
            print(f"    ✅ Rank: #{rank}")
        else:
            print(f"    ⚠️ Kunde inte hitta rank (Leta efter '{label}')")
            # await page.screenshot(path=f"error_{item['excel_name']}.png")
        
        return rank

    except Exception as e:
        print(f"    ❌ Error: {e}")
        return 0
    finally:
        await page.close()

def append_to_excel(data_dict):
    file_exists = os.path.exists(XLSX_PATH)
    headers = ["Date"] + [t["excel_name"] for t in TARGETS]

    if not file_exists:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(headers)
    else:
        wb = openpyxl.load_workbook(XLSX_PATH)
        if SHEET_NAME in wb.sheetnames:
            ws = wb[SHEET_NAME]
        else:
            ws = wb.create_sheet(SHEET_NAME)
            ws.append(headers)

    row = [data_dict["Date"]]
    for t in TARGETS:
        row.append(data_dict.get(t["excel_name"], 0))
    
    ws.append(row)
    wb.save(XLSX_PATH)
    print(f"💾 Data saved to {XLSX_PATH}")

async def run():
    today = date.today().isoformat()
    results = {"Date": today}

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=HEADLESS)
        
        print(f"--- Starting Simple Rank Scraping ({today}) ---")

        for item in TARGETS:
            # Vi skapar en ny context för varje marknad för att få rätt Locale (språkinställning)
            context = await browser.new_context(
                locale=item['locale'], # sätter en-US eller de-DE
                viewport={"width": 1920, "height": 1080},
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
            )
            
            rank = await get_rank(context, item)
            results[item['excel_name']] = rank
            await context.close()
        
        await browser.close()

    append_to_excel(results)

if __name__ == "__main__":
    asyncio.run(run())