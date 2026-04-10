#!/usr/bin/env python3
"""
track_tu_brands.py

Tracks new brands appearing on the Technische Unie brands overview page
(https://www.c.technischeunie.nl/merken-overzicht.html).

Methodology
-----------
Each daily run:
  1. Fetches the brands overview page and extracts:
       a. Full A-Z brand list  – display names from  zoeken?brand=  links (~1 600 brands)
       b. Featured brand slugs – brands that have a dedicated sub-page at
          /merken-overzicht/<slug>.html  (~58 brands)
  2. Compares both lists against the stored state to detect new entries.
  3. For each new brand, attempts to match it to a featured-brand slug and,
     if found, fetches that dedicated page to collect product sub-categories.
  4. If new brands are found, appends one row to the Excel file:
       Column A : run date  (ISO-8601, e.g. 2026-04-10)
       Column B+: one column per new brand – cell value is the brand name
                  optionally followed by a newline and its product categories,
                  e.g.  "ABB\nInstallatiekasten, Noodverlichting, EV-laadsystemen"
  5. Saves the updated state so the next run only reports genuinely new entries.

First-run behaviour
-------------------
On the very first run the state file does not exist.  The script populates it
with the current brand snapshot and exits without writing to Excel, because
every brand would otherwise look "new".

State file  : data/tu_brands_state.json
Excel output: data/tu_brands.xlsx
"""

import json
import os
import re
import tempfile
import time
import urllib.parse
from datetime import date
from pathlib import Path

import requests
from openpyxl import Workbook, load_workbook

# ── Configuration ──────────────────────────────────────────────────────────────
BRANDS_URL      = "https://www.c.technischeunie.nl/merken-overzicht.html"
BRAND_PAGE_TPL  = "https://www.c.technischeunie.nl/merken-overzicht/{slug}.html"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )
}

REQUEST_TIMEOUT  = 30   # seconds
CATEGORY_DELAY_S = 0.3  # pause between brand-page requests

SCRIPT_DIR = Path(__file__).resolve().parent
STATE_FILE = (SCRIPT_DIR / ".." / "data" / "tu_brands_state.json").resolve()
XLSX_PATH  = (SCRIPT_DIR / ".." / "data" / "tu_brands.xlsx").resolve()

# ── HTTP helpers ───────────────────────────────────────────────────────────────

def fetch_html(url: str) -> str:
    resp = requests.get(url, headers=HEADERS, timeout=REQUEST_TIMEOUT)
    resp.raise_for_status()
    return resp.text


# ── Brand extraction ───────────────────────────────────────────────────────────

def extract_all_brands(html: str) -> set[str]:
    """Return all brand display-names embedded in zoeken?brand= links."""
    raw = re.findall(r'zoeken\?brand=([^"&\s]+)', html)
    return {urllib.parse.unquote_plus(b) for b in raw}


def extract_featured_slugs(html: str) -> set[str]:
    """Return slugs of brands that have a dedicated /merken-overzicht/<slug>.html page."""
    return set(re.findall(r'/merken-overzicht/([A-Za-z0-9][A-Za-z0-9-]+)\.html', html))


def get_brand_categories(slug: str) -> list[str]:
    """
    Fetch a featured brand's dedicated page and return its product sub-categories
    as human-readable strings (e.g. ['Installatiekasten', 'Noodverlichting']).
    Returns an empty list if the page is unavailable or has no sub-categories.
    """
    url = BRAND_PAGE_TPL.format(slug=slug)
    try:
        html = fetch_html(url)
        cat_slugs = re.findall(
            rf'/merken-overzicht/{re.escape(slug)}/([a-z0-9][a-z0-9-]+)\.html',
            html,
        )
        if not cat_slugs:
            return []
        return [c.replace("-", " ").title() for c in sorted(set(cat_slugs))]
    except Exception:
        return []


def brand_name_to_slug(name: str) -> str:
    """
    Convert a brand display-name to its most likely URL slug.
    e.g. 'Schneider Electric' → 'schneider-electric'
         'ABB Busch-Jaeger'   → 'abb-busch-jaeger'
    """
    slug = name.lower()
    slug = re.sub(r'[^a-z0-9]+', '-', slug)
    return slug.strip('-')


# ── State I/O ──────────────────────────────────────────────────────────────────

def load_state() -> dict:
    if STATE_FILE.exists():
        with open(STATE_FILE, encoding="utf-8") as fh:
            return json.load(fh)
    return {}   # empty dict signals first run


def save_state(brands: set[str], featured_slugs: set[str]) -> None:
    STATE_FILE.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "known_brands":         sorted(brands),
        "known_featured_slugs": sorted(featured_slugs),
    }
    with open(STATE_FILE, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, ensure_ascii=False, indent=2)


# ── Excel output ───────────────────────────────────────────────────────────────

def _extend_headers(ws, needed_brand_cols: int) -> None:
    """Ensure the header row has enough 'Brand N' columns."""
    header_row = ws[1]
    current_cols = len(header_row)
    for i in range(current_cols, needed_brand_cols + 1):  # col index 1-based
        ws.cell(row=1, column=i + 1).value = f"Brand {i}"


def append_to_excel(run_date: str, new_brands_data: list[tuple[str, list[str]]]) -> None:
    """
    Append one row.  new_brands_data is a list of (brand_name, [category, ...]).
    Each brand occupies one cell; categories are joined on a newline after the name.
    """
    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)

    if XLSX_PATH.exists():
        wb = load_workbook(XLSX_PATH)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "New Brands"
        ws.append(["Date"])   # minimal header; brand columns added below

    # Extend headers if this run has more brands than any previous run
    _extend_headers(ws, len(new_brands_data))

    row = [run_date]
    for brand_name, categories in new_brands_data:
        if categories:
            cell_value = f"{brand_name}\n{', '.join(categories)}"
        else:
            cell_value = brand_name
        row.append(cell_value)

    ws.append(row)
    wb.save(XLSX_PATH)


# ── Main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    print("Fetching Technische Unie brands overview page…")
    html = fetch_html(BRANDS_URL)

    current_brands         = extract_all_brands(html)
    current_featured_slugs = extract_featured_slugs(html)

    print(
        f"Found {len(current_brands):,} brands, "
        f"{len(current_featured_slugs)} featured brand pages"
    )

    state = load_state()

    # ── First run: just seed the state, no Excel output ────────────────────────
    if not state:
        print(
            "\nFirst run – seeding state file with current brand snapshot.\n"
            "No Excel row written (all current brands would look 'new').\n"
            "From the next run onwards, genuinely new brands will be logged."
        )
        save_state(current_brands, current_featured_slugs)
        print(f"✓ State saved to {STATE_FILE}")
        return

    known_brands         = set(state.get("known_brands", []))
    known_featured_slugs = set(state.get("known_featured_slugs", []))

    new_brands         = sorted(current_brands         - known_brands)
    new_featured_slugs = sorted(current_featured_slugs - known_featured_slugs)

    # ── Report & log ───────────────────────────────────────────────────────────
    if new_brands:
        print(f"\n{len(new_brands)} new brand(s) detected:")
        new_brands_data: list[tuple[str, list[str]]] = []

        for brand_name in new_brands:
            slug       = brand_name_to_slug(brand_name)
            categories: list[str] = []

            if slug in current_featured_slugs:
                categories = get_brand_categories(slug)
                time.sleep(CATEGORY_DELAY_S)

            new_brands_data.append((brand_name, categories))

            if categories:
                print(f"  • {brand_name}  →  {', '.join(categories)}")
            else:
                print(f"  • {brand_name}  →  (no dedicated category page)")

        today = date.today().isoformat()
        append_to_excel(today, new_brands_data)
        print(f"\n✓ Row appended to {XLSX_PATH}")

        # Write alert file for the GitHub Actions issue-creation step
        alert_dir  = Path(os.environ.get("RUNNER_TEMP", tempfile.gettempdir()))
        alert_file = alert_dir / "tu_brands_alert.json"
        title = f"New TU brand(s) detected ({today}): {', '.join(b[0] for b in new_brands_data)}"
        body_lines = []
        for brand_name, cats in new_brands_data:
            line = f"**{brand_name}**"
            if cats:
                line += f"  \nCategories: {', '.join(cats)}"
            body_lines.append(line)
        body = (
            f"The following new brand(s) appeared on the "
            f"[Technische Unie brands page](https://www.c.technischeunie.nl/merken-overzicht.html) "
            f"on {today}:\n\n"
            + "\n\n".join(body_lines)
        )
        with open(alert_file, "w", encoding="utf-8") as fh:
            json.dump({"title": title, "body": body}, fh, ensure_ascii=False)
        print(f"✓ Alert file written → {alert_file}")
    else:
        print("No new brands in the A-Z list.")

    if new_featured_slugs:
        print(
            f"\nNew featured brand pages (dedicated pages added): "
            f"{', '.join(new_featured_slugs)}"
        )

    # ── Save updated state ─────────────────────────────────────────────────────
    save_state(current_brands, current_featured_slugs)
    print(f"✓ State updated → {STATE_FILE}")


if __name__ == "__main__":
    main()
