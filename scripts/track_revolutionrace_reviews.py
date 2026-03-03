#!/usr/bin/env python3
"""
track_revolutionrace_reviews.py

Tracks Revolution Race product reviews via their public GraphQL reviews API
(https://reviews.revolutionrace.com/revolutionrace/graphql).

Methodology
-----------
1. Pull current total review count per base-product with a single efficient
   GraphQL facets call.
2. For any product not yet seen, batch-resolve its name + product-page URL via
   GraphQL alias queries (50 products per request). Extracts both the Swedish
   URL (SE_CHANNEL_UUID) and the German URL (DE_CHANNEL_UUID) as fallback.
3. Fetch / periodically refresh the price in parallel (8 workers):
     - SE URL  → price in SEK (direct)
     - DE URL  → price in EUR, converted to SEK via live ECB rate
     Reviews are already aggregated at baseProduct level, so there is NO
     double-counting across colour variants.
4. Compare against the previous stored snapshot → new_reviews per product.
5. Sales-activity proxy  =  Σ  price_sek × new_reviews   (all products)

State file    :  data/revolutionrace_state.json
Excel output  :  data/revolutionrace_reviews.xlsx
  Sheet "Summary"  – date | total_reviews | total_new_reviews | proxy_value_sek
"""

import ast
import json
import re
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date
from pathlib import Path
from typing import Optional

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

# ── Configuration ──────────────────────────────────────────────────────────────
GRAPHQL_URL      = "https://reviews.revolutionrace.com/revolutionrace/graphql"
SE_CHANNEL_UUID  = "3963b20d-4d89-4ddb-92dc-d0c897dc149a"  # Swedish store (SEK)
DE_CHANNEL_UUID  = "76302142-cd49-4c57-a48e-9217cf41c8b5"  # German store  (EUR)
EUR_SEK_FALLBACK = 11.5                                     # fallback rate if API unavailable

SCRIPT_DIR       = Path(__file__).resolve().parent
STATE_FILE       = (SCRIPT_DIR / ".." / "data" / "revolutionrace_state.json").resolve()
XLSX_PATH        = (SCRIPT_DIR / ".." / "data" / "revolutionrace_reviews.xlsx").resolve()

PRICE_REFRESH_DAYS  = 7   # re-fetch price if stored value is older than this
INFO_BATCH_SIZE     = 50  # products per batched GraphQL aliases query
PRICE_WORKERS       = 8   # parallel threads for price fetching
PRICE_CHUNK         = 80  # report progress every N products
REQUEST_DELAY_S     = 0.2 # pause between non-threaded HTTP requests

# Channel UUID → ISO country code (identified via dominant review language)
CHANNEL_COUNTRIES: dict[str, str] = {
    "76302142-cd49-4c57-a48e-9217cf41c8b5": "DE",
    "3963b20d-4d89-4ddb-92dc-d0c897dc149a": "SE",
    "4b897b07-7369-4665-abf3-ca895d31ecf3": "FI",
    "89853c2f-41a4-46db-b1ae-8b9977f1e4ec": "DK",
    "206916a9-9467-4fda-a02e-19a9307aec04": "PL",
    "977bb77c-8cb2-443c-bbd0-fad62f3c46b5": "NL",
    "65963f84-c285-4df7-9f6a-5876288db01c": "AT",
    "d0b1658e-ba32-4283-847f-aa8accd33fbf": "UK",
    "0f5ea666-e2bf-4c57-9ff4-6b849652f42d": "CZ",
    "719beded-b644-46be-98cc-4f35b90c0263": "NO",
    "03bbe5f0-b12f-406b-8817-eedb455cbfa2": "FR",
    "30bd1080-c713-49ac-8762-993cb31fae61": "CH",
    "c3b6bb57-4b88-4a2c-977d-29aa97a0d89d": "IT",
    "9c9dfccb-2138-4357-99da-2cec70102091": "IE",
}
# Ordered lists used for consistent Excel column ordering
TRACKED_COUNTRIES   = ["DE", "SE", "FI", "DK", "PL", "NL", "AT", "UK", "CZ", "NO", "FR", "CH", "IT", "IE"]
TRACKED_CATEGORIES  = ["PANTS", "TOPS", "JACKETS", "ACCS", "FOOTWEAR", "BASELAYERS", "BAGS"]
TRACKED_GENDERS     = ["MEN", "WOMEN", "UNISEX"]

GQL_HEADERS = {
    "Content-Type": "application/json",
    "User-Agent":   "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
}
WEB_HEADERS = {
    "User-Agent":      "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept-Language": "sv-SE,sv;q=0.9",
}


# ── GraphQL helpers ────────────────────────────────────────────────────────────

def gql(query: str, retries: int = 3) -> dict:
    """POST a GraphQL query; retry with exponential back-off on failure."""
    for attempt in range(retries):
        try:
            resp = requests.post(
                GRAPHQL_URL,
                json={"query": query},
                headers=GQL_HEADERS,
                timeout=30,
            )
            resp.raise_for_status()
            return resp.json()
        except Exception:
            if attempt == retries - 1:
                raise
            time.sleep(2 ** attempt)


def fetch_all_review_counts() -> dict[str, int]:
    """
    Single GraphQL facets call → {base_product_id: total_review_count}.

    The facets endpoint returns up to the platform's default bucket limit
    (typically several hundred distinct products).
    """
    data = gql("""
    {
      publicReviews(take: 0, facets: { item_baseProduct: [] }) {
        total
        facets { item_baseProduct { value count } }
      }
    }
    """)
    facets = data["data"]["publicReviews"]["facets"]["item_baseProduct"]
    return {item["value"]: item["count"] for item in facets}


def fetch_global_aggregates() -> dict:
    """
    Single API call returning:
      - average rating (0–1 scale)
      - rating breakdown {value_str: count}
      - review counts by channel UUID  {uuid: count}
      - review counts by product category {PANTS: count, …}
      - review counts by gender {MEN: count, …}
    ~350 ms, 2 KB response.
    """
    data = gql("""
    {
      publicReviews(take: 0, facets: {
        channelId: [],
        item_parentItemCategory: [],
        item_gender: [],
        rating: [],
        language: []
      }) {
        total
        average
        breakdown { value count }
        facets {
          channelId            { value count }
          item_parentItemCategory { value count }
          item_gender          { value count }
          language             { value count }
        }
      }
    }
    """)
    r = data["data"]["publicReviews"]
    return {
        "avg_rating":  r["average"],
        "breakdown":   {item["value"]: item["count"] for item in r["breakdown"]},
        "channels":    {item["value"]: item["count"] for item in r["facets"]["channelId"]},
        "categories":  {item["value"]: item["count"] for item in r["facets"]["item_parentItemCategory"]},
        "genders":     {item["value"]: item["count"] for item in r["facets"]["item_gender"]},
        "languages":   {item["value"]: item["count"] for item in r["facets"]["language"]},
    }


def _gql_alias(bp: str) -> str:
    """Sanitize a base-product ID into a valid GraphQL alias (no hyphens etc)."""
    return "p" + re.sub(r"[^a-zA-Z0-9]", "_", bp)


def _parse_productUrls(raw: Optional[str]) -> dict:
    """
    The API returns productUrls as a Python-dict-literal string, but with JSON
    null values instead of Python None. Fix that before parsing.
    """
    if not raw:
        return {}
    try:
        fixed = raw.replace(": null", ": None").replace(":null", ": None")
        result = ast.literal_eval(fixed)
        return result if isinstance(result, dict) else {}
    except Exception:
        return {}


def _parse_item_fields(item: dict, base_product: str) -> dict:
    """Extract Swedish name, SE URL and DE URL from a raw item dict."""
    name = base_product
    try:
        display_dict = ast.literal_eval(item["displayName"])
        name = (
            display_dict.get("sv-SE")
            or next(iter(display_dict.values()), base_product)
        )
    except Exception:
        pass

    urls = _parse_productUrls(item.get("productUrls"))
    return {
        "name":   name,
        "se_url": urls.get(SE_CHANNEL_UUID),  # None if not sold in Sweden
        "de_url": urls.get(DE_CHANNEL_UUID),  # None if not sold in Germany
    }


def fetch_product_info_batch(base_products: list[str]) -> dict[str, dict]:
    """
    Fetch display name + SE URL for a batch of base-products in ONE GraphQL
    request using query aliases (up to INFO_BATCH_SIZE products at a time).

    Returns {base_product_id: {"name": ..., "se_url": ...}}.
    The API stores displayName and productUrls as Python-dict-literal strings
    (single-quoted), parsed here with ast.literal_eval.
    """
    results: dict[str, dict] = {}

    for chunk_start in range(0, len(base_products), INFO_BATCH_SIZE):
        chunk = base_products[chunk_start : chunk_start + INFO_BATCH_SIZE]

        # Build a multi-alias query: one alias per product
        aliases = "\n".join(
            f'  {_gql_alias(bp)}: publicReviews(filter: {{ item_baseProduct: ["{bp}"] }}, take: 1) {{'
            f'    hits {{ item {{ displayName productUrls }} }}'
            f'  }}'
            for bp in chunk
        )
        data = gql(f"{{ {aliases} }}")

        for bp in chunk:
            alias_data = data["data"].get(_gql_alias(bp), {})
            hits = alias_data.get("hits", [])
            if hits:
                results[bp] = _parse_item_fields(hits[0]["item"], bp)
            else:
                results[bp] = {"name": bp, "se_url": None, "de_url": None}

        time.sleep(REQUEST_DELAY_S)

    return results


# ── EUR/SEK exchange rate ────────────────────────────────────────────────────────

def fetch_eur_sek_rate() -> float:
    """
    Fetch today's EUR/SEK rate from the Frankfurter API (free, ECB-sourced).
    Falls back to EUR_SEK_FALLBACK if unavailable.
    """
    try:
        resp = requests.get(
            "https://api.frankfurter.app/latest?from=EUR&to=SEK",
            timeout=10,
        )
        resp.raise_for_status()
        rate = resp.json()["rates"]["SEK"]
        print(f"  EUR/SEK rate: {rate:.4f}")
        return float(rate)
    except Exception:
        print(f"  EUR/SEK rate: using fallback {EUR_SEK_FALLBACK}")
        return EUR_SEK_FALLBACK


# ── Price fetching ─────────────────────────────────────────────────────────────

def _fetch_price_from_page(url: str, eur_sek: float) -> Optional[float]:
    """
    Fetch price from a product page. Reads the JSON-LD Product schema.
    For EUR-priced pages (revolutionrace.de / .at / .ch / .eu etc.),
    converts the EUR price to SEK using the supplied rate.
    Returns the price in SEK.
    """
    try:
        resp = requests.get(url, headers=WEB_HEADERS, timeout=15)
        resp.raise_for_status()
    except Exception:
        return None
    soup = BeautifulSoup(resp.text, "html.parser")
    for script in soup.find_all("script", type="application/ld+json"):
        try:
            ld = json.loads(script.string)
            if not isinstance(ld, dict) or ld.get("@type") != "Product":
                continue
            offers = ld.get("offers", {})
            if isinstance(offers, list):
                offers = offers[0]
            price    = offers.get("price")
            currency = offers.get("priceCurrency", "SEK")
            if price is None:
                continue
            price = float(price)
            if currency == "EUR":
                price = round(price * eur_sek, 2)
            return price
        except Exception:
            continue
    return None


def fetch_prices_parallel(
    url_map: dict[str, str],  # {base_product: url (SE or DE)}
    eur_sek: float,
    workers: int = PRICE_WORKERS,
    chunk_size: int = PRICE_CHUNK,
) -> dict[str, Optional[float]]:
    """
    Fetch prices for many products in parallel using a thread pool.
    Processes in chunks to show progress and avoid connection pool exhaustion.
    Returns {base_product: price_sek_or_None}.
    """
    results: dict[str, Optional[float]] = {}
    items = list(url_map.items())
    total = len(items)

    for chunk_start in range(0, total, chunk_size):
        chunk = items[chunk_start : chunk_start + chunk_size]
        with ThreadPoolExecutor(max_workers=workers) as pool:
            future_to_bp = {
                pool.submit(_fetch_price_from_page, url, eur_sek): bp
                for bp, url in chunk
            }
            for future in as_completed(future_to_bp, timeout=120):
                bp = future_to_bp[future]
                try:
                    results[bp] = future.result()
                except Exception:
                    results[bp] = None
        done = min(chunk_start + chunk_size, total)
        ok   = sum(1 for bp, _ in chunk if results.get(bp) is not None)
        print(f"    {done:>4}/{total}  (chunk ok: {ok}/{len(chunk)})")

    return results




# ── State management ───────────────────────────────────────────────────────────

def load_state() -> dict:
    if STATE_FILE.exists():
        with open(STATE_FILE, encoding="utf-8") as f:
            return json.load(f)
    return {"products": {}, "runs": []}


def save_state(state: dict) -> None:
    STATE_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(STATE_FILE, "w", encoding="utf-8") as f:
        json.dump(state, f, ensure_ascii=False, indent=2)


# ── Excel output ───────────────────────────────────────────────────────────────

def _append_df(wb: Workbook, sheet_name: str, df: pd.DataFrame) -> None:
    """Append a DataFrame to an existing or new sheet (header written once)."""
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in dataframe_to_rows(df, index=False, header=False):
            ws.append(row)
    else:
        ws = wb.create_sheet(title=sheet_name)
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)


def write_to_excel(summary_row: dict) -> None:
    """
    Appends one summary row to the 'Summary' sheet.
    Per-product history lives in revolutionrace_state.json and is not
    written to Excel (avoids the file growing by 1000+ rows per run).
    """
    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)

    if XLSX_PATH.exists():
        wb = load_workbook(XLSX_PATH)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    _append_df(wb, "Summary", pd.DataFrame([summary_row]))

    wb.save(XLSX_PATH)
    print(f"  → Saved: {XLSX_PATH}")


# ── Main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    today = date.today().isoformat()
    print(f"[{today}] Revolution Race – review tracker")

    state          = load_state()
    products_state = state.setdefault("products", {})
    runs           = state.setdefault("runs", [])

    # ── 1. Current review counts + global aggregates (2 API calls) ──────────────
    print("  Fetching review counts from GraphQL …")
    current_counts = fetch_all_review_counts()
    total_now = sum(current_counts.values())
    print(f"  {len(current_counts):,} products · {total_now:,} total reviews")

    print("  Fetching global aggregates (country / category / gender / rating) …")
    agg_now = fetch_global_aggregates()

    # Store daily aggregate snapshots; update in-place if re-running same day.
    agg_state = state.setdefault("aggregates", [])
    prior_agg  = next((a for a in reversed(agg_state) if a["date"] != today), None)
    today_snap = {
        "date":       today,
        "avg_rating": agg_now["avg_rating"],
        "channels":   agg_now["channels"],
        "categories": agg_now["categories"],
        "genders":    agg_now["genders"],
        "languages":  agg_now["languages"],
    }
    if agg_state and agg_state[-1]["date"] == today:
        agg_state[-1] = today_snap
    else:
        agg_state.append(today_snap)

    # Deltas vs prior day
    prev_ch  = prior_agg["channels"]   if prior_agg else {}
    prev_cat = prior_agg["categories"] if prior_agg else {}
    prev_gen = prior_agg["genders"]    if prior_agg else {}
    ch_delta  = {k: max(0, v - prev_ch.get(k, 0))  for k, v in agg_now["channels"].items()}
    cat_delta = {k: max(0, v - prev_cat.get(k, 0)) for k, v in agg_now["categories"].items()}
    gen_delta = {k: max(0, v - prev_gen.get(k, 0)) for k, v in agg_now["genders"].items()}

    # Per-country new reviews using CHANNEL_COUNTRIES map
    country_new: dict[str, int] = {}
    other_country_new = 0
    for uuid, delta in ch_delta.items():
        code = CHANNEL_COUNTRIES.get(uuid)
        if code:
            country_new[code] = country_new.get(code, 0) + delta
        else:
            other_country_new += delta

    # ── Expansion detection: new channel UUIDs or new languages ──────────────
    # On first run, pre-seed from any prior aggregate snapshots so we don't
    # false-positive on already-established channels/languages.
    is_first_channels  = "seen_channels"  not in state
    is_first_languages = "seen_languages" not in state
    all_seen_channels  = state.setdefault("seen_channels",  {})
    all_seen_languages = state.setdefault("seen_languages", {})

    # Seed from prior agg snapshots if this is the first time we're tracking.
    if is_first_channels:
        for snap in agg_state[:-1]:  # all snapshots except today's
            for uuid in snap.get("channels", {}):
                if uuid not in all_seen_channels:
                    all_seen_channels[uuid] = {"first_seen": snap["date"], "first_count": snap["channels"][uuid]}
    if is_first_languages:
        for snap in agg_state[:-1]:
            for lang in snap.get("languages", {}):
                if lang not in all_seen_languages:
                    all_seen_languages[lang] = {"first_seen": snap["date"], "first_count": snap["languages"][lang]}
        # If still empty (no prior language data at all), seed silently from current
        if not all_seen_languages:
            for lang, count in agg_now["languages"].items():
                all_seen_languages[lang] = {"first_seen": today, "first_count": count}
    if is_first_channels and not all_seen_channels:
        for uuid, count in agg_now["channels"].items():
            all_seen_channels[uuid] = {"first_seen": today, "first_count": count}

    new_channel_uuids  = []
    new_language_codes = []
    for uuid, count in agg_now["channels"].items():
        if uuid not in all_seen_channels:
            new_channel_uuids.append(uuid)
            all_seen_channels[uuid] = {"first_seen": today, "first_count": count}
    for lang, count in agg_now["languages"].items():
        if lang not in all_seen_languages:
            new_language_codes.append(lang)
            all_seen_languages[lang] = {"first_seen": today, "first_count": count}
    if new_channel_uuids:
        print(f"  *** EXPANSION SIGNAL: {len(new_channel_uuids)} new channel(s) detected: {new_channel_uuids}")
    if new_language_codes:
        print(f"  *** EXPANSION SIGNAL: {len(new_language_codes)} new language(s) detected: {new_language_codes}")

    # ── 2. Batch-fetch product info for new products + products missing de_url ──
    # Also re-fetch any product that predates the de_url feature (key absent).
    new_products = [
        bp for bp in current_counts
        if not products_state.get(bp, {}).get("name")
        or "de_url" not in products_state.get(bp, {})
    ]
    if new_products:
        print(f"  Resolving info for {len(new_products)} new products "
              f"({len(new_products) // INFO_BATCH_SIZE + 1} batch requests) …")
        info_map = fetch_product_info_batch(new_products)
        for bp, info in info_map.items():
            p = products_state.setdefault(bp, {
                "name": None, "se_url": None, "de_url": None,
                "price_sek": None, "price_updated": None, "counts": [],
            })
            p["name"]   = info["name"]
            p["se_url"] = info["se_url"]
            p["de_url"] = info["de_url"]
        print(f"  Done.")
        save_state(state)  # checkpoint: avoid re-fetching info on retry

    # ── 3. Parallel-fetch prices for products that need a refresh ────────────
    # Priority: SE URL (native SEK) → DE URL (EUR × EUR/SEK rate).
    def _needs_refresh(p: dict) -> bool:
        return (
            p.get("price_sek") is None
            or p.get("price_updated") is None
            or (date.today() - date.fromisoformat(p["price_updated"])).days > PRICE_REFRESH_DAYS
        )

    needs_price = {
        bp: (products_state[bp].get("se_url") or products_state[bp].get("de_url"))
        for bp in current_counts
        if _needs_refresh(products_state.get(bp, {}))
        and (products_state.get(bp, {}).get("se_url") or products_state.get(bp, {}).get("de_url"))
    }
    if needs_price:
        eur_sek = fetch_eur_sek_rate()
        se_count = sum(1 for bp, url in needs_price.items() if url and "revolutionrace.se" in url)
        de_count = len(needs_price) - se_count
        print(f"  Fetching prices for {len(needs_price)} products "
              f"(SE: {se_count}, DE/other: {de_count}, {PRICE_WORKERS} workers) …")
        price_results = fetch_prices_parallel(needs_price, eur_sek=eur_sek)
        for bp, price in price_results.items():
            if price is not None:
                products_state[bp]["price_sek"]     = price
                products_state[bp]["price_updated"] = today
        fetched_ok = sum(1 for p in price_results.values() if p is not None)
        print(f"  Prices fetched: {fetched_ok}/{len(needs_price)}")

    # ── 4. Compute per-product deltas ─────────────────────────────────────────
    per_product_rows: list[dict] = []
    total_proxy       = 0.0
    total_new_reviews = 0

    for base_product, current_count in sorted(current_counts.items()):
        p = products_state.setdefault(base_product, {
            "name": None, "se_url": None,
            "price_sek": None, "price_updated": None, "counts": [],
        })

        # Previous count = last entry that is NOT today (so re-runs on the same
        # day don't accumulate phantom new reviews).
        prior_entries = [e for e in p["counts"] if e["date"] != today]
        prev_count  = prior_entries[-1]["count"] if prior_entries else 0
        new_reviews = max(0, current_count - prev_count)

        # Update today's entry in-place rather than appending duplicates.
        if p["counts"] and p["counts"][-1]["date"] == today:
            p["counts"][-1]["count"] = current_count
        else:
            p["counts"].append({"date": today, "count": current_count})

        price = p.get("price_sek") or 0.0
        proxy = price * new_reviews
        total_proxy       += proxy
        total_new_reviews += new_reviews

        per_product_rows.append({
            "date":          today,
            "base_product":  base_product,
            "name":          p.get("name") or base_product,
            "price_sek":     price,
            "total_reviews": current_count,
            "new_reviews":   new_reviews,
            "proxy_value":   round(proxy),
        })

    # ── 5. Summary ─────────────────────────────────────────────────────────────
    # Rating metrics: API uses 0–1 scale; convert avg to 0–5 for readability.
    avg_5star = round(agg_now["avg_rating"] * 5, 3)
    five_star_count = agg_now["breakdown"].get("1.0", agg_now["breakdown"].get("1", 0))
    total_breakdown = sum(agg_now["breakdown"].values())
    pct_5star = round(five_star_count / total_breakdown * 100, 1) if total_breakdown else 0.0

    summary_row: dict = {
        "date":              today,
        "total_reviews":     total_now,
        "total_new_reviews": total_new_reviews,
        "proxy_value_sek":   round(total_proxy),
        "avg_rating":        avg_5star,
        "pct_5star":         pct_5star,
    }
    # New reviews by country
    for code in TRACKED_COUNTRIES:
        summary_row[f"new_{code}"] = country_new.get(code, 0)
    summary_row["new_other_country"] = other_country_new
    # New reviews by category
    for cat in TRACKED_CATEGORIES:
        summary_row[f"new_{cat}"] = cat_delta.get(cat, 0)
    other_cat_delta = sum(v for k, v in cat_delta.items() if k not in TRACKED_CATEGORIES)
    summary_row["new_other_cat"] = other_cat_delta
    # New reviews by gender
    for g in TRACKED_GENDERS:
        summary_row[f"new_{g}"] = gen_delta.get(g, 0)
    # Expansion signals
    summary_row["new_channel_uuids"]  = ",".join(new_channel_uuids)  if new_channel_uuids  else ""
    summary_row["new_language_codes"] = ",".join(new_language_codes) if new_language_codes else ""

    runs.append(summary_row)

    print(f"\n  ── Summary {'─' * 38}")
    print(f"     Total reviews       {total_now:>12,}")
    print(f"     New reviews today   {total_new_reviews:>12,}")
    print(f"     Proxy value (SEK)   {total_proxy:>12,.0f}")
    print(f"     Avg rating          {avg_5star:>12.3f}  /5")
    print(f"     % 5-star            {pct_5star:>11.1f}%")
    print(f"     New by country      " + "  ".join(f"{c}:{country_new.get(c,0)}" for c in TRACKED_COUNTRIES[:6]))
    print(f"                         " + "  ".join(f"{c}:{country_new.get(c,0)}" for c in TRACKED_COUNTRIES[6:]))
    print(f"     New by category     " + "  ".join(f"{cat}:{cat_delta.get(cat,0)}" for cat in TRACKED_CATEGORIES))
    print(f"     New by gender       " + "  ".join(f"{g}:{gen_delta.get(g,0)}" for g in TRACKED_GENDERS))
    print(f"     Channels total      {len(agg_now['channels']):>12,}  (known: {sum(1 for u in agg_now['channels'] if u in CHANNEL_COUNTRIES)})")
    print(f"     Languages total     {len(agg_now['languages']):>12,}  ({', '.join(sorted(agg_now['languages'])[:8])}{'...' if len(agg_now['languages'])>8 else ''})")

    # ── 6. Persist state + write Excel ────────────────────────────────────────
    save_state(state)
    write_to_excel(summary_row)


if __name__ == "__main__":
    main()
