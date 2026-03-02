import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path
import openpyxl
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError

# ==========================================
# Configuration & Setup
# ==========================================
SCRIPT_NAME = "youtube_diy_trends.py"
DATA_FILENAME = "youtube_diy_sentiment.xlsx"
PRIMARY_QUERY = "Gaming PC Build"
API_KEY = "AIzaSyCHwNxI4HSv5cbLx3praqwLv7w_1YdGeCM"

# SET THIS TO TRUE TO SEE DETAILED OUTPUT IN CONSOLE
DEBUG = False

# Window for "top recent videos by view count" (stable, consistent set day-to-day)
RECENT_WINDOW_DAYS = 30
MAX_RECENT_VIDEOS = 200

# Window for counting new upload velocity (how many videos published recently)
UPLOAD_WINDOW_DAYS = 7
MAX_UPLOAD_COUNT = 2000

# ==========================================
# Helpers
# ==========================================

def get_iso_date(days_ago: int) -> str:
    dt = datetime.now(timezone.utc) - timedelta(days=days_ago)
    return dt.strftime("%Y-%m-%dT%H:%M:%SZ")

def chunked_list(lst, n):
    for i in range(0, len(lst), n):
        yield lst[i:i + n]

def fetch_video_ids(youtube, query, published_after, order, max_results):
    """Page through search results and return (video_ids, top_title)."""
    video_ids = []
    top_title = "N/A"
    next_page_token = None

    while len(video_ids) < max_results:
        fetch_count = min(max_results - len(video_ids), 50)
        resp = youtube.search().list(
            q=query,
            part="id,snippet",
            maxResults=fetch_count,
            type="video",
            order=order,
            publishedAfter=published_after,
            pageToken=next_page_token,
        ).execute()

        items = resp.get("items", [])
        if not items:
            break

        if not next_page_token and items:
            top_title = items[0]["snippet"]["title"]

        for item in items:
            video_ids.append(item["id"]["videoId"])

        next_page_token = resp.get("nextPageToken")
        if not next_page_token:
            break

    return video_ids, top_title

def fetch_view_counts(youtube, video_ids):
    """Fetch viewCount for a list of video IDs. Returns list of ints."""
    view_counts = []
    for chunk in chunked_list(video_ids, 50):
        resp = youtube.videos().list(
            part="statistics,snippet",
            id=",".join(chunk),
        ).execute()
        for video in resp.get("items", []):
            views = int(video["statistics"].get("viewCount", 0))
            view_counts.append(views)
            if DEBUG:
                print(f"    [Views: {views:,}] {video['snippet']['title']}")
    return view_counts

# ==========================================
# Signal 1 & 2 – YouTube data
# ==========================================

def get_youtube_data():
    """
    Collects two YouTube signals that are far more stable than the old
    'top-500 by relevance' approach:

    Signal 2 – Top-viewed videos published in the last 30 days
        Uses order=viewCount so the same high-performing videos appear
        in the set day after day.  Average / total views on this fixed
        window changes slowly and reflects genuine audience size.

    Signal 3 – New upload velocity (last 7 days)
        Counts how many new 'Gaming PC Build' videos were published in the
        past week.  When interest rises, more creators publish → count goes
        up.  This is a demand-driven supply signal that does not depend on
        YouTube's relevance algorithm at all.
    """
    try:
        youtube = build("youtube", "v3", developerKey=API_KEY)

        # --- Signal 2: top recent videos sorted by viewCount ---
        published_after_30d = get_iso_date(RECENT_WINDOW_DAYS)
        print(
            f"[{SCRIPT_NAME}] Fetching top-viewed videos from last "
            f"{RECENT_WINDOW_DAYS} days (order=viewCount)..."
        )
        if DEBUG:
            print(f"\n  -- Recent top-viewed videos --")

        recent_ids, top_title = fetch_video_ids(
            youtube, PRIMARY_QUERY, published_after_30d,
            order="viewCount", max_results=MAX_RECENT_VIDEOS
        )
        recent_views = fetch_view_counts(youtube, recent_ids)

        avg_recent_views   = int(sum(recent_views) / len(recent_views)) if recent_views else 0
        total_recent_views = sum(recent_views)
        print(
            f"[{SCRIPT_NAME}] {len(recent_ids)} recent videos | "
            f"avg views: {avg_recent_views:,} | total: {total_recent_views:,}"
        )

        # --- Signal 3: new upload count in last 7 days ---
        published_after_7d = get_iso_date(UPLOAD_WINDOW_DAYS)
        print(f"[{SCRIPT_NAME}] Counting new uploads in last {UPLOAD_WINDOW_DAYS} days...")
        upload_ids, _ = fetch_video_ids(
            youtube, PRIMARY_QUERY, published_after_7d,
            order="date", max_results=MAX_UPLOAD_COUNT
        )
        new_upload_count = len(upload_ids)
        print(f"[{SCRIPT_NAME}] New uploads (last {UPLOAD_WINDOW_DAYS}d): {new_upload_count}")

        return {
            "recent_video_count":  len(recent_ids),
            "avg_recent_views":    avg_recent_views,
            "total_recent_views":  total_recent_views,
            "top_recent_title":    top_title,
            "new_uploads_7d":      new_upload_count,
        }

    except HttpError as e:
        print(f"[{SCRIPT_NAME}] YouTube API HttpError: {e}")
        return None
    except Exception as e:
        print(f"[{SCRIPT_NAME}] YouTube error: {e}")
        return None

# ==========================================
# Excel output
# ==========================================

HEADERS = [
    "Date",
    "Query",
    # Signal 1 – YouTube: top-viewed recent videos (stable set; sorted by viewCount)
    "Avg_Views_Top200_Last30d",
    "Total_Views_Top200_Last30d",
    "Video_Count_Last30d",
    "Top_Recent_Video_Title",
    # Signal 2 – YouTube: upload velocity (demand-driven supply proxy)
    "New_Uploads_Last7d",
]

def update_excel(yt_metrics):
    script_dir = Path(__file__).resolve().parent
    data_dir   = script_dir.parent / "data"
    data_dir.mkdir(parents=True, exist_ok=True)
    file_path  = data_dir / DATA_FILENAME

    if not file_path.exists():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "YouTube Trends"
        ws.append(HEADERS)
    else:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        # Refresh header row if the file predates this script version
        existing = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
        if existing != HEADERS:
            for col, h in enumerate(HEADERS, 1):
                ws.cell(row=1, column=col).value = h

    ws.append([
        datetime.now().strftime("%Y-%m-%d"),
        PRIMARY_QUERY,
        yt_metrics["avg_recent_views"],
        yt_metrics["total_recent_views"],
        yt_metrics["recent_video_count"],
        yt_metrics["top_recent_title"],
        yt_metrics["new_uploads_7d"],
    ])

    wb.save(file_path)
    print(f"[{SCRIPT_NAME}] Data saved to {file_path}")

# ==========================================
# Entry point
# ==========================================

def main():
    yt_metrics = get_youtube_data()

    if yt_metrics:
        update_excel(yt_metrics)
    else:
        print(f"[{SCRIPT_NAME}] Skipped Excel update due to missing YouTube data.")

if __name__ == "__main__":
    main()