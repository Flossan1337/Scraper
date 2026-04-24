import asyncio
import random
import re
import os
from datetime import date
from pathlib import Path
from playwright.async_api import async_playwright

try:
    import openpyxl
except ImportError:
    print("Du måste installera openpyxl: pip install openpyxl")
    exit()

# ── KONFIGURATION ──────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).resolve().parent
DATA_DIR = (SCRIPT_DIR / ".." / "data").resolve()
DATA_DIR.mkdir(parents=True, exist_ok=True)

XLSX_PATH = str((DATA_DIR / "anoto_amazon_data.xlsx").resolve())
SHEET_NAME = "Tracking"

HEADLESS = True

# --- PRODUKT ---
PRODUCT_NAME = "inq Smart Writing Set"
ASIN = "B0FW7G4KXP"
DOMAIN = "amazon.com"
GL = "US"
LOCALE = "en-US"
ACCEPT_LANG = "en-US,en;q=0.9"
COOKIE_NAME = "i18n-prefs"
COOKIE_VALUE = "USD"

UAS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_0) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4 Safari/605.1.15",
]

BOUGHT_SELECTORS = [
    "#social-proofing-faceout-title-tk_bought span.a-text-bold",
    "#social-proofing-faceout-title-tk_bought",
    "div.social-proofing-faceout span.a-text-bold",
]

# ───────────────────────────────────────────────────────────────────────────

def parse_number(text: str) -> int:
    """
    Handles '50+', '1K+', '1.5K', '2M+' and converts to int.
    K/M multiplier is only applied if it directly follows the number
    and is NOT followed by another letter (avoids words like 'month').
    """
    if not text:
        return 0

    match = re.search(r"([0-9]+(?:[.,][0-9]+)?)\s*([kKmM])(?![a-zA-Z])", text)
    if match:
        num_str = match.group(1).replace(",", ".")
        suffix = match.group(2).lower()
        try:
            val = float(num_str)
            if suffix == "k":
                return int(val * 1000)
            elif suffix == "m":
                return int(val * 1_000_000)
        except ValueError:
            return 0

    plain = re.search(r"([0-9]+)", text)
    if plain:
        try:
            return int(plain.group(1))
        except ValueError:
            return 0

    return 0


def extract_bought_from_html(html_content: str) -> int:
    """Scans raw HTML for patterns like '>50+ bought' or '>1K+ bought'."""
    match = re.search(r">([0-9.,]+[kKmM]?)\+?\s*bought", html_content, re.IGNORECASE)
    if match:
        return parse_number(match.group(1))
    return 0


def extract_rank_from_row_text(row_text: str) -> int:
    """
    Extracts Best Sellers Rank from a row of text.
    Ignores 'Top 100' links to avoid picking up that number.
    """
    clean_text = re.sub(r"top\s*100", "", row_text, flags=re.IGNORECASE)
    regex = r"(?:(?:Nr\.?|#)\s*([0-9.,]+))|([0-9.,]+)\s+in\s+"
    matches = re.findall(regex, clean_text, re.IGNORECASE)

    candidates = []
    for m in matches:
        raw_num = m[0] if m[0] else m[1]
        clean_str = raw_num.replace(",", "").replace(".", "")
        if clean_str.isdigit():
            val = int(clean_str)
            if 0 < val < 10_000_000:
                candidates.append(val)

    return min(candidates) if candidates else 0


async def handle_blockers(page):
    cookie_selectors = [
        "#sp-cc-accept",
        "input[name='accept']",
        "button[name='accept']",
        "text=Accept Cookies",
    ]
    for selector in cookie_selectors:
        try:
            if await page.locator(selector).is_visible(timeout=1000):
                print(f"    -> Cookie banner found ({selector}). Dismissing...")
                await page.click(selector)
                try:
                    await page.locator(selector).wait_for(state="hidden", timeout=3000)
                except Exception:
                    pass
                await page.wait_for_timeout(1000)
                break
        except Exception:
            continue

    upsells = ['text="Continue shopping"']
    for sel in upsells:
        try:
            if await page.locator(sel).is_visible(timeout=500):
                await page.click(sel)
                await page.wait_for_timeout(500)
        except Exception:
            pass


async def get_product_data(context) -> tuple[int, int]:
    page = await context.new_page()
    url = f"https://www.{DOMAIN}/dp/{ASIN}?th=1&psc=1&gl={GL}"
    print(f"  Fetching {PRODUCT_NAME} ({DOMAIN})...")

    bought_val = 0
    rank_val = 0

    try:
        await page.goto(url, wait_until="domcontentloaded", timeout=30000)
        await handle_blockers(page)

        full_html = await page.content()

        # --- 1. BOUGHT PAST MONTH ---
        found_bought_text = ""
        for sel in BOUGHT_SELECTORS:
            try:
                el = await page.wait_for_selector(sel, state="attached", timeout=1000)
                if el:
                    found_bought_text = (await el.inner_text()).strip()
                    if found_bought_text:
                        break
            except Exception:
                continue

        if found_bought_text:
            bought_val = parse_number(found_bought_text)
            print(f"    Bought (CSS): {bought_val} (Raw: '{found_bought_text}')")
        else:
            bought_val = extract_bought_from_html(full_html)
            if bought_val > 0:
                print(f"    Bought (Regex): {bought_val}")
            else:
                print(f"    Bought: Not found")

        # --- 2. BEST SELLERS RANK ---
        try:
            rank_header = page.locator("text=/Best Sellers Rank|Best Seller Rank/i").first

            if await rank_header.count() > 0:
                row_element = rank_header.locator(
                    "xpath=ancestor::tr | ancestor::li | ancestor::div[contains(@class, 'db_row')]"
                ).first

                if await row_element.count() > 0:
                    row_text = await row_element.inner_text()
                    rank_val = extract_rank_from_row_text(row_text)
                else:
                    row_text = await rank_header.evaluate("el => el.parentElement.innerText")
                    rank_val = extract_rank_from_row_text(row_text)

            if rank_val > 0:
                print(f"    Best Sellers Rank: #{rank_val}")
            else:
                body_text = await page.inner_text("body")
                rank_val = extract_rank_from_row_text(body_text)
                if rank_val > 0:
                    print(f"    Best Sellers Rank (Fallback): #{rank_val}")
                else:
                    print(f"    Best Sellers Rank: Not found")

        except Exception as e:
            print(f"    Rank error: {e}")

        return bought_val, rank_val

    except Exception as e:
        print(f"    Error: {e}")
        return 0, 0
    finally:
        await page.close()


def append_to_excel(today: str, bought: int, rank: int):
    file_exists = os.path.exists(XLSX_PATH)

    headers = ["Date", "Bought Past Month", "Best Sellers Rank"]

    if not file_exists:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(headers)
    else:
        wb = openpyxl.load_workbook(XLSX_PATH)
        if SHEET_NAME in wb.sheetnames:
            ws = wb[SHEET_NAME]
        else:
            ws = wb.create_sheet(SHEET_NAME)
            ws.append(headers)

    ws.append([today, bought, rank])
    wb.save(XLSX_PATH)
    print(f"Data saved to {XLSX_PATH}")


async def run_once():
    today = date.today().isoformat()

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=HEADLESS, args=["--start-maximized"])
        print(f"--- Anoto Amazon Data Fetcher ({today}) ---")

        context = await browser.new_context(
            locale=LOCALE,
            user_agent=random.choice(UAS),
            java_script_enabled=True,
            viewport={"width": 1920, "height": 1080},
            extra_http_headers={"Accept-Language": ACCEPT_LANG},
        )
        await context.add_cookies([{
            "name": COOKIE_NAME,
            "value": COOKIE_VALUE,
            "domain": f".{DOMAIN}",
            "path": "/",
            "secure": True,
            "httpOnly": False,
        }])

        bought, rank = await get_product_data(context)

        await context.close()
        await browser.close()

    print(f"\nResults for {PRODUCT_NAME} (US):")
    print(f"  Bought Past Month : {bought}")
    print(f"  Best Sellers Rank : {rank}")

    append_to_excel(today, bought, rank)


if __name__ == "__main__":
    asyncio.run(run_once())
