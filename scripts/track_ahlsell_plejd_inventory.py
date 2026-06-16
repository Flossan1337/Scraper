#!/usr/bin/env python3
"""
track_ahlsell_plejd_inventory.py

Spårar Plejds produktlager på Ahlsell.se per butik och centrallager.

Flöde
-----
  1. Hämta alla Plejd-produkter via sök-API (brand=Plejd, ~38 produkter)
  2. För produkter med fler än 1 variant, hämta samtliga artikelnummer
     via varianter-API (t.ex. TRM-01 har 6 färgvarianter)
  3. Hämta och cachelagra butikskatalog (~100 butiker)
  4. Hämta lagersaldo per artikelnummer och butik
  5. Spara daglig snapshot i JSON-tillståndsfil
  6. Exportera till Excel:
       • "Totalt"   — tidsserie per artikel (rader) × datum (kolumner)
       • "Butiker"  — senaste dagets butikslager i wide-format

API:er (ingen autentisering krävs)
-----------------------------------
  Sök:       GET https://www.ahlsell.se/api/search
             ?searchPhrase=plejd&pageSize=100
  Varianter: GET https://www.ahlsell.se/api/search/variants
             ?productCode={code}&activeVariantNumber={variantNumber}
  Butiker:   GET https://www.ahlsell.se/api/warehouses
  Lager:     GET https://www.ahlsell.se/api/warehouses/stock
             ?variantNumber={articleNumber}

Tillståndsfil  : data/ahlsell_plejd_state.json
Excel-utdata   : data/ahlsell_plejd_inventory.xlsx
"""

import json
import time
from datetime import date
from pathlib import Path

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Konfiguration ──────────────────────────────────────────────────────────────
BASE_URL       = "https://www.ahlsell.se"
SEARCH_URL     = f"{BASE_URL}/api/search"
VARIANTS_URL   = f"{BASE_URL}/api/search/variants"
WAREHOUSES_URL = f"{BASE_URL}/api/warehouses"
STOCK_URL      = f"{BASE_URL}/api/warehouses/stock"

SEARCH_PHRASE  = "plejd"
BRAND_FILTER   = "Plejd"
REQUEST_DELAY  = 0.3   # sekunder mellan anrop

STATE_FILE = Path(__file__).parent.parent / "data" / "ahlsell_plejd_state.json"
EXCEL_FILE = Path(__file__).parent.parent / "data" / "ahlsell_plejd_inventory.xlsx"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept": "application/json",
    "Accept-Language": "sv-SE,sv;q=0.9",
}

# ── API-anrop ──────────────────────────────────────────────────────────────────

def fetch_products() -> list[dict]:
    """Hämtar alla Plejd-produkter från sök-API:et (filtrerar på brand-klientsidan)."""
    resp = requests.get(
        SEARCH_URL,
        params={"searchPhrase": SEARCH_PHRASE, "pageSize": 100},
        headers=HEADERS,
        timeout=30,
    )
    resp.raise_for_status()
    cards = resp.json().get("productCards", [])
    return [c for c in cards if c.get("brand", "").upper() == BRAND_FILTER.upper()]


def fetch_all_variant_numbers(product_code: str, active_variant: str) -> list[str]:
    """
    Returnerar artikelnummer för samtliga varianter av en produkt
    (t.ex. alla 6 färger av TRM-01).
    """
    resp = requests.get(
        VARIANTS_URL,
        params={"productCode": product_code, "activeVariantNumber": active_variant},
        headers=HEADERS,
        timeout=30,
    )
    resp.raise_for_status()
    return [item["code"] for item in resp.json().get("items", []) if item.get("code")]


def fetch_warehouses() -> dict[str, dict]:
    """Hämtar butikskatalog och returnerar som {warehouseId: metadata}."""
    resp = requests.get(WAREHOUSES_URL, headers=HEADERS, timeout=30)
    resp.raise_for_status()
    return {
        str(w["id"]): {
            "name":        w.get("name", ""),
            "city":        w.get("city", ""),
            "address":     w.get("address", ""),
            "phone":       w.get("phoneNumber", ""),
            "url_segment": w.get("urlSegment", ""),
        }
        for w in resp.json()
    }


def fetch_stock(variant_number: str) -> dict[str, float]:
    """
    Hämtar lagersaldo per butik för ett artikelnummer.
    Returnerar {warehouseId: quantity} — bara poster med quantity > 0.
    """
    resp = requests.get(
        STOCK_URL,
        params={"variantNumber": variant_number},
        headers=HEADERS,
        timeout=30,
    )
    resp.raise_for_status()
    return {
        str(entry["id"]): entry["stock"]["quantity"]
        for entry in resp.json()
        if (entry.get("stock", {}).get("quantity") or 0) > 0
    }


# ── Insamling ──────────────────────────────────────────────────────────────────

def collect_snapshot() -> tuple[dict, dict, dict]:
    """
    Kör alla API-anrop och returnerar:
      products   : {articleNumber: {product_name, brand, product_code, page_url}}
      warehouses : {warehouseId: {name, city, address, ...}}
      stock      : {articleNumber: {warehouseId: quantity}}
    """
    # 1. Produktlista
    print("Hämtar produkter...")
    product_cards = fetch_products()
    print(f"  {len(product_cards)} Plejd-produkter hittade")

    # 2. Expandera till alla varianter
    products: dict[str, dict] = {}
    for card in product_cards:
        product_code    = card["code"]
        product_name    = card["name"]
        brand           = card["brand"]
        most_relevant   = str(card["variantNumber"])
        num_variants    = card.get("numberOfVariants", 1)
        page_url        = card.get("firstVariationPageUrl", "")

        if num_variants > 1:
            time.sleep(REQUEST_DELAY)
            try:
                variant_numbers = fetch_all_variant_numbers(product_code, most_relevant)
            except Exception as exc:
                print(f"  Varning: kunde ej hämta varianter för {product_name}: {exc}")
                variant_numbers = [most_relevant]
        else:
            variant_numbers = [most_relevant]

        for vn in variant_numbers:
            products[str(vn)] = {
                "product_name": product_name,
                "brand":        brand,
                "product_code": product_code,
                "page_url":     page_url,
            }

    print(f"  Totalt {len(products)} artikelnummer (inkl. alla varianter)")

    # 3. Butikskatalog
    print("Hämtar butikskatalog...")
    warehouses = fetch_warehouses()
    print(f"  {len(warehouses)} butiker")

    # 4. Lagersaldo per artikel
    print("Hämtar lagersaldo...")
    stock: dict[str, dict] = {}
    for i, art_num in enumerate(products):
        time.sleep(REQUEST_DELAY)
        try:
            stock[art_num] = fetch_stock(art_num)
        except Exception as exc:
            print(f"  Varning: lagerfel för {art_num}: {exc}")
            stock[art_num] = {}
        if (i + 1) % 10 == 0:
            print(f"  {i + 1}/{len(products)} artiklar klara...")

    total_entries = sum(len(v) for v in stock.values())
    print(f"  Klart — {total_entries} butiksposter med lager > 0")
    return products, warehouses, stock


# ── Tillstånd ──────────────────────────────────────────────────────────────────

def load_state() -> dict:
    if STATE_FILE.exists():
        return json.loads(STATE_FILE.read_text(encoding="utf-8"))
    return {"warehouses": {}, "products": {}, "snapshots": {}}


def save_state(state: dict) -> None:
    STATE_FILE.write_text(
        json.dumps(state, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def update_state(
    state: dict,
    products: dict,
    warehouses: dict,
    stock: dict,
) -> None:
    today = date.today().isoformat()
    state["warehouses"]    = warehouses
    state["products"]      = products
    state["last_updated"]  = today

    snapshot = {
        art: {
            "warehouses": wh_stock,
            "total":      sum(wh_stock.values()),
        }
        for art, wh_stock in stock.items()
    }
    state.setdefault("snapshots", {})[today] = snapshot


# ── Kategorisering ─────────────────────────────────────────────────────────────

# Ordningen spelar roll – LED-Panel kollas alltid först via artikelnummer
CATEGORIES      = ["Dimmer", "Armaturer", "Termostat", "LED-Panel", "Övrigt"]
LED_PANEL_ARTS  = {"7077777"}   # LPN-01 — ska INTE in under Armaturer
OVRIGT_ARTS     = {             # Vridadaptrar för dosdimmer — INTE Dimmer
    "1377733", "1377734", "1377735",
    "1377736", "1377737", "1377738",
}


def categorize(article: str, product_name: str) -> str:
    """Mappar en artikel till rätt kategorikolumn i Excel."""
    if article in LED_PANEL_ARTS:
        return "LED-Panel"
    if article in OVRIGT_ARTS:
        return "Övrigt"
    name = product_name.lower()
    if "dimmer" in name:
        return "Dimmer"
    if "led-panel" in name or "led panel" in name:
        return "LED-Panel"
    if any(w in name for w in ("downlight", "plafond", "väggarmatur")):
        return "Armaturer"
    if "termostat" in name:
        return "Termostat"
    return "Övrigt"


# ── Deltaberäkning ─────────────────────────────────────────────────────────────

def compute_deltas(
    snapshots: dict,
    products: dict,
) -> tuple[dict, dict]:
    """
    Beräknar dagliga deltan per artikel per butik och grupperar i kategorier.

    Logik: för varje konsekutivt datumpar jämförs lagersaldo per butik
    och artikel separat. Lagerminsk­ningar och lagerökningar hålls isär
    så att t.ex. -5 i Göteborg och +5 i Stockholm INTE kvittas mot
    varandra – båda registreras i sina respektive sheet.

    Returnerar:
      sales_out : {date: {kategori: enheter}}  – kunder köper från Ahlsell
      sales_in  : {date: {kategori: enheter}}  – Plejd säljer in till Ahlsell
    """
    sorted_dates = sorted(snapshots.keys())
    sales_out: dict[str, dict[str, float]] = {}
    sales_in:  dict[str, dict[str, float]] = {}

    for i in range(1, len(sorted_dates)):
        d_prev = sorted_dates[i - 1]
        d_curr = sorted_dates[i]

        snap_prev = snapshots[d_prev]
        snap_curr = snapshots[d_curr]

        out_by_cat: dict[str, float] = {c: 0.0 for c in CATEGORIES}
        in_by_cat:  dict[str, float] = {c: 0.0 for c in CATEGORIES}

        for art in set(snap_prev) | set(snap_curr):
            meta     = products.get(art, {})
            cat      = categorize(art, meta.get("product_name", ""))
            prev_wh  = snap_prev.get(art, {}).get("warehouses", {})
            curr_wh  = snap_curr.get(art, {}).get("warehouses", {})

            for wid in set(prev_wh) | set(curr_wh):
                delta = curr_wh.get(wid, 0.0) - prev_wh.get(wid, 0.0)
                if delta < 0:
                    out_by_cat[cat] += abs(delta)   # lagerminskning = försäljning till kund
                elif delta > 0:
                    in_by_cat[cat]  += delta        # lagerökning = inköp från Plejd

        sales_out[d_curr] = out_by_cat
        sales_in[d_curr]  = in_by_cat

    return sales_out, sales_in


# ── Excel ──────────────────────────────────────────────────────────────────────

_EVEN_FILL = PatternFill("solid", fgColor="F2F7FD")
_CENTER    = Alignment(horizontal="center", vertical="center")


def _write_delta_sheet(
    ws,
    delta_data: dict,
    header_color: str,
    subheader_color: str,
) -> None:
    """
    Skriver ett delta-sheet till ett befintligt worksheet-objekt.

    Struktur:
      Rad 1 : kolumnrubriker  (Datum | Dimmer | Armaturer | Termostat | LED-Panel | Övrigt)
      Rad 2+: ett datum per rad
    """
    hdr_fill = PatternFill("solid", fgColor=header_color)
    hdr_font = Font(color="FFFFFF", bold=True, size=11)

    # Rubrikrad
    headers = ["Datum"] + CATEGORIES
    for col, label in enumerate(headers, start=1):
        cell = ws.cell(1, col, label)
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = _CENTER

    ws.row_dimensions[1].height = 22

    # Datarader
    for row, d in enumerate(sorted(delta_data.keys()), start=2):
        ws.cell(row, 1, d).alignment = _CENTER
        for col, cat in enumerate(CATEGORIES, start=2):
            val = delta_data[d].get(cat, 0.0)
            ws.cell(row, col, round(val) if val else None)
        if row % 2 == 0:
            for col in range(1, 2 + len(CATEGORIES)):
                ws.cell(row, col).fill = _EVEN_FILL

    # Kolumnbredder
    ws.column_dimensions["A"].width = 13
    for col in range(2, 2 + len(CATEGORIES)):
        ws.column_dimensions[get_column_letter(col)].width = 15

    ws.freeze_panes = "B2"


def write_excel(state: dict) -> None:
    snapshots = state.get("snapshots", {})
    products  = state.get("products", {})

    if not snapshots:
        print("Ingen data att exportera.")
        return

    sales_out, sales_in = compute_deltas(snapshots, products)

    wb = Workbook()

    # ── Sheet 1: AHLSELL SALES OUT ─────────────────────────────────────────────
    ws_out = wb.active
    ws_out.title = "AHLSELL SALES OUT"
    _write_delta_sheet(ws_out, sales_out, header_color="C00000", subheader_color="FF0000")

    # ── Sheet 2: PLEJD SALES IN ────────────────────────────────────────────────
    ws_in = wb.create_sheet("PLEJD SALES IN")
    _write_delta_sheet(ws_in, sales_in, header_color="375623", subheader_color="70AD47")

    # ── Sheet 3: Kategorier ────────────────────────────────────────────────────
    ws_cat = wb.create_sheet("Kategorier")

    cat_hdr_fill = PatternFill("solid", fgColor="4A4A4A")
    cat_hdr_font = Font(color="FFFFFF", bold=True, size=11)

    for col, label in enumerate(["Artikel", "Produktnamn", "Kategori"], start=1):
        cell = ws_cat.cell(1, col, label)
        cell.fill      = cat_hdr_fill
        cell.font      = cat_hdr_font
        cell.alignment = _CENTER

    ws_cat.row_dimensions[1].height = 22

    # Sortera artiklar per kategori, sedan produktnamn
    sorted_articles = sorted(
        products.items(),
        key=lambda kv: (
            categorize(kv[0], kv[1].get("product_name", "")),
            kv[1].get("product_name", ""),
        ),
    )

    cat_colors = {
        "Dimmer":    "FFF2CC",
        "Armaturer": "DDEBF7",
        "Termostat": "E2EFDA",
        "LED-Panel": "FCE4D6",
        "Övrigt":    "F2F2F2",
    }

    for row, (art, meta) in enumerate(sorted_articles, start=2):
        cat = categorize(art, meta.get("product_name", ""))
        fill = PatternFill("solid", fgColor=cat_colors.get(cat, "FFFFFF"))
        for col, val in enumerate([art, meta.get("product_name", ""), cat], start=1):
            cell = ws_cat.cell(row, col, val)
            cell.fill = fill

    ws_cat.column_dimensions["A"].width = 14
    ws_cat.column_dimensions["B"].width = 48
    ws_cat.column_dimensions["C"].width = 15
    ws_cat.freeze_panes = "A2"

    wb.save(EXCEL_FILE)
    print(f"Excel sparad: {EXCEL_FILE}")


# ── Main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    today = date.today().isoformat()
    print(f"=== Ahlsell Plejd lageruppföljning — {today} ===\n")

    state = load_state()

    if today in state.get("snapshots", {}):
        print(f"Snapshot för {today} finns redan. Exporterar Excel...")
        write_excel(state)
        return

    products, warehouses, stock = collect_snapshot()
    update_state(state, products, warehouses, stock)
    save_state(state)
    print(f"\nTillstånd sparat: {STATE_FILE}")

    write_excel(state)
    print("\nKlart!")


if __name__ == "__main__":
    main()
