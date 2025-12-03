import asyncio
import random
import re
import os
from datetime import date
from pathlib import Path
from playwright.async_api import async_playwright

try:
    import openpyxl
except ImportError:
    print("Du måste installera openpyxl: pip install openpyxl")
    exit()

# ── KONFIGURATION ──────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).resolve().parent
DATA_DIR = (SCRIPT_DIR / ".." / "data").resolve()
DATA_DIR.mkdir(parents=True, exist_ok=True)

# Sparar till en ny fil för att bara ha ranking-data
XLSX_PATH = str((DATA_DIR / "fractal_scape_refine_ranks.xlsx").resolve())
SHEET_NAME = "Rankings"

HEADLESS = True

# --- PRODUKTER ---
PRODUCTS = [
    # --- SCAPE ---
    ("Scape Dark",        "B0D5HGK3C2"),
    ("Scape Light",       "B0D5HK6JRS"),
    # --- REFINE ---
    ("Refine Mesh Light",   "B0CSYXY8FD"),
    ("Refine Fabric Light", "B0CSYXYX39"),
    ("Refine Mesh Dark",    "B0CSYYMTT4"),
    ("Refine Fabric Dark",  "B0CSYWWRSV"),
]

# --- LÄNDER ---
COUNTRIES = [
    # code, domain, Accept-Language, gl, locale, cookie_name, cookie_value
    ("US", "amazon.com", "en-US,en;q=0.9", "US", "en-US", "i18n-prefs", "USD"),
    ("DE", "amazon.de",  "de-DE,de;q=0.9,en;q=0.6", "DE", "de-DE", "lc-acbde", "de_DE"), 
]

UAS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_0) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4 Safari/605.1.15",
]

# ───────────────────────────────────────────────────────────────────────────

def extract_rank_from_row_text(row_text: str) -> int:
    """
    Hämtar ranking, ignorerar 'Top 100'.
    Letar efter siffror med # eller Nr. ELLER siffror följt av 'in'.
    """
    # 1. Städa bort "Top 100" länkar så vi inte råkar ta siffran 100
    clean_text = re.sub(r"top\s*100", "", row_text, flags=re.IGNORECASE)
    
    # 2. Regex som fångar:
    # Fall A: Prefix (# eller Nr.) följt av siffra. Ex: "#45" eller "Nr. 45"
    # Fall B: Siffra följt av "in". Ex: "141 in"
    regex = r"(?:(?:Nr\.?|#)\s*([0-9.,]+))|([0-9.,]+)\s+in\s+"
    
    matches = re.findall(regex, clean_text, re.IGNORECASE)
    
    candidates = []
    for m in matches:
        # m är en tuple ('', '141') eller ('45', '')
        raw_num = m[0] if m[0] else m[1]
        
        clean_str = raw_num.replace(",", "").replace(".", "")
        if clean_str.isdigit():
            val = int(clean_str)
            # Rimlighetsfilter (Rank > 0 och < 10 miljoner)
            if 0 < val < 10000000:
                candidates.append(val)
                
    if candidates:
        return min(candidates) # Välj den lägsta (bästa) rankingen
    return 0

async def handle_blockers(page):
    try:
        # Försök klicka bort cookies
        if await page.locator("#sp-cc-accept").is_visible(timeout=2000):
            await page.click("#sp-cc-accept")
            await page.wait_for_timeout(500)
        
        # Försök klicka bort "Fortsätt handla" popups
        upsells = ['text="Continue shopping"', 'text="Weiter shoppen"', 'text="Weiter einkaufen"']
        for sel in upsells:
            if await page.locator(sel).is_visible(timeout=500):
                await page.click(sel)
                await page.wait_for_timeout(500)
    except:
        pass

async def get_product_rank(context, asin: str, domain: str, gl: str, name: str):
    page = await context.new_page()
    url = f"https://www.{domain}/dp/{asin}?th=1&psc=1&gl={gl}"
    print(f"  Fetching Rank for {name} ({domain})...")

    rank_val = 0

    try:
        await page.goto(url, wait_until="domcontentloaded", timeout=30000)
        await handle_blockers(page)

        # --- RANKING LOGIK ---
        try:
            # Leta efter element som innehåller "Best Seller" eller "Bestseller"
            # Vi letar brett efter texten, sen hittar vi föräldern (raden)
            rank_header = page.locator("text=/Best Sellers Rank|Bestseller-Rang|Best Seller Rank/i").first
            
            if await rank_header.count() > 0:
                # Hitta raden/sektionen (tr, li, eller div)
                row_element = rank_header.locator("xpath=ancestor::tr | ancestor::li | ancestor::div[contains(@class, 'db_row')]").first
                
                if await row_element.count() > 0:
                    row_text = await row_element.inner_text()
                    rank_val = extract_rank_from_row_text(row_text)
                else:
                    # Fallback: ta förälderns text om vi inte hittar en tydlig rad
                    row_text = await rank_header.evaluate("el => el.parentElement.innerText")
                    rank_val = extract_rank_from_row_text(row_text)

            if rank_val > 0:
                print(f"    🏆 Rank: #{rank_val}")
            else:
                # Sista utväg: Sök i hela kroppen om vi missade tabellen
                body_text = await page.inner_text("body")
                rank_val = extract_rank_from_row_text(body_text)
                if rank_val > 0:
                    print(f"    🏆 Rank (Fallback): #{rank_val}")
                else:
                    print(f"    ⚠️ Rank: Hittades ej")

        except Exception as e:
            print(f"    ⚠️ Rank Error: {e}")

        return rank_val

    except Exception as e:
        print(f"    ❌ Error: {e}")
        return 0
    finally:
        await page.close()

def append_to_excel(data_dict):
    file_exists = os.path.exists(XLSX_PATH)
    
    headers = ["Date"]
    keys_order = []
    
    # Skapa headers ENBART för Ranking
    for country_code, *_ in COUNTRIES:
        for prod_name, _ in PRODUCTS:
            key_rank = f"{prod_name} {country_code} Rank"
            headers.append(key_rank)
            keys_order.append(key_rank)

    if not file_exists:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(headers)
    else:
        wb = openpyxl.load_workbook(XLSX_PATH)
        if SHEET_NAME in wb.sheetnames:
            ws = wb[SHEET_NAME]
        else:
            ws = wb.create_sheet(SHEET_NAME)
            if ws.max_row == 1: pass 

    row = [data_dict["Date"]]
    for key in keys_order:
        row.append(data_dict.get(key, 0))
    
    ws.append(row)
    wb.save(XLSX_PATH)
    print(f"💾 Data saved to {XLSX_PATH}")

async def run_once():
    today = date.today().isoformat()
    results = {"Date": today}

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=HEADLESS)
        print(f"--- Starting Rank-Only Scraper ({today}) ---")

        for code, domain, accept_lang, gl, locale, ck_name, ck_value in COUNTRIES:
            print(f"\n--- Market: {code} ({domain}) ---")
            
            context = await browser.new_context(
                locale=locale,
                user_agent=random.choice(UAS),
                java_script_enabled=True,
                viewport={"width": 1366, "height": 850},
                extra_http_headers={"Accept-Language": accept_lang},
            )
            await context.add_cookies([{
                "name": ck_name, "value": ck_value,
                "domain": f".{domain}", "path": "/",
                "secure": True, "httpOnly": False
            }])

            for prod_name, asin in PRODUCTS:
                full_name = f"{prod_name} {code}"
                
                # Hämtar enbart rank
                rank = await get_product_rank(context, asin, domain, gl, full_name)
                
                results[f"{full_name} Rank"] = rank
                
                await asyncio.sleep(random.randint(1, 3))

            await context.close()
        await browser.close()
    append_to_excel(results)

if __name__ == "__main__":
    asyncio.run(run_once())