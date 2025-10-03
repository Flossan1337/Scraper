\
"""
excel_utils.py
Återanvändbara Excel-hjälpfunktioner.

Regler:
- Skriv alltid resultat till en .xlsx-fil.
- Om filen inte finns → skapa den.
- Om filen finns men fliken inte finns → skapa fliken.
- Om fliken finns → lägg till ny rad (append).
"""

from __future__ import annotations
import time
from pathlib import Path
from typing import Dict

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

DEFAULT_RETRY = 3
RETRY_SLEEP_S = 0.5


def _ensure_workbook(path: Path) -> None:
    """Skapar en tom arbetsbok om filen saknas (med en standardflik)."""
    if not path.exists():
        path.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()  # behåll default-fliken så filen alltid är giltig
        wb.save(path)


def _append_df_to_sheet(xlsx_path: Path, sheet_name: str, df: pd.DataFrame) -> None:
    """Appendar en DataFrame till en flik. Skapar fil/flik om de saknas."""
    _ensure_workbook(xlsx_path)
    try:
        wb = load_workbook(xlsx_path)
    except (InvalidFileException, KeyError):
        # Om filen korrupt: återskapa
        wb = Workbook()

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        is_empty = (ws.max_row == 0)
    else:
        ws = wb.create_sheet(title=sheet_name)
        is_empty = True

    # Skriv header vid behov
    if is_empty and len(df.columns) > 0:
        ws.append(list(df.columns))

    # Skriv rader
    for _, row in df.iterrows():
        ws.append([row.get(col, None) for col in df.columns])

    wb.save(xlsx_path)


def append_row(xlsx_path: str, sheet_name: str, row_dict: Dict, retries: int = DEFAULT_RETRY) -> None:
    """Lägg till en rad (dict) i en Excel-flik."""
    df = pd.DataFrame([row_dict])
    _attempt(lambda: _append_df_to_sheet(Path(xlsx_path), sheet_name, df), retries=retries)


def append_df(xlsx_path: str, sheet_name: str, df: pd.DataFrame, retries: int = DEFAULT_RETRY) -> None:
    """Lägg till en DataFrame i en Excel-flik."""
    _attempt(lambda: _append_df_to_sheet(Path(xlsx_path), sheet_name, df), retries=retries)


def _attempt(fn, retries: int = DEFAULT_RETRY):
    last_err = None
    for _ in range(max(1, retries)):
        try:
            return fn()
        except PermissionError as e:
            last_err = e
            time.sleep(RETRY_SLEEP_S)
        except Exception as e:
            last_err = e
            break
    if last_err:
        raise last_err
