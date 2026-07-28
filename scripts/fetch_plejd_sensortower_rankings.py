# fetch_plejd_sensortower_rankings.py
#
# Fetches Plejd's daily App Store category ranking (Lifestyle, Top Free iPhone)
# from the public Sensor Tower overview page — no login required.
# Appends one row per run to data/plejd_sensortower_rankings.xlsx.

import re
import time
import random
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeoutError

from excel_utils import append_row
from core.db import safe_insert

# ── CONFIG ─────────────────────────────────────────────────────────────────────
APP_ID     = "1032689423"
CATEGORY   = "6012"   # Lifestyle
CHART_TYPE = "free"   # Top Free
DEVICE     = "iphone"

COUNTRIES = ["SE", "NO", "FI", "NL", "DE", "DK", "ES"]



REPO_ROOT  = Path(__file__).resolve().parent.parent
DATA_DIR   = REPO_ROOT / "data"
DATA_DIR.mkdir(exist_ok=True)
XLSX_PATH  = str(DATA_DIR / "plejd_sensortower_rankings.xlsx")
SHEET_NAME = "category_rankings"

# Overview page — contains a KPI card with the current category ranking.
# The card element has a stable aria-labelledby attribute we can target directly.
OVERVIEW_URL = (
    "https://app.sensortower.com/overview/{app_id}"
    "?os=ios&country={country}&category={category}"
    "&device={device}&chart_type={chart_type}"
    "&start_date={date}&end_date={date}"
    "&granularity=daily"
)

# Rotate user agents to avoid a fixed fingerprint
_USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
]

# Browser args that reduce headless detection signals
_BROWSER_ARGS = [
    "--disable-blink-features=AutomationControlled",
    "--no-sandbox",
    "--disable-dev-shm-usage",
]

# Selector for the ranking KPI card — stable aria attribute visible in the DOM
# (renders as e.g. "#270", "#13")
_KPI_SELECTOR = '[aria-labelledby="app-overview-unified-kpi-category-ranking"]'

# ── HELPERS ────────────────────────────────────────────────────────────────────

def get_row_for_date(target_date: str) -> dict | None:
    """Returns the row dict for target_date if already in the Excel sheet, else None."""
    path = Path(XLSX_PATH)
    if not path.exists():
        return None
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        if SHEET_NAME not in wb.sheetnames:
            return None
        ws = wb[SHEET_NAME]
        rows_iter = ws.iter_rows(values_only=True)
        headers = next(rows_iter)
        for row in rows_iter:
            if row and row[0] is not None and str(row[0])[:10] == target_date:
                return dict(zip(headers, row))
    except Exception:
        pass
    return None


def write_rankings_to_db(row: dict) -> tuple[int | None, str | None]:
    """Best-effort: write one row per country to Postgres. Must never raise."""
    snapshot_date = str(row["Date"])
    values = [
        (snapshot_date, country, int(row[country]))
        for country in COUNTRIES
        if row.get(country) is not None
    ]
    return safe_insert(
        table="plejd_sensortower_rankings",
        columns=["snapshot_date", "country", "rank"],
        rows=values,
        conflict_columns=["snapshot_date", "country"],
    )


def fetch_rank(page, country: str) -> int | None:
    """Navigate to the Sensor Tower overview page and extract the KPI ranking card."""
    today_str = str(date.today())
    url = OVERVIEW_URL.format(
        app_id=APP_ID,
        country=country,
        category=CATEGORY,
        device=DEVICE,
        chart_type=CHART_TYPE,
        date=today_str,
    )
    print(f"  [{country}] Loading {url}")
    try:
        page.goto(url, timeout=30_000)
    except PWTimeoutError:
        print(f"  [{country}] Navigation timed out.")
        return None

    if "/users/sign_in" in page.url:
        print(f"  [{country}] Redirected to sign-in.")
        return None

    # Wait for the KPI ranking card to render
    try:
        page.wait_for_selector(_KPI_SELECTOR, timeout=20_000)
    except PWTimeoutError:
        print(f"  [{country}] KPI ranking card not found — app likely unranked here.")
        return None

    raw = page.locator(_KPI_SELECTOR).first.inner_text()
    # Text is e.g. "#270\nLifestyle - Downloads" — extract the leading integer
    match = re.search(r"#(\d+)", raw)
    if match:
        rank = int(match.group(1))
        print(f"  [{country}] Rank: {rank}")
        return rank

    print(f"  [{country}] Could not parse rank from: {raw!r}")
    return None


# ── MAIN ───────────────────────────────────────────────────────────────────────

def main():
    today_str = str(date.today())
    existing_row = get_row_for_date(today_str)

    if existing_row is not None:
        print(f"Today ({today_str}) is already written to {XLSX_PATH}. Skipping fetch, writing to DB only.")
        row = existing_row
    else:
        # Random startup delay (0–45 s) so the run time varies each day and
        # doesn’t leave a perfectly fixed pattern in Sensor Tower’s logs.
        startup_delay = random.uniform(0, 45)
        print(f"Startup delay: {startup_delay:.1f}s")
        time.sleep(startup_delay)

        ranks: dict[str, int | None] = {}

        with sync_playwright() as pw:
            browser = pw.chromium.launch(
                headless=True,
                args=_BROWSER_ARGS,
            )
            context = browser.new_context(
                user_agent=random.choice(_USER_AGENTS),
                locale="en-US",
                viewport={"width": random.choice([1280, 1366, 1440, 1920]), "height": random.choice([800, 900, 1080])},
                java_script_enabled=True,
            )
            # Mask the navigator.webdriver property
            context.add_init_script(
                "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
            )
            page = context.new_page()

            for country in COUNTRIES:
                ranks[country] = fetch_rank(page, country)
                # Polite delay between requests
                time.sleep(random.uniform(2.0, 4.0))

            browser.close()

        row = {"Date": today_str} | {c: ranks[c] for c in COUNTRIES}
        append_row(XLSX_PATH, SHEET_NAME, row)
        print(f"\nDone. Row written to {XLSX_PATH}:")
        for k, v in row.items():
            print(f"  {k}: {v}")

    db_rows_written, db_error = write_rankings_to_db(row)
    if db_error is not None:
        print(f"Databas: MISSLYCKADES – {db_error}")
    else:
        print(f"Databas: {db_rows_written} rader skrivna")


if __name__ == "__main__":
    main()
