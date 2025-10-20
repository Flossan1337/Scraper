from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
from statistics import median
from datetime import date
import os, re, sys, requests, json, time, argparse

# NEW: Excel writing
from openpyxl import Workbook, load_workbook

# -------- Config --------
BASE_ROOT = "https://secure.adtraction.com"
BASE = f"{BASE_ROOT}/partner"
STATE_PATH = "adtraction_state.json"

# NEW: write to data/ as .xlsx
DATA_DIR  = "data"
XLSX_PATH = os.path.join(DATA_DIR, "nonfinance_median_epc_SEK_wide.xlsx")

CACHE_PATH = "adtraction_category_cache.json"
CACHE_TTL_DAYS = 30

COUNTRY_URLS = {
    "Sweden":       f"{BASE}/programs.htm?cid=1&asonly=false",
    "Denmark":      f"{BASE}/programs.htm?cid=12&asonly=false",
    "Finland":      f"{BASE}/programs.htm?cid=14&asonly=false",
    "Norway":       f"{BASE}/programs.htm?cid=33&asonly=false",
    "France":       f"{BASE}/programs.htm?cid=15&asonly=false",
    "Germany":      f"{BASE}/programs.htm?cid=16&asonly=false",
    "Italy":        f"{BASE}/programs.htm?cid=22&asonly=false",
    "Spain":        f"{BASE}/programs.htm?cid=42&asonly=false",
    "Netherlands":  f"{BASE}/programs.htm?cid=32&asonly=false",
    "Poland":       f"{BASE}/programs.htm?cid=34&asonly=false",
    "Switzerland":  f"{BASE}/programs.htm?cid=44&asonly=false",
}

COUNTRY_CCY = {
    "Sweden": "SEK","Denmark": "DKK","Finland": "EUR","Norway": "NOK",
    "France": "EUR","Germany": "EUR","Italy": "EUR","Spain": "EUR",
    "Netherlands": "EUR","Poland": "PLN","Switzerland": "CHF",
}

COUNTRY_ORDER = [
    "Sweden","Denmark","Finland","Norway",
    "France","Germany","Italy","Spain",
    "Netherlands","Poland","Switzerland"
]

FINANCE_KEYWORDS = {"finans","finance","finanzen","finanza","finanze","financiën","finanse","finanzas","rahoitus"}
NUMBER_RE   = re.compile(r"(\d[\d\s\u00A0]*[.,]\d+|\d[\d\s\u00A0]*)")
CURR_CODE_RE = re.compile(r"\b(SEK|EUR|DKK|NOK|PLN|CHF)\b", re.IGNORECASE)
HAS_EUR_SYM  = re.compile(r"€")
HAS_PLN_SYM  = re.compile(r"zł", re.IGNORECASE)
HAS_CHF_SYM  = re.compile(r"\bCHF\b|(?<!\w)Fr(?!\w)")
HAS_KR_SYM   = re.compile(r"\bkr\.?\b", re.IGNORECASE)

# ---------- Cache helpers ----------
def load_cache():
    if not os.path.exists(CACHE_PATH): return {}
    try:
        with open(CACHE_PATH, "r", encoding="utf-8") as f: return json.load(f)
    except Exception: return {}

def save_cache(cache):
    with open(CACHE_PATH, "w", encoding="utf-8") as f: json.dump(cache, f, ensure_ascii=False, indent=2)

def is_entry_fresh(entry):
    try: ts = float(entry.get("ts", 0))
    except Exception: return False
    return (time.time() - ts)/86400.0 <= CACHE_TTL_DAYS
# -----------------------------------

def parse_number(text: str):
    if not text: return None
    t = " ".join(text.split()).strip().lower()
    if t in {"ingen data","no data","-","—",""}: return None
    m = NUMBER_RE.search(t)
    if not m: return None
    raw = m.group(1).replace("\u00A0"," ").replace(" ","")
    if "," in raw and "." in raw: raw = raw.replace(".","")
    try: val = float(raw.replace(",",".")) 
    except ValueError: return None
    if abs(val) < 1e-12: return None
    return val

def detect_currency(cell_text: str, country_name: str):
    if not cell_text: return None
    t = " ".join(cell_text.split()).strip()
    m = CURR_CODE_RE.search(t)
    if m: return m.group(1).upper()
    if HAS_EUR_SYM.search(t): return "EUR"
    if HAS_PLN_SYM.search(t): return "PLN"
    if HAS_CHF_SYM.search(t): return "CHF"
    if HAS_KR_SYM.search(t) or "SEK" in t.upper():
        if "SEK" in t.upper(): return "SEK"
        if country_name == "Sweden": return "SEK"
        if country_name == "Denmark": return "DKK"
        if country_name == "Norway":  return "NOK"
        return "SEK"
    return None

def scrape_epc_values_from_list(page, country_name: str):
    out = []
    tables = page.locator("table")
    for i in range(tables.count()):
        table = tables.nth(i)
        headers = table.locator("thead tr th")
        if not headers.count(): continue
        heads = [headers.nth(j).inner_text().strip().lower() for j in range(headers.count())]
        if not any("epc" in h for h in heads): continue
        epc_idx = next(idx for idx,h in enumerate(heads) if "epc" in h)
        try: table.locator("tbody tr").first.wait_for(state="visible", timeout=6000)
        except PWTimeout: pass
        rows = table.locator("tbody tr")
        for r in range(rows.count()):
            cells = rows.nth(r).locator("td")
            if cells.count() <= epc_idx: continue
            cell_text = cells.nth(epc_idx).inner_text()
            val = parse_number(cell_text)
            if val is None: continue
            ccy = detect_currency(cell_text, country_name)
            out.append((val, ccy))
    return out

def extract_categories_via_dom(page):
    cats = page.evaluate("""
(() => {
  const out = [], seen = new Set();
  const rx = /category\\(\\s*['"]?(-?\\d+)['"]?\\s*\\)/i;
  for (const el of document.querySelectorAll('*')) {
    let id=null, label=null;
    const oc = el.getAttribute && el.getAttribute('onclick');
    if (oc && rx.test(oc)) { id = oc.match(rx)[1];
      if (id !== '-1' && !seen.has(id)) { seen.add(id); label = (el.textContent||'').trim().replace(/\\s+/g,' ');
        out.push({id, label}); } continue; }
    const href = el.getAttribute && el.getAttribute('href');
    if (href && rx.test(href)) { id = href.match(rx)[1];
      if (id !== '-1' && !seen.has(id)) { seen.add(id); label = (el.textContent||'').trim().replace(/\\s+/g,' ');
        out.push({id, label}); } continue; }
    const ds = el.dataset || {};
    for (const key of ['category','categoryId','cid','cat','cId']) {
      if (ds[key]) { id = String(ds[key]);
        if (id !== '-1' && !seen.has(id)) { seen.add(id); label = (el.textContent||'').trim().replace(/\\s+/g,' ');
          out.push({id, label}); } }
    }
  } return out;
})()
""")
    return [(c.get("id"), c.get("label","")) for c in cats]

def guess_finance_id(categories):
    for cid, label in categories:
        if any(k in (label or "").lower() for k in FINANCE_KEYWORDS):
            return cid
    return None

def to_abs(href: str) -> str:
    if not href: return ""
    if href.startswith("http"): return href
    if href.startswith("/"):    return BASE_ROOT + href
    return BASE_ROOT + "/" + href.lstrip("./")

def discover_pagination_urls(page, cid: str):
    hrefs = page.evaluate("""() => Array.from(document.querySelectorAll('a[href]'),a=>a.getAttribute('href'))""")
    urls = set([page.url])
    if hrefs:
        for h in hrefs:
            if h and "listadvertprograms.htm" in h and f"cId={cid}" in h and "page=" in h:
                urls.add(to_abs(h))
    def page_key(u):
        m = re.search(r"[?&]page=(\d+)", u)
        return int(m.group(1)) if m else 1
    return sorted(urls, key=page_key)

# ---------- Login helpers (NEW) ----------
def looks_like_login(page):
    url = page.url.lower()
    if "login" in url or "signin" in url: return True
    if page.locator('input[type="password"]').count() > 0: return True
    if page.locator('text=/logga in|sign in|log in/i').count() > 0: return True
    return False

def auto_login_and_save_state(p, email: str, password: str, headless: bool = True):
    """Attempt email+password login; save storage state to STATE_PATH."""
    browser = p.chromium.launch(headless=headless)
    context = browser.new_context(
        user_agent=("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                    "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
    )
    page = context.new_page()
    page.goto(f"{BASE}/programs.htm?asonly=false", wait_until="domcontentloaded")

    if not looks_like_login(page):
        context.storage_state(path=STATE_PATH)
        context.close(); browser.close()
        return True

    def fill_if_exists(selector, value):
        loc = page.locator(selector)
        if loc.count():
            loc.first.fill(value)
            return True
        return False

    # Try obvious login route
    try:
        page.locator('a:has-text("Partner")').first.click(timeout=2000)
    except Exception:
        pass

    email_filled = (
        fill_if_exists('input[type="email"]', email) or
        fill_if_exists('input[name="email"]', email) or
        fill_if_exists('input#email', email) or
        fill_if_exists('input[id*="email" i]', email) or
        fill_if_exists('input[name*="user" i]', email)
    )
    pwd_filled = (
        fill_if_exists('input[type="password"]', password) or
        fill_if_exists('input[name="password"]', password) or
        fill_if_exists('input#password', password) or
        fill_if_exists('input[id*="pass" i]', password)
    )
    if not (email_filled and pwd_filled):
        page.goto(f"{BASE}/login.htm", wait_until="domcontentloaded")
        email_filled = (
            fill_if_exists('input[type="email"]', email) or
            fill_if_exists('input[name="email"]', email) or
            fill_if_exists('input#email', email) or
            fill_if_exists('input[id*="email" i]', email) or
            fill_if_exists('input[name*="user" i]', email)
        )
        pwd_filled = (
            fill_if_exists('input[type="password"]', password) or
            fill_if_exists('input[name="password"]', password) or
            fill_if_exists('input#password', password) or
            fill_if_exists('input[id*="pass" i]', password)
        )

    submitted = False
    for sel in ['button[type="submit"]','input[type="submit"]',
                'button:has-text("Logga in")','button:has-text("Sign in")',
                'text=Logga in','text=Sign in']:
        try:
            page.locator(sel).first.click(timeout=1500)
            submitted = True
            break
        except Exception:
            continue
    if not submitted:
        try: page.keyboard.press("Enter")
        except Exception: pass

    try:
        page.wait_for_load_state("networkidle", timeout=10000)
    except PWTimeout:
        pass

    page.goto(f"{BASE}/programs.htm?asonly=false", wait_until="domcontentloaded")
    ok = not looks_like_login(page)
    if ok:
        context.storage_state(path=STATE_PATH)
    context.close(); browser.close()
    return ok

def interactive_login_and_save_state(p):
    print("\nOpening a visible browser for one-time login…")
    browser = p.chromium.launch(headless=False, slow_mo=250)
    context = browser.new_context()
    page = context.new_page()
    page.goto(f"{BASE}/programs.htm?asonly=false", wait_until="domcontentloaded")
    input("Log in in the browser, then press ENTER here… ")
    context.storage_state(path=STATE_PATH)
    context.close(); browser.close()
# ----------------------------------------

def fetch_fx_local_to_sek(currencies):
    need = sorted({c for c in currencies if c and c != "SEK"})
    out = {c: 1.0 for c in currencies}
    if not need: return out
    try:
        r = requests.get("https://open.er-api.com/v6/latest/SEK", timeout=15)
        r.raise_for_status(); data = r.json()
        if data.get("result") == "success":
            rates = data.get("rates", {})
            for c in need:
                per_sek = float(rates[c]); out[c] = 1.0/per_sek if per_sek else 1.0
            print("\nFX (open.er-api.com):")
            for c in need: print(f"  {c}→SEK = {out[c]:.4f}")
            return out
    except Exception as e:
        print(f"FX provider 1 failed ({e}); trying fallback…")
    symbols = ",".join(need)
    r = requests.get(f"https://api.exchangerate.host/latest?base=SEK&symbols={symbols}", timeout=15)
    r.raise_for_status(); rates = r.json().get("rates", {})
    for c in need:
        per_sek = float(rates[c]); out[c] = 1.0/per_sek if per_sek else 1.0
    print("\nFX (exchangerate.host):")
    for c in need: print(f"  {c}→SEK = {out[c]:.4f}")
    return out

# ===== Excel helpers (replace CSV) =====
def ensure_xlsx_header(path):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        header = ["date", "All excl. Finance (SEK)"] + [f"{c} excl. Finance (SEK)" for c in COUNTRY_ORDER]
        ws.append(header)
        wb.save(path)

def append_row_xlsx(path, dt, all_med, per_country):
    ensure_xlsx_header(path)
    wb = load_workbook(path)
    ws = wb.active
    row = [dt, (None if all_med is None else round(all_med, 2))] + [
        (None if per_country.get(c) is None else round(per_country.get(c), 2)) for c in COUNTRY_ORDER
    ]
    ws.append(row)
    wb.save(path)
# ======================================

def main():
    parser = argparse.ArgumentParser(add_help=False)
    parser.add_argument("countries", nargs="*", help="Run for specific countries (partial names OK)")
    parser.add_argument("--refresh-cache", "-R", action="store_true", help="Re-scan category IDs even if cached")
    parser.add_argument("--headful", action="store_true", help="Show the browser window during run")
    args = parser.parse_args()

    if args.countries:
        sel = []
        for r in args.countries:
            rlow = r.lower()
            sel.extend([c for c in COUNTRY_ORDER if rlow in c.lower()])
        seen=set(); countries=[c for c in sel if not (c in seen or seen.add(c))]
        if not countries: countries = COUNTRY_ORDER
    else:
        countries = COUNTRY_ORDER

    fx = fetch_fx_local_to_sek({COUNTRY_CCY[c] for c in countries})
    cache = load_cache()

    with sync_playwright() as p:
        # ---- NEW: try existing state; else auto-login via env; else interactive once ----
        need_login = not os.path.exists(STATE_PATH)
        if not need_login:
            tmp_b = p.chromium.launch(headless=True)
            tmp_ctx = tmp_b.new_context(storage_state=STATE_PATH)
            tmp_page = tmp_ctx.new_page()
            tmp_page.goto(f"{BASE}/programs.htm?asonly=false", wait_until="domcontentloaded")
            need_login = looks_like_login(tmp_page)
            tmp_ctx.close(); tmp_b.close()

        if need_login:
            email = os.environ.get("ADTRACTION_EMAIL", "")
            password = os.environ.get("ADTRACTION_PASSWORD", "")
            if email and password:
                ok = auto_login_and_save_state(p, email, password, headless=True)
                if not ok:
                    print("Auto-login misslyckades – faller tillbaka till manuell inloggning.")
                    interactive_login_and_save_state(p)
            else:
                print("ADTRACTION_EMAIL/ADTRACTION_PASSWORD saknas – använder manuell engångsinloggning.")
                interactive_login_and_save_state(p)
        # -------------------------------------------------------------------------------

        browser = p.chromium.launch(headless=not args.headful, slow_mo=150 if args.headful else 0)
        context = browser.new_context(storage_state=STATE_PATH, viewport={"width":1440,"height":900} if args.headful else None)
        page = context.new_page()

        today = date.today().isoformat()
        all_vals_sek, per_country_med = [], {}

        for country in countries:
            ccy_expected = COUNTRY_CCY[country]
            try:
                page.goto(COUNTRY_URLS[country], wait_until="domcontentloaded")
                try: page.wait_for_load_state("networkidle", timeout=8000)
                except PWTimeout: pass
                print(f"[{country}] Country context set via Programs page.")

                entry = cache.get(country)
                if entry and not args.refresh_cache and is_entry_fresh(entry):
                    cats = [(c["id"], c.get("label","")) for c in entry.get("categories", [])]
                    finance_id = entry.get("finance_id") or guess_finance_id(cats)
                    print(f"[{country}] Using cached categories ({len(cats)}); finance cId={finance_id or 'unknown'}")
                else:
                    cats = extract_categories_via_dom(page)
                    finance_id = guess_finance_id(cats)
                    cache[country] = {
                        "ts": time.time(),
                        "categories": [{"id": cid, "label": label} for cid, label in cats],
                        "finance_id": finance_id,
                    }
                    save_cache(cache)
                    print(f"[{country}] Cached {len(cats)} categories; finance cId={finance_id or 'unknown'}")

                non_finance = [(cid,label) for cid,label in cats if cid != finance_id]

                page.goto(COUNTRY_URLS[country], wait_until="domcontentloaded")
                try: page.wait_for_load_state("networkidle", timeout=8000)
                except PWTimeout: pass

                values_local_currency = []
                for cid, label in non_finance:
                    list_url = f"{BASE}/listadvertprograms.htm?cId={cid}&asonly=false"
                    page.goto(list_url, wait_until="domcontentloaded")
                    try: page.wait_for_selector("table", timeout=8000)
                    except PWTimeout: continue

                    page_urls = discover_pagination_urls(page, cid)
                    for u in page_urls:
                        page.goto(u, wait_until="domcontentloaded")
                        try: page.wait_for_selector("table", timeout=8000)
                        except PWTimeout: continue
                        values_local_currency.extend(scrape_epc_values_from_list(page, country))

                if not values_local_currency:
                    print(f"[{country}] No non-finance EPC values found.")
                    per_country_med[country] = None
                    continue

                ccy_counts = {}
                vals_sek = []
                for val, ccy_detected in values_local_currency:
                    ccy = (ccy_detected or ccy_expected).upper()
                    ccy_counts[ccy] = ccy_counts.get(ccy, 0) + 1
                    rate = fx.get(ccy, 1.0)
                    vals_sek.append(val * rate)
                print(f"   [CCY mix {country}] {ccy_counts}")

                med_country = median(vals_sek)
                per_country_med[country] = med_country
                all_vals_sek.extend(vals_sek)
                print(f"[{country}] n={len(values_local_currency)} median_ex_fin={med_country:.2f} SEK")
            except Exception as e:
                print(f"[{country}] Error: {e}")
                per_country_med[country] = None

        all_med = median(all_vals_sek) if all_vals_sek else None
        if all_med is not None:
            print(f"\n[ALL COUNTRIES excl. Finance] n={len(all_vals_sek)} median={all_med:.2f} SEK")
        else:
            print("\n[ALL COUNTRIES excl. Finance] No values.")

        # NEW: write to Excel
        append_row_xlsx(XLSX_PATH, today, all_med, per_country_med)
        print(f"\nAppended row for {today} to {XLSX_PATH}")

        context.close(); browser.close()

if __name__ == "__main__":
    main()
