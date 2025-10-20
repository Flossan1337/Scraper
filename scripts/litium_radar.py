# scripts/litium_radar.py
# -*- coding: utf-8 -*-

"""
Litium Radar (manual run)
- Månadsvis (Jan 2024 -> current month) hämtar kandidater via Common Crawl (CC)
- Validerar kandidater med heuristiker (HTML/JS/headers/cookies)
- Räknar "High" för varje månad
- Skriver Excel: data/litium_radar.xlsx
  - Sheet "Summary": Month | HighCount
  - Sheet "NewDomains": en kolumn per månad med NYA domäner (första-gångs-fynd)
Kör: python scripts/litium_radar.py
"""

import os
import sys
import json
import math
import time
import asyncio
import calendar
import datetime as dt
from collections import OrderedDict, defaultdict
from urllib.parse import urlparse, urljoin

import requests
from urllib3.util.retry import Retry
from requests.adapters import HTTPAdapter

import aiohttp
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# ---------------------------------
# Konfig
# ---------------------------------

# Var vi skriver Excel
DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data")
XLSX_PATH = os.path.join(DATA_DIR, "litium_radar.xlsx")

# Tidsperiod
SINCE = dt.date(2024, 1, 1)  # från och med Jan 2024
UNTIL = dt.date.today().replace(day=1)  # till och med nuvarande månad (början på månaden)

# Begränsa vilka toppdomäner vi bryr oss om (Norden)
TLD_FILTER = [".se", ".no", ".dk", ".fi"]  # kan utökas/ändras

# Common Crawl söktermer (unika Litium-spår)
CC_TERMS = [
    "data-litium-block-id",
    "window._litium",
    "litium-request-context",
    "litium.constants",
]

# Hur många sidor per term vi begär från CC-index per månad (kontrollerar volym/hastighet)
CC_PAGE_SIZE = 1000
CC_MAX_PAGES_PER_TERM = 3  # 3*1000*4 terms ≈ 12k råträffar/månad (dedupas hårt)

# Validerings-inställningar
VALIDATION_CONCURRENCY = 20
VALIDATION_TIMEOUT = 12
MAX_JS_FETCH = 5  # hämta upp till N script-filer per sajt

# Litium-signaler (regex/enkla substrings)
DETECTION_SUBSTRINGS = [
    ("js_window._litium", "window._litium", 3),
    ("js_litium.constants", "litium.constants", 2),
    ("js_litium.cache", "litium.cache", 2),
    ("js_litium.bootstrapComponent", "litium.bootstrapComponent", 3),
    ("attr_data-litium-block-id", "data-litium-block-id", 2),
    ("header_or_js_litium-request-context", "litium-request-context", 2),
]

# Statiska CC-index (fallback OCH mapping mot månader)
# Vi väljer "närmast" index för varje månad (ordningen viktigast – nyast först)
CC_INDEX_POOL = [
    "CC-MAIN-2025-39", "CC-MAIN-2025-34", "CC-MAIN-2025-30",
    "CC-MAIN-2025-25", "CC-MAIN-2025-20", "CC-MAIN-2025-14",
    "CC-MAIN-2025-10", "CC-MAIN-2025-06", "CC-MAIN-2025-01",
    "CC-MAIN-2024-49", "CC-MAIN-2024-43", "CC-MAIN-2024-38",
    "CC-MAIN-2024-34", "CC-MAIN-2024-31", "CC-MAIN-2024-27",
    "CC-MAIN-2024-22", "CC-MAIN-2024-18", "CC-MAIN-2024-10",
    "CC-MAIN-2024-06", "CC-MAIN-2024-01",
]

# ---------------------------------
# Hjälpare
# ---------------------------------

def month_iter(start_date: dt.date, end_date: dt.date):
    """Yield (year, month) från start_date (inkl) till end_date (inkl om samma månad)."""
    y, m = start_date.year, start_date.month
    last = (end_date.year, end_date.month)
    while (y, m) <= last:
        yield y, m
        if m == 12:
            y += 1
            m = 1
        else:
            m += 1

def ensure_data_dir():
    os.makedirs(DATA_DIR, exist_ok=True)

def _make_requests_session():
    sess = requests.Session()
    retry = Retry(
        total=5, connect=5, read=5,
        backoff_factor=1.0,
        status_forcelist=[429, 500, 502, 503, 504, 520, 521, 522, 524],
        allowed_methods=["GET"],
        raise_on_status=False,
    )
    adapter = HTTPAdapter(max_retries=retry, pool_connections=10, pool_maxsize=10)
    sess.mount("http://", adapter)
    sess.mount("https://", adapter)
    sess.headers.update({
        "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                       "AppleWebKit/537.36 (KHTML, like Gecko) "
                       "Chrome/120.0.0.0 Safari/537.36")
    })
    return sess

def fetch_collinfo():
    """Försök hämta dynamisk lista över CC-index. Faller tillbaka till CC_INDEX_POOL."""
    url = "https://index.commoncrawl.org/collinfo.json"
    sess = _make_requests_session()
    try:
        r = sess.get(url, timeout=30)
        r.raise_for_status()
        info = r.json()
        # sortera nyaste först, och ta bara CC-MAIN
        ids = [c["id"] for c in info if c.get("id", "").startswith("CC-MAIN-")]
        ids = sorted(ids, reverse=True)
        return ids
    except Exception as e:
        print(f"[warn] Could not fetch collinfo dynamically, using static pool. ({e})")
        return CC_INDEX_POOL

def pick_index_for_month(all_ids, year, month):
    """
    Välj ett "rimligt" index för en given månad.
    Vi väljer helt enkelt den index-ID vars år ligger närmast och som finns i vår lista.
    (Detta räcker bra i praktiken för månadsupplösning.)
    """
    # Approximation: välj första index som börjar med samma år eller närmast efter.
    ym_tag = f"{year}"
    for id_ in all_ids:
        if ym_tag in id_:
            return id_
    # annars ta närmast nyare
    return all_ids[0] if all_ids else None

def cdx_query(index_id, term, page, page_size):
    """
    Returnerar rader (str) från CC CDX för given index/term/sida.
    """
    # Ex: https://index.commoncrawl.org/CC-MAIN-2024-43-index?url=*data-litium-block-id*&output=json&page=0&limit=1000&filter=status:200
    base = f"https://index.commoncrawl.org/{index_id}-index"
    params = {
        "url": f"*{term}*",
        "output": "json",
        "page": page,
        "limit": page_size,
        "filter": "status:200"
    }
    sess = _make_requests_session()
    r = sess.get(base, params=params, timeout=40)
    if r.status_code != 200:
        return []
    return r.text.splitlines()

def extract_host_from_cdx_line(line: str):
    try:
        obj = json.loads(line)
        url = obj.get("url")
        if not url:
            return None
        host = urlparse(url).netloc.lower()
        if ":" in host:
            host = host.split(":")[0]
        return host
    except Exception:
        return None

def passes_tld(host: str):
    if not TLD_FILTER:
        return True
    return any(host.endswith(tld) for tld in TLD_FILTER)

# ---------------------------------
# Validering (asynkron)
# ---------------------------------

def make_litium_score(html_text: str, headers: dict, cookies: dict, resource_hits: list):
    text = html_text or ""
    score = 0
    evidence = []

    # direkta substrings
    low_text = text.lower()
    for tag, needle, pts in DETECTION_SUBSTRINGS:
        if needle.lower() in low_text:
            score += pts
            evidence.append(tag)

    # headers
    if headers:
        if any("litium" in str(v).lower() for v in headers.values()):
            score += 2
            evidence.append("response_header_contains_litium")

    # cookies
    if cookies:
        ck = {k: (getattr(v, "value", str(v)) if v else "") for k, v in cookies.items()}
        if any("litium" in (k.lower() + str(v).lower()) for k, v in ck.items()):
            score += 2
            evidence.append("cookie_contains_litium")

    # resursnamn
    if resource_hits:
        score += 2
        evidence.append("resource_name_contains_litium")

    # dom-attribut
    if "data-litium-block-id" in low_text:
        score += 2
        if "attr_data-litium-block-id" not in evidence:
            evidence.append("dom_selector_[data-litium-block-id]")

    confidence = "high" if score >= 7 else ("medium" if score >= 4 else "low")
    return score, confidence, evidence

async def fetch_text(session: aiohttp.ClientSession, url: str, timeout: int):
    try:
        async with session.get(url, allow_redirects=True, timeout=timeout) as resp:
            txt = await resp.text(errors="ignore")
            return resp.status, dict(resp.headers), txt, resp.cookies
    except Exception:
        return None, {}, "", {}

async def validate_host(session: aiohttp.ClientSession, host: str):
    url = "https://" + host
    status, headers, html, cookies = await fetch_text(session, url, VALIDATION_TIMEOUT)
    if not html:
        return {"host": host, "status": status, "score": 0, "confidence": "low", "evidence": []}

    soup = BeautifulSoup(html, "html.parser")

    # samla lokala script-länkar (samma origin)
    js_urls = []
    origin = urlparse(url).netloc
    for s in soup.find_all("script", src=True):
        src = s.get("src") or ""
        absu = urljoin(url, src)
        if urlparse(absu).netloc == origin:
            js_urls.append(absu)
    js_urls = js_urls[:MAX_JS_FETCH]

    resource_hits = [u for u in js_urls if "litium" in u.lower()]
    js_texts = []
    for ju in js_urls:
        st, _, jtxt, _ = await fetch_text(session, ju, VALIDATION_TIMEOUT)
        if st and jtxt:
            js_texts.append(jtxt)

    combined = html + "\n" + "\n".join(js_texts)
    score, conf, evidence = make_litium_score(combined, headers, cookies, resource_hits)
    return {"host": host, "status": status, "score": score, "confidence": conf, "evidence": evidence}

async def validate_many(hosts: list):
    connector = aiohttp.TCPConnector(limit=VALIDATION_CONCURRENCY, ssl=False)
    headers = {
        "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                       "AppleWebKit/537.36 (KHTML, like Gecko) "
                       "Chrome/120.0.0.0 Safari/537.36")
    }
    async with aiohttp.ClientSession(connector=connector, headers=headers) as session:
        tasks = [validate_host(session, h) for h in hosts]
        results = []
        for chunk in [tasks[i:i+VALIDATION_CONCURRENCY] for i in range(0, len(tasks), VALIDATION_CONCURRENCY)]:
            results.extend(await asyncio.gather(*chunk))
        return results

# ---------------------------------
# Huvudlogik
# ---------------------------------

def collect_candidates_for_month(index_id: str, year: int, month: int):
    """Från CC-index + termer -> kandidathosts för denna månad (dedupad, TLD-filtrerad)."""
    uniq = OrderedDict()
    for term in CC_TERMS:
        pages = 0
        while pages < CC_MAX_PAGES_PER_TERM:
            lines = cdx_query(index_id, term, pages, CC_PAGE_SIZE)
            if not lines:
                break
            got = 0
            for L in lines:
                host = extract_host_from_cdx_line(L)
                if not host:
                    continue
                if not passes_tld(host):
                    continue
                uniq[host] = None
                got += 1
            pages += 1
            # liten paus för att vara snäll
            time.sleep(0.6)
            if got == 0:
                break
    return list(uniq.keys())

def write_excel(summary_rows, new_domains_by_month):
    """
    summary_rows: list of (YYYY-MM, high_count)
    new_domains_by_month: dict {'YYYY-MM': [domains...]}
    """
    ensure_data_dir()
    wb = Workbook()

    # Summary sheet
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Month", "HighCount"])
    for ym, cnt in summary_rows:
        ws.append([ym, cnt])
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12

    # NewDomains sheet: en kolumn per månad
    ws2 = wb.create_sheet("NewDomains")
    months = [ym for ym, _ in summary_rows]
    for col_idx, ym in enumerate(months, start=1):
        ws2.cell(row=1, column=col_idx, value=ym)
        doms = new_domains_by_month.get(ym, [])
        for r, host in enumerate(doms, start=2):
            ws2.cell(row=r, column=col_idx, value=host)
        # auto-bredd typ
        ws2.column_dimensions[get_column_letter(col_idx)].width = 28

    wb.save(XLSX_PATH)
    print(f"[ok] Wrote Excel → {XLSX_PATH}")

def main():
    print(f"[i] Period: {SINCE.strftime('%Y-%m')} → {UNTIL.strftime('%Y-%m')} | TLD-filter: {TLD_FILTER}")

    # 1) Hämta CC-indexlista (dynamiskt, fallback statiskt)
    cc_ids = fetch_collinfo()
    if not cc_ids:
        cc_ids = CC_INDEX_POOL

    # 2) Per månad: plocka index-id och samla kandidater
    month_hosts = {}  # 'YYYY-MM' -> set(hosts)
    for y, m in month_iter(SINCE, UNTIL):
        ym = f"{y:04d}-{m:02d}"
        idx = pick_index_for_month(cc_ids, y, m)
        if not idx:
            print(f"[warn] No CC index for {ym}, skipping")
            month_hosts[ym] = set()
            continue

        print(f"[.] {ym}: querying Common Crawl index {idx} …")
        hosts = collect_candidates_for_month(idx, y, m)
        month_hosts[ym] = set(hosts)
        print(f"    → candidates: {len(hosts)}")

    # 3) Validera varje månads kandidater (High/Medium/Low) – vi validerar unionen och mappar tillbaka
    all_candidates = sorted(set().union(*month_hosts.values()))
    print(f"[.] Validating total unique hosts: {len(all_candidates)} … (this can take a bit)")
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    results = loop.run_until_complete(validate_many(all_candidates))
    loop.close()

    # Gör en lookup: host -> confidence
    conf_by_host = {r["host"]: r["confidence"] for r in results}

    # 4) Räkna High per månad + hitta NYA domäner (första gången de blir High)
    seen_high = set()
    summary_rows = []
    new_domains_by_month = defaultdict(list)

    for ym in sorted(month_hosts.keys()):
        hosts = month_hosts[ym]
        highs = sorted(h for h in hosts if conf_by_host.get(h) == "high")
        # nya = sådana som inte varit high tidigare
        new_highs = [h for h in highs if h not in seen_high]
        for h in new_highs:
            seen_high.add(h)
        summary_rows.append((ym, len(highs)))
        new_domains_by_month[ym] = new_highs
        print(f"[{ym}] High={len(highs)} | New={len(new_highs)}")

    # 5) Skriv Excel
    write_excel(summary_rows, new_domains_by_month)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nInterrupted.")
        sys.exit(1)
